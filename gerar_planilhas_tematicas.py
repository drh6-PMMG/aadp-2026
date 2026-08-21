# -*- coding: utf-8 -*-
"""
Gera 3 planilhas temáticas AADP 2026:
  1. Militares com Nota + Punição
  2. Militares com Nota + Prisão
  3. Militares com Nota + Movimentação

Cada planilha possui:
  - Aba "Geral" com todos os registros do tema
  - Uma aba por Unidade RPM do avaliado
"""

import csv, sys, io, os, re, unicodedata, zipfile
from datetime import datetime, date
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# ─── CONFIGURAÇÕES ────────────────────────────────────────────────────────────
BASE_DIR   = r"C:\Users\guilh\Downloads\analise AADP 2026"
GERAL_FILE = os.path.join(BASE_DIR, "geral.csv")
SIGEF_FILE = os.path.join(BASE_DIR, "SIGEF.csv")
MOV_FILE   = os.path.join(BASE_DIR, "MOVIMENTAÇÕES.xlsx")
PRESO_FILE = os.path.join(BASE_DIR, "PRESO.xlsx")
PUN_FILE   = os.path.join(BASE_DIR, "PUNIÇÃO.xlsx")
OUT_DIR    = os.path.join(BASE_DIR, "Resultado_AADP_2026")
TODAY      = date.today()

os.makedirs(OUT_DIR, exist_ok=True)

SITUACOES_ALVO = {
    "ATIV. DIRECAO GERAL", "ATIV. FIM DESTACADO", "ATIV. FIM NA SEDE",
    "ATIV. MEIO", "ATIVIDADE MEIO", "DISP MED DEFINITIVA",
    "GESTANTE/LAC/ADOTANT", "QUADRO ESPECIALISTA",
}

CONCEITO_FAIXA = {
    "nivel superior de desempenho":       (9.00, 10.00),
    "nivel alto de desempenho":           (7.00,  8.99),
    "nivel intermediario de desempenho":  (6.00,  6.99),
    "nivel baixo de desempenho":          (3.00,  5.99),
    "nivel inferior de desempenho":       (0.00,  2.99),
}

# ─── FUNÇÕES UTILITÁRIAS ──────────────────────────────────────────────────────
def normaliza(texto):
    t = unicodedata.normalize("NFD", str(texto).lower())
    return "".join(c for c in t if unicodedata.category(c) != "Mn")

def is_empty(v):
    if v == 0 or v == 0.0:
        return False
    return not v or str(v).strip() in ("", "-", "nan", "none", "NaT")

def concordam(conceito, nota_str):
    if is_empty(conceito) or is_empty(nota_str):
        return None
    try:
        nota = float(str(nota_str).replace(",", "."))
    except:
        return None
    faixa = CONCEITO_FAIXA.get(normaliza(conceito.strip()))
    if faixa is None:
        return None
    return faixa[0] <= nota <= faixa[1]

def calc_cert_hom(j, l):
    if is_empty(j) or is_empty(l):
        return "-"
    c = concordam(j, l)
    return "NÃO" if c is True else ("SIM" if c is False else "-")

def calc_status(j, l, n):
    if is_empty(j):
        return "Aberta"
    if is_empty(l):
        return "Parcialmente Encerrada"
    c = concordam(j, l)
    if c is True:
        return "Encerrada"
    elif c is False:
        return "Encerrada" if not is_empty(n) else "Homologação"
    return "Encerrada"

def fmt_date(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%d/%m/%Y")
    s = str(val).strip()
    if s in ("", "NaT", "nan", "00/00/0000"):
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10], fmt[:10]).strftime("%d/%m/%Y")
        except:
            pass
    return s

def parse_date_safe(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, float) and pd.isna(val):
        return None
    s = str(val).strip()
    if s in ("", "NaT", "nan", "00/00/0000", "None"):
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10], fmt[:10]).date()
        except:
            pass
    return None

def safe_sheet_name(name, used=None):
    """Cria nome de aba válido para Excel (máx 31 chars, sem chars especiais)."""
    s = re.sub(r'[\\/*?:\[\]]', '_', str(name).strip())
    s = s[:31]
    if used is not None:
        base = s
        i = 1
        while s in used:
            suffix = f"_{i}"
            s = base[:31-len(suffix)] + suffix
            i += 1
        used.add(s)
    return s

# ─── 1. SIGEF ────────────────────────────────────────────────────────────────
print("Carregando SIGEF.csv ...")
sigef_unidade = {}
with open(SIGEF_FILE, encoding="cp1252", errors="replace") as f:
    reader = csv.reader(f, delimiter=";")
    next(reader)
    for row in reader:
        if len(row) > 9:
            nrpm = row[0].strip().lstrip("0") or "0"
            sigef_unidade[nrpm] = row[9].strip()

print("Processando Data do CDP para Situação Comissão...")
pm_max_cdp = {}
with open(GERAL_FILE, encoding="cp1252", errors="replace") as f:
    reader = csv.reader(f, delimiter=";")
    next(reader)
    for row in reader:
        if len(row) > 11:
            sit = row[11].strip()
            if sit in SITUACOES_ALVO:
                nrpm = row[1].strip().lstrip("0") or "0"
                dt_cdp = parse_date_safe(row[5].strip())
                if dt_cdp:
                    if nrpm not in pm_max_cdp or dt_cdp > pm_max_cdp[nrpm]:
                        pm_max_cdp[nrpm] = dt_cdp

# ─── 2. MOVIMENTAÇÕES ────────────────────────────────────────────────────────
print("Carregando MOVIMENTAÇÕES.xlsx ...")
df_mov = pd.read_excel(MOV_FILE, dtype=str)
df_mov.columns = [c.strip() for c in df_mov.columns]
mov_mais_recente = {}
mov_map = defaultdict(list)
for _, row in df_mov.iterrows():
    nrpm = str(row.get("Numero Servidor", "")).strip().lstrip("0") or "0"
    dt_raw = row.get("Dt movimentacao", row.get(" Dt movimentacao", ""))
    motivo = str(row.get("Motivo movimentacao", "")).strip()
    dt = parse_date_safe(dt_raw)
    if nrpm and nrpm != "0":
        mov_map[nrpm].append({"data": dt, "motivo": motivo})
for nrpm, lista in mov_map.items():
    lista_valida = [x for x in lista if x["data"] is not None]
    if lista_valida:
        mov_mais_recente[nrpm] = max(lista_valida, key=lambda x: x["data"])
    elif lista:
        mov_mais_recente[nrpm] = lista[0]

# ─── 3. PRESO ────────────────────────────────────────────────────────────────
print("Carregando PRESO.xlsx ...")
df_preso = pd.read_excel(PRESO_FILE, dtype=str)
df_preso.columns = [c.strip() for c in df_preso.columns]
preso_map = defaultdict(list)
for _, row in df_preso.iterrows():
    nrpm = str(row.get("NUMERO", "")).strip().lstrip("0") or "0"
    data_inicio = parse_date_safe(row.get("DATA INICIO", ""))
    data_fim    = parse_date_safe(row.get("DATA TERMINO", ""))
    sit_func    = str(row.get("SITUACAO FUNCIONAL", "")).strip()
    if nrpm and nrpm != "0":
        preso_map[nrpm].append({"data_inicio": data_inicio, "data_fim": data_fim, "sit_func": sit_func})
preso_info = {}
for nrpm, lista in preso_map.items():
    lista_valida = [x for x in lista if x["data_inicio"] is not None]
    if lista_valida:
        preso_info[nrpm] = max(lista_valida, key=lambda x: x["data_inicio"])

# ─── 4. PUNIÇÃO ──────────────────────────────────────────────────────────────
print("Carregando PUNIÇÃO.xlsx ...")
df_pun = pd.read_excel(PUN_FILE, dtype=str)
df_pun.columns = [c.strip() for c in df_pun.columns]
pun_map = defaultdict(lambda: {"qtd_punicoes": 0, "qtd_ativadas": 0, "pontos": 0})
for _, row in df_pun.iterrows():
    nrpm = str(row.get("MATRICULA", "")).strip().lstrip("0") or "0"
    if not nrpm or nrpm == "0":
        continue
    data_pun = parse_date_safe(row.get("DATA PUNICAO", ""))
    if data_pun is None:
        continue
    pun_map[nrpm]["qtd_punicoes"] += 1
    data_ativ_raw = str(row.get("DATA ATIVACAO", "")).strip()
    if data_ativ_raw not in ("00/00/0000", "", "nan", "None", "NaT"):
        if parse_date_safe(data_ativ_raw) is not None:
            pun_map[nrpm]["qtd_ativadas"] += 1
    try:
        pontos = int(float(str(row.get("PONTOS", "0")).replace(",", ".")))
    except:
        pontos = 0
    pun_map[nrpm]["pontos"] += pontos

# ─── 5. GERAL.CSV ────────────────────────────────────────────────────────────
print("Carregando geral.csv ...")
rows = []
with open(GERAL_FILE, encoding="cp1252", errors="replace") as f:
    reader = csv.reader(f, delimiter=";")
    next(reader)
    for row in reader:
        while len(row) < 198:
            row.append("")
        sit = row[11].strip()
        if sit not in SITUACOES_ALVO:
            continue
        nrpm  = row[1].strip()
        local = row[9].strip()
        nrpm_key = nrpm.lstrip("0") or "0"
        j = row[46].strip()
        l = row[47].strip()
        n = row[70].strip()
        sigef_unit = sigef_unidade.get(nrpm_key, "")
        dt_cdp = parse_date_safe(row[5].strip())
        if dt_cdp and nrpm_key in pm_max_cdp:
            sc = "Comissão Atual" if dt_cdp >= pm_max_cdp[nrpm_key] else "Nota Provisória"
        else:
            sc = "Comissão Atual" if local.upper().strip() == sigef_unit.upper().strip() else "Nota Provisória"

        # Enriquecimento — Movimentação
        mov = mov_mais_recente.get(nrpm_key)
        data_transf  = fmt_date(mov["data"]) if mov and mov["data"] else ""
        motivo_transf = mov["motivo"] if mov else ""

        # Enriquecimento — Preso
        preso = preso_info.get(nrpm_key)
        if preso:
            dt_ini = preso["data_inicio"]
            dt_fim = preso["data_fim"]
            data_ini_preso = fmt_date(dt_ini) if dt_ini else ""
            data_fim_preso = fmt_date(dt_fim) if dt_fim else ""
            sit_preso = preso["sit_func"]
            if dt_ini:
                ref = dt_fim if dt_fim else TODAY
                dias_preso = max(0, (ref - dt_ini).days)
            else:
                dias_preso = ""
        else:
            data_ini_preso = data_fim_preso = sit_preso = ""
            dias_preso = ""

        # Enriquecimento — Punição
        pun = pun_map.get(nrpm_key)
        if pun and pun["qtd_punicoes"] > 0:
            qtd_pun   = pun["qtd_punicoes"]
            qtd_ativ  = pun["qtd_ativadas"]
            pontos    = pun["pontos"]
        else:
            qtd_pun = qtd_ativ = pontos = ""

        rows.append({
            "nrPM (Avaliado)":                nrpm,
            "Nome Completo (Avaliado)":        row[2].strip(),
            "Posto/Graduação (Avaliado)":      row[6].strip(),
            "Unidade RPM (Avaliado)":          row[7].strip(),
            "Unidade Principal (Avaliado)":    row[8].strip(),
            "Local/Unidade (Avaliado)":        local,
            "Quadro Atual (Avaliado)":         row[10].strip(),
            "Situação Funcional Atual":        sit,
            "Data Avaliação 1":                row[36].strip(),
            "Conceito Geral":                  j,
            "Data Avaliação 2":                row[71].strip() if not is_empty(n) else row[45].strip(),
            "Nota Geral":                      n if not is_empty(n) else l,
            "Certificação Homologador":        calc_cert_hom(j, l),
            "Data Homologação":                row[71].strip(),
            "Nota Homologação":                n,
            "Competência 1":                   row[48].strip(),
            "Conceito (Comp. 1)":              row[49].strip(),
            "Nota (Comp. 1)":                  row[50].strip(),
            "Competência 2":                   row[51].strip(),
            "Conceito (Comp. 2)":              row[52].strip(),
            "Nota (Comp. 2)":                  row[53].strip(),
            "Competência 3":                   row[54].strip(),
            "Conceito (Comp. 3)":              row[55].strip(),
            "Nota (Comp. 3)":                  row[56].strip(),
            "Competência 4":                   row[57].strip(),
            "Conceito (Comp. 4)":              row[58].strip(),
            "Nota (Comp. 4)":                  row[59].strip(),
            "nrPM (Avaliador 1)":              row[28].strip(),
            "Nome (Avaliador 1)":              row[29].strip(),
            "Posto (Avaliador 1)":             row[30].strip(),
            "RPM (Avaliador 1)":               row[31].strip(),
            "Unid. Principal (Avaliador 1)":   row[32].strip(),
            "Local (Avaliador 1)":             row[33].strip(),
            "Quadro (Avaliador 1)":            row[34].strip(),
            "Situação (Avaliador 1)":          row[35].strip(),
            "nrPM (Avaliador 2)":              row[37].strip(),
            "Nome (Avaliador 2)":              row[38].strip(),
            "Posto (Avaliador 2)":             row[39].strip(),
            "RPM (Avaliador 2)":               row[40].strip(),
            "Unid. Principal (Avaliador 2)":   row[41].strip(),
            "Local (Avaliador 2)":             row[42].strip(),
            "Quadro (Avaliador 2)":            row[43].strip(),
            "Situação (Avaliador 2)":          row[44].strip(),
            "nrPM (Homologador)":              row[62].strip(),
            "Nome (Homologador)":              row[63].strip(),
            "Posto (Homologador)":             row[64].strip(),
            "RPM (Homologador)":               row[65].strip(),
            "Unid. Principal (Homologador)":   row[66].strip(),
            "Local (Homologador)":             row[67].strip(),
            "Quadro (Homologador)":            row[68].strip(),
            "Situação (Homologador)":          row[69].strip(),
            "Situação Comissão":               sc,
            "Status Avaliação":                calc_status(j, l, n),
            "Recurso Fase 1":                  row[73].strip(),
            "Nota (Fase 1)":                   row[74].strip(),
            "Data Recurso 1":                  row[76].strip(),
            "Recurso Fase 2":                  row[77].strip(),
            "Nota (Fase 2)":                   row[78].strip(),
            "Data Recurso 2":                  row[80].strip(),
            "Recurso Fase 3":                  row[81].strip(),
            "Nota (Fase 3)":                   row[82].strip(),
            "Data Recurso 3":                  row[84].strip(),
            "Recurso Fase 4":                  row[85].strip(),
            "Nota (Fase 4)":                   row[86].strip(),
            "Data Recurso 4":                  row[88].strip(),
            # Novas colunas temáticas
            "DATA DA TRANSFERÊNCIA":           data_transf,
            "MOTIVO DA TRANSFERÊNCIA":         motivo_transf,
            "DATA INÍCIO DA PRISÃO":           data_ini_preso,
            "DATA FIM DA PRISÃO":              data_fim_preso,
            "SITUAÇÃO FUNCIONAL (PRESO)":      sit_preso,
            "DIAS PRESO":                      dias_preso,
            "QTDADE DE PUNIÇÕES":              qtd_pun,
            "QTDADE DE ATIVADAS":              qtd_ativ,
            "PONTOS PERDIDOS":                 pontos,
        })

df_full = pd.DataFrame(rows)
print(f"  {len(df_full):,} registros processados")

# ─── FILTROS TEMÁTICOS ────────────────────────────────────────────────────────
# "Receberam nota" = Nota Geral preenchida
df_full["_tem_nota"] = df_full["Nota Geral"].replace("", pd.NA).notna()

df_pun_tema   = df_full[df_full["_tem_nota"] & (df_full["QTDADE DE PUNIÇÕES"] != "")].copy()
df_preso_tema = df_full[df_full["_tem_nota"] & (df_full["DATA INÍCIO DA PRISÃO"] != "")].copy()
df_mov_tema   = df_full[df_full["_tem_nota"] & (df_full["DATA DA TRANSFERÊNCIA"] != "")].copy()

print(f"  Militares com nota + punição   : {len(df_pun_tema):,}")
print(f"  Militares com nota + prisão    : {len(df_preso_tema):,}")
print(f"  Militares com nota + transf.   : {len(df_mov_tema):,}")

# Remove coluna auxiliar
df_full.drop(columns=["_tem_nota"], inplace=True)
df_pun_tema.drop(columns=["_tem_nota"], inplace=True)
df_preso_tema.drop(columns=["_tem_nota"], inplace=True)
df_mov_tema.drop(columns=["_tem_nota"], inplace=True)

# ─── COLUNAS POR TEMA ─────────────────────────────────────────────────────────
COLS_BASE = [
    "nrPM (Avaliado)", "Nome Completo (Avaliado)", "Posto/Graduação (Avaliado)",
    "Unidade RPM (Avaliado)", "Unidade Principal (Avaliado)", "Local/Unidade (Avaliado)",
    "Quadro Atual (Avaliado)", "Situação Funcional Atual",
    "Data Avaliação 1", "Conceito Geral",
    "Data Avaliação 2", "Nota Geral",
    "Certificação Homologador", "Data Homologação", "Nota Homologação",
    "Competência 1", "Conceito (Comp. 1)", "Nota (Comp. 1)",
    "Competência 2", "Conceito (Comp. 2)", "Nota (Comp. 2)",
    "Competência 3", "Conceito (Comp. 3)", "Nota (Comp. 3)",
    "Competência 4", "Conceito (Comp. 4)", "Nota (Comp. 4)",
    "nrPM (Avaliador 1)", "Nome (Avaliador 1)", "Posto (Avaliador 1)",
    "RPM (Avaliador 1)", "Unid. Principal (Avaliador 1)",
    "nrPM (Avaliador 2)", "Nome (Avaliador 2)", "Posto (Avaliador 2)",
    "RPM (Avaliador 2)", "Unid. Principal (Avaliador 2)",
    "nrPM (Homologador)", "Nome (Homologador)", "Posto (Homologador)",
    "RPM (Homologador)", "Unid. Principal (Homologador)",
    "Situação Comissão", "Status Avaliação",
    "Recurso Fase 1", "Nota (Fase 1)", "Data Recurso 1",
    "Recurso Fase 2", "Nota (Fase 2)", "Data Recurso 2",
    "Recurso Fase 3", "Nota (Fase 3)", "Data Recurso 3",
    "Recurso Fase 4", "Nota (Fase 4)", "Data Recurso 4",
]

COLS_PUNICAO = COLS_BASE + [
    "QTDADE DE PUNIÇÕES", "QTDADE DE ATIVADAS", "PONTOS PERDIDOS"
]

COLS_PRISAO = COLS_BASE + [
    "DATA INÍCIO DA PRISÃO", "DATA FIM DA PRISÃO",
    "SITUAÇÃO FUNCIONAL (PRESO)", "DIAS PRESO"
]

COLS_MOVIMENTACAO = COLS_BASE + [
    "DATA DA TRANSFERÊNCIA", "MOTIVO DA TRANSFERÊNCIA"
]

# ─── ESTILOS ──────────────────────────────────────────────────────────────────
STATUS_BG = {
    "Encerrada":              "70AD47",
    "Homologação":            "FFD966",
    "Parcialmente Encerrada": "FF8C00",
    "Aberta":                 "FF4444",
}
SIT_BG = {"Comissão Atual": "4472C4", "Nota Provisória": "FFC000"}

thin = Side(border_style="thin", color="CCCCCC")
BRD  = Border(left=thin, right=thin, top=thin, bottom=thin)

TEMA_CORES = {
    "pun":  {"hdr": "7B2D00", "nova": "C0392B", "nova_fundo": "FADBD8"},  # Vermelho
    "pres": {"hdr": "1A237E", "nova": "2980B9", "nova_fundo": "D6EAF8"},  # Azul escuro
    "mov":  {"hdr": "1B5E20", "nova": "27AE60", "nova_fundo": "D5F5E3"},  # Verde
}

NOVAS_PUN  = {"QTDADE DE PUNIÇÕES", "QTDADE DE ATIVADAS", "PONTOS PERDIDOS"}
NOVAS_PRES = {"DATA INÍCIO DA PRISÃO", "DATA FIM DA PRISÃO", "SITUAÇÃO FUNCIONAL (PRESO)", "DIAS PRESO"}
NOVAS_MOV  = {"DATA DA TRANSFERÊNCIA", "MOTIVO DA TRANSFERÊNCIA"}

def write_theme_sheet(ws, df, titulo, cols, cor_tema, novas_cols, numero=None, total=None):
    """Escreve uma aba temática formatada."""
    hdr_hex  = cor_tema["hdr"]
    nova_hex = cor_tema["nova"]
    fundo_hex = cor_tema["nova_fundo"]

    actual_cols = [c for c in cols if c in df.columns]
    n_cols = len(actual_cols)

    # Linha 1 — Título
    if n_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(1, 1, titulo)
    c.fill = PatternFill("solid", fgColor=hdr_hex)
    c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Linha 2 — Sub-info (qtde de registros)
    if numero is not None and total is not None:
        info = f"Registros nesta unidade: {numero} | Total geral: {total}"
    elif numero is not None:
        info = f"Total de registros: {numero}"
    else:
        info = ""
    if info:
        if n_cols > 1:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        ci = ws.cell(2, 1, info)
        ci.fill = PatternFill("solid", fgColor="F5F5F5")
        ci.font = Font(italic=True, name="Calibri", size=9, color="555555")
        ci.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 16
        hdr_row = 3
    else:
        hdr_row = 2

    # Cabeçalho
    for ci, col in enumerate(actual_cols, 1):
        fill_hex = nova_hex if col in novas_cols else hdr_hex
        cell = ws.cell(hdr_row, ci, col)
        cell.fill = PatternFill("solid", fgColor=fill_hex)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BRD
    ws.row_dimensions[hdr_row].height = 34
    ws.freeze_panes = ws.cell(hdr_row + 1, 1).coordinate

    # Dados
    df_out = df[actual_cols].reset_index(drop=True)
    for r_idx, (_, row) in enumerate(df_out.iterrows(), hdr_row + 1):
        for ci, col in enumerate(actual_cols, 1):
            val = row.get(col, "")
            txt = "" if (isinstance(val, float) and pd.isna(val)) else (str(val) if val != "" else "")
            cell = ws.cell(r_idx, ci, txt)
            cell.font  = Font(name="Calibri", size=9)
            cell.border = BRD
            cell.alignment = Alignment(horizontal="center", vertical="center") if ci > 8 else Alignment(vertical="center")

            if col == "Status Avaliação" and val in STATUS_BG:
                cell.fill = PatternFill("solid", fgColor=STATUS_BG[val])
                cell.font = Font(bold=True, name="Calibri", size=9,
                                 color="FFFFFF" if val == "Aberta" else "000000")
            elif col == "Situação Comissão" and val in SIT_BG:
                cell.fill = PatternFill("solid", fgColor=SIT_BG[val])
                cell.font = Font(bold=True, name="Calibri", size=9,
                                 color="FFFFFF" if val == "Comissão Atual" else "000000")
            elif col in novas_cols and txt:
                cell.fill = PatternFill("solid", fgColor=fundo_hex)
                cell.font = Font(bold=True, name="Calibri", size=9, color="000000")

    # Largura automática
    for ci, col in enumerate(actual_cols, 1):
        max_len = len(str(col))
        if col in df_out.columns and not df_out.empty:
            m = df_out[col].astype(str).str.len().max()
            if pd.notna(m):
                max_len = max(max_len, int(m))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len * 0.9, 8), 45)

    return len(df_out)

def rpm_sort_key(name):
    m = re.match(r'^(\d+)\s+RPM', str(name))
    return (0, int(m.group(1)), "") if m else (1, 0, str(name))

def build_workbook(df_tema, titulo_wb, cols_tema, cor_tema, novas_cols, nome_arquivo):
    """Cria o workbook completo com aba Geral + abas por Unidade RPM."""
    wb = Workbook()
    total_geral = len(df_tema)

    # ── Aba GERAL ──────────────────────────────────────────────────────────────
    ws_geral = wb.active
    ws_geral.title = "Geral"
    write_theme_sheet(
        ws_geral, df_tema,
        f"{titulo_wb} — TODOS ({total_geral} registros) — {TODAY.strftime('%d/%m/%Y')}",
        cols_tema, cor_tema, novas_cols, numero=total_geral
    )
    print(f"    Aba 'Geral': {total_geral} registros")

    # ── Abas por Unidade RPM ───────────────────────────────────────────────────
    rpms = sorted(df_tema["Unidade RPM (Avaliado)"].dropna().unique(), key=rpm_sort_key)
    used_names = {"Geral"}

    for rpm in rpms:
        df_rpm = df_tema[df_tema["Unidade RPM (Avaliado)"] == rpm].copy()
        if df_rpm.empty:
            continue
        tab_name = safe_sheet_name(rpm, used_names)
        ws = wb.create_sheet(tab_name)
        n = write_theme_sheet(
            ws, df_rpm,
            f"{titulo_wb} — {rpm} ({len(df_rpm)} registros)",
            cols_tema, cor_tema, novas_cols,
            numero=len(df_rpm), total=total_geral
        )
        print(f"    Aba '{tab_name}': {n} registros")

    # ── Salva ──────────────────────────────────────────────────────────────────
    out_path = os.path.join(OUT_DIR, nome_arquivo)
    wb.save(out_path)
    print(f"  ✅ Salvo: {out_path}")
    return out_path

# ─── GERAR AS 3 PLANILHAS ─────────────────────────────────────────────────────
print("\n=== [1/3] Gerando planilha de PUNIÇÕES ===")
f1 = build_workbook(
    df_pun_tema,
    "MILITARES COM NOTA + PUNIÇÃO — AADP 2026",
    COLS_PUNICAO,
    TEMA_CORES["pun"],
    NOVAS_PUN,
    "Tematica_Punicao_AADP2026.xlsx"
)

print("\n=== [2/3] Gerando planilha de PRISÃO ===")
f2 = build_workbook(
    df_preso_tema,
    "MILITARES COM NOTA + PRISÃO — AADP 2026",
    COLS_PRISAO,
    TEMA_CORES["pres"],
    NOVAS_PRES,
    "Tematica_Prisao_AADP2026.xlsx"
)

print("\n=== [3/3] Gerando planilha de MOVIMENTAÇÃO ===")
f3 = build_workbook(
    df_mov_tema,
    "MILITARES COM NOTA + MOVIMENTAÇÃO — AADP 2026",
    COLS_MOVIMENTACAO,
    TEMA_CORES["mov"],
    NOVAS_MOV,
    "Tematica_Movimentacao_AADP2026.xlsx"
)

print(f"""
╔══════════════════════════════════════════════════════════╗
║          GERAÇÃO CONCLUÍDA — {TODAY.strftime('%d/%m/%Y')}                  ║
╠══════════════════════════════════════════════════════════╣
║  1. Tematica_Punicao_AADP2026.xlsx                       ║
║     → {len(df_pun_tema):,} militares com nota + punição                 
║  2. Tematica_Prisao_AADP2026.xlsx                        ║
║     → {len(df_preso_tema):,} militares com nota + prisão                   
║  3. Tematica_Movimentacao_AADP2026.xlsx                  ║
║     → {len(df_mov_tema):,} militares com nota + movimentação             
╚══════════════════════════════════════════════════════════╝
""")
