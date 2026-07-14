"""


AADP 2026 — Dashboard de Análise de Avaliações


Versão 3.0 — Google Drive + Streamlit Cloud + Geração em memória


"""


import time


_prof_start = time.time()


import streamlit as st


import pandas as pd


import plotly.express as px


import plotly.graph_objects as go


import io, os, re, json, subprocess, unicodedata, csv, tempfile, zipfile, sqlite3, hashlib


from pathlib import Path


from collections import defaultdict


from datetime import datetime, timezone, timedelta





def now_br():


    """Retorna datetime no fuso de Brasília (UTC-3)."""


    return datetime.now(timezone(timedelta(hours=-3)))


@st.cache_data(ttl=600)


def get_last_updated_time(av_f, drive_av_id=None):


    """Retorna a data e hora de consolidação dos dados em horário de Brasília."""


    import requests, email.utils, os


    dt_utc = None


    if drive_av_id:


        try:


            url = f"https://drive.google.com/uc?id={drive_av_id}&export=download"


            r = requests.head(url, allow_redirects=True, timeout=5)


            last_mod = r.headers.get("Last-Modified")


            if last_mod:


                dt_utc = email.utils.parsedate_to_datetime(last_mod)


        except Exception:


            pass


            


    if not dt_utc and os.path.exists(av_f):


        try:


            mtime = os.path.getmtime(av_f)


            dt_utc = datetime.fromtimestamp(mtime, timezone.utc)


        except Exception:


            pass


            


    if dt_utc:


        tz_br = timezone(timedelta(hours=-3))


        dt_br = dt_utc.astimezone(tz_br)


        return dt_br.strftime("%d/%m/%Y, às %H:%M horas")


        


    return "Data/Hora indisponível"







def normalize_pm(v):
    if pd.isna(v): return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.lstrip("0")

def parse_float(v):
    if pd.isna(v): return None
    s = str(v).strip().replace(",", ".")
    if s in ("", "-", "nan", "none"): return None
    try:
        return float(s)
    except ValueError:
        return None

@st.cache_resource(show_spinner=False)
def run_grades_audit(xlsx_path, csv_path, drive_geral_id=None, drive_master_xlsx_id=None):
    import pandas as pd
    import numpy as np
    import csv
    import os
    
    # Se estamos no modo Drive e temos os IDs configurados, baixar os arquivos primeiro
    if drive_geral_id:
        try:
            if not os.path.exists(csv_path) or os.path.getsize(csv_path) == 0:
                _baixar_drive(drive_geral_id, csv_path)
        except Exception as e:
            return None, f"Falha ao baixar geral.csv do Google Drive (ID: {drive_geral_id}): {str(e)}"
            
    if drive_master_xlsx_id:
        try:
            if not os.path.exists(xlsx_path) or os.path.getsize(xlsx_path) == 0:
                _baixar_drive(drive_master_xlsx_id, xlsx_path)
        except Exception as e:
            return None, f"Falha ao baixar master Excel do Google Drive (ID: {drive_master_xlsx_id}): {str(e)}"

    CONCEITO_FAIXA = {
        "nivel superior de desempenho":       (9.00, 10.00),
        "nivel alto de desempenho":           (7.00,  8.99),
        "nivel intermediario de desempenho":  (6.00,  6.99),
        "nivel baixo de desempenho":          (3.00,  5.99),
        "nivel inferior de desempenho":       (0.00,  2.99),
    }

    def normaliza(texto: str) -> str:
        if not isinstance(texto, str): return ""
        import unicodedata
        t = unicodedata.normalize("NFD", texto.lower())
        return "".join(c for c in t if unicodedata.category(c) != "Mn")

    try:
        df_master = pd.read_excel(xlsx_path)
        df_master['NR PM'] = df_master['NR PM'].apply(normalize_pm)
    except Exception as e:
        return None, f"Erro ao ler excel master (caminho: {xlsx_path}): {str(e)}"

    if not os.path.exists(csv_path):
        return None, f"Arquivo geral.csv nao encontrado no caminho: {csv_path}."

    cols_to_keep = {
        'nrPM (Avaliado)': 1, 'Nome Completo (Avaliado)': 2, 'Conceito Geral': 46, 'Nota Geral': 47,
        'Nota (Competência 1)': 50, 'Nota (Competência 2)': 53, 'Nota (Competência 3)': 56, 'Nota (Competência 4)': 59,
        'Nota da Homologação': 70, 'Data da Homologação': 71,
        'Recurso Fase 2': 77, 'Nota (Fase 2)': 78, 'Recurso Fase 3': 81, 'Nota (Fase 3)': 82,
        'Data da Avaliação 1': 36, 'Data da Avaliação 2': 45
    }
    
    rows = []
    try:
        with open(csv_path, "r", encoding="cp1252", errors="ignore") as f:
            reader = csv.reader(f, delimiter=";")
            header = next(reader)
            col_indices = {col: header.index(col) for col in cols_to_keep if col in header}
            for r in reader:
                row_clean = r[:len(header)]
                while len(row_clean) < len(header):
                    row_clean.append("")
                extracted_row = {}
                for col_name, col_idx in col_indices.items():
                    extracted_row[col_name] = row_clean[col_idx]
                rows.append(extracted_row)
        df_geral = pd.DataFrame(rows)
        df_geral['nrPM (Avaliado)'] = df_geral['nrPM (Avaliado)'].apply(normalize_pm)
    except Exception as e:
        return None, f"Erro ao ler geral.csv (caminho: {csv_path}): {str(e)}"

    discrepancies = []

    # 1. Auditoria de Qtd de Avaliações
    master_counts = df_master.set_index('NR PM')['Qtd Avaliações'].to_dict()
    geral_counts = df_geral['nrPM (Avaliado)'].value_counts().to_dict()
    all_pms = set(master_counts.keys()).union(set(geral_counts.keys()))

    for pm in sorted(list(all_pms)):
        if pm in ('nan', '', 'None'): continue
        m_c = master_counts.get(pm, 0)
        g_c = geral_counts.get(pm, 0)
        if m_c != g_c:
            name = ""
            if pm in master_counts:
                name = df_master[df_master['NR PM'] == pm]['Nome Completo'].values[0]
            else:
                matching_rows = df_geral[df_geral['nrPM (Avaliado)'] == pm]
                if not matching_rows.empty:
                    name = matching_rows['Nome Completo (Avaliado)'].values[0]
            discrepancies.append({
                "PM": pm,
                "Nome": name,
                "Tipo": "Divergência de Qtd de Avaliações",
                "Detalhe": f"Excel mestre diz {m_c} avaliações, mas geral.csv possui {g_c} registros."
            })

    # 2. Auditorias de Notas por registro de geral.csv
    for idx, row in df_geral.iterrows():
        pm = row['nrPM (Avaliado)']
        name = row['Nome Completo (Avaliado)']
        
        n_g = parse_float(row['Nota Geral'])
        c1 = parse_float(row['Nota (Competência 1)'])
        c2 = parse_float(row['Nota (Competência 2)'])
        c3 = parse_float(row['Nota (Competência 3)'])
        c4 = parse_float(row['Nota (Competência 4)'])
        
        if n_g is not None and all(x is not None for x in [c1, c2, c3, c4]):
            avg_comp = (c1 + c2 + c3 + c4) / 4.0
            if abs(avg_comp - n_g) > 0.01:
                discrepancies.append({
                    "PM": pm,
                    "Nome": name,
                    "Tipo": "Divergência de Média de Competências",
                    "Detalhe": f"Média das Competências = {avg_comp:.2f} (C1={c1}, C2={c2}, C3={c3}, C4={c4}) vs Nota Geral informada = {n_g}"
                })
                
        concept = row['Conceito Geral']
        n_hom = parse_float(row['Nota da Homologação'])
        
        if n_g is not None and not pd.isna(concept) and concept != '-':
            concept_norm = normaliza(str(concept))
            faixa = CONCEITO_FAIXA.get(concept_norm)
            if faixa:
                is_divergent = not (faixa[0] <= n_g <= faixa[1])
                if is_divergent:
                    if n_hom is None:
                        discrepancies.append({
                            "PM": pm,
                            "Nome": name,
                            "Tipo": "Divergência de Nota de Homologação",
                            "Detalhe": f"Divergência entre Conceito Geral ('{concept}') e Nota Geral ({n_g}), mas sem Nota de Homologação cadastrada."
                        })

    return pd.DataFrame(discrepancies), None




def build_audit_data_from_geral(csv_path):
    import csv
    import os
    import pandas as pd
    import numpy as np
    import math
    import unicodedata

    SITUACOES_ALVO = {
        "ATIV. DIRECAO GERAL", "ATIV. FIM DESTACADO", "ATIV. FIM NA SEDE",
        "ATIV. MEIO", "ATIVIDADE MEIO", "DISP MED DEFINITIVA",
        "GESTANTE/LAC/ADOTANT", "QUADRO ESPECIALISTA",
    }

    CONCEITO_FAIXA = {
        "nivel superior de desempenho":       (9.00, 10.00),
        "nivel alto de desempenho":           (7.00,  8.99),
        "nivel intermediario de desempenho":  (6.00,  6.99),
        "nivel baixo de desempenho":          (3.00,  5.99),
        "nivel inferior de desempenho":       (0.00,  2.99),
    }

    def normaliza(texto: str) -> str:
        if not isinstance(texto, str): return ""
        t = unicodedata.normalize("NFD", texto.lower())
        return "".join(c for c in t if unicodedata.category(c) != "Mn")

    def is_empty(v) -> bool:
        if v == 0 or v == 0.0:
            return False
        return not v or str(v).strip() in ("", "-", "nan", "none", "None", "<NA>")

    def parse_float(s):
        try:
            if is_empty(s): return None
            return float(str(s).replace(",", "."))
        except ValueError:
            return None

    def concordam(conceito: str, nota_str: str):
        if is_empty(conceito) or is_empty(nota_str):
            return None
        nota = parse_float(nota_str)
        if nota is None:
            return None
        faixa = CONCEITO_FAIXA.get(normaliza(conceito.strip()))
        if faixa is None:
            return None
        return faixa[0] <= nota <= faixa[1]

    def calc_status(j: str, l: str, n: str) -> str:
        if is_empty(j):
            return "Aberta"
        if is_empty(l):
            return "Parcialmente Encerrada"
        c = concordam(j, l)
        if c is True:
            return "Encerrada"
        elif c is False:
            return "Encerrada" if not is_empty(n) else "Homologação"
        return "Parcialmente Encerrada"

    def normalize_pm(pm):
        try:
            if is_empty(pm): return ""
            return str(int(float(str(pm).strip())))
        except Exception:
            return str(pm).strip()

    def find_col(header, pattern):
        pattern_norm = normaliza(pattern)
        for idx, col in enumerate(header):
            if pattern_norm in normaliza(col):
                return idx
        raise ValueError(f"Coluna contendo '{pattern}' não encontrada no geral.csv.")

    pm_evals = {}
    
    with open(csv_path, "r", encoding="cp1252", errors="ignore") as f:
        reader = csv.reader(f, delimiter=";")
        header = next(reader)
        
        c_pm = find_col(header, "nrPM (Avaliado)")
        c_name = find_col(header, "Nome Completo (Avaliado)")
        c_rank = find_col(header, "Posto/Grad")
        c_rpm = find_col(header, "Unidade RPM (Avaliado)")
        c_unit = find_col(header, "Unidade Principal (Avaliado)")
        c_quadro = find_col(header, "Quadro Atual (Avaliado)")
        c_sit = find_col(header, "Situação Funcional")
        c_dt_av1 = find_col(header, "Data da Avaliação 1")
        c_dt_av2 = find_col(header, "Data da Avaliação 2")
        c_concept = find_col(header, "Conceito Geral")
        c_grade = find_col(header, "Nota Geral")
        c_n_hom = find_col(header, "Nota da Homologação")
        c_dt_hom = find_col(header, "Data da Homologação")
        
        c_n_f1 = find_col(header, "Nota (Fase 1)")
        c_n_f2 = find_col(header, "Nota (Fase 2)")
        c_n_f3 = find_col(header, "Nota (Fase 3)")
        c_n_f4 = find_col(header, "Nota (Fase 4)")
        
        for row in reader:
            if len(row) < len(header):
                continue
            sit = row[c_sit].strip()
            if sit not in SITUACOES_ALVO:
                continue
                
            pm = normalize_pm(row[c_pm])
            if not pm:
                continue
                
            j = row[c_concept].strip()
            l = row[c_grade].strip()
            n = row[c_n_hom].strip()
            status = calc_status(j, l, n)
            
            n_f4 = parse_float(row[c_n_f4])
            n_f3 = parse_float(row[c_n_f3])
            n_f2 = parse_float(row[c_n_f2])
            n_f1 = parse_float(row[c_n_f1])
            
            final_grade = None
            houve_recurso = "-"
            fase_recurso = "-"
            nota_recurso = "-"
            
            if n_f4 is not None:
                final_grade = n_f4
                houve_recurso = "SIM"
                fase_recurso = "FASE 4"
                nota_recurso = n_f4
            elif n_f3 is not None:
                final_grade = n_f3
                houve_recurso = "SIM"
                fase_recurso = "FASE 3"
                nota_recurso = n_f3
            elif n_f2 is not None:
                final_grade = n_f2
                houve_recurso = "SIM"
                fase_recurso = "FASE 2"
                nota_recurso = n_f2
            elif n_f1 is not None:
                final_grade = n_f1
                houve_recurso = "SIM"
                fase_recurso = "FASE 1"
                nota_recurso = n_f1
            elif not is_empty(n):
                final_grade = parse_float(n)
            else:
                final_grade = parse_float(l)
                
            dt_av = row[c_dt_av2].strip() or row[c_dt_av1].strip() or row[c_dt_hom].strip() or "-"
            
            eval_data = {
                "date": dt_av,
                "status": status.upper(),
                "grade": parse_float(l) if not is_empty(l) else "-",
                "houve_recurso": houve_recurso,
                "fase_recurso": fase_recurso,
                "nota_recurso": nota_recurso,
                "final_grade": final_grade
            }
            
            if pm not in pm_evals:
                pm_evals[pm] = {
                    "NR PM": pm,
                    "Posto/Graduação": row[c_rank].strip(),
                    "Nome Completo": row[c_name].strip(),
                    "Nome RPM": row[c_rpm].strip(),
                    "Nome Unidade Principal": row[c_unit].strip(),
                    "Quadro": row[c_quadro].strip(),
                    "Sit. Funcional": sit,
                    "evals": []
                }
            pm_evals[pm]["evals"].append(eval_data)
            
    rows_audit = []
    for pm, data in pm_evals.items():
        evals = data["evals"]
        qtd = len(evals)
        todas_encerradas = "SIM" if all(e["status"] == "ENCERRADA" for e in evals) else "NAO"
        
        if todas_encerradas == "SIM":
            grades = [e["final_grade"] for e in evals if e["final_grade"] is not None]
            if len(grades) == qtd:
                final_avg = sum(grades) / float(qtd)
                final_avg_rounded = math.floor(final_avg * 100 + 0.5) / 100.0
            else:
                final_avg_rounded = "-"
        else:
            final_avg_rounded = "-"
            
        r_audit = {
            "NR PM": data["NR PM"],
            "Posto/Graduação": data["Posto/Graduação"],
            "Nome Completo": data["Nome Completo"],
            "Nome RPM": data["Nome RPM"],
            "Nome Unidade Principal": data["Nome Unidade Principal"],
            "Quadro": data["Quadro"],
            "Sit. Funcional": data["Sit. Funcional"],
            "Qtd Avaliações": qtd,
            "Todas Avaliações Foram Encerradas?": todas_encerradas,
            "Nota Final - Média Aritmética": final_avg_rounded,
        }
        
        for i in range(1, 5):
            if i <= qtd:
                ev = evals[i-1]
                r_audit[f"Data Avaliação {i}"] = ev["date"]
                r_audit[f"Fase Avaliação {i}"] = ev["status"]
                r_audit[f"Nota Avaliação {i}"] = ev["grade"]
                r_audit[f"Houve Recurso? {i}"] = ev["houve_recurso"]
                r_audit[f"Fase Recurso {i}"] = ev["fase_recurso"]
                r_audit[f"Nota Fase 2 ou 3 {i}"] = ev["nota_recurso"]
            else:
                r_audit[f"Data Avaliação {i}"] = np.nan
                r_audit[f"Fase Avaliação {i}"] = np.nan
                r_audit[f"Nota Avaliação {i}"] = np.nan
                r_audit[f"Houve Recurso? {i}"] = np.nan
                r_audit[f"Fase Recurso {i}"] = np.nan
                r_audit[f"Nota Fase 2 ou 3 {i}"] = np.nan
        rows_audit.append(r_audit)
        
    return pd.DataFrame(rows_audit)


@st.cache_resource(show_spinner=False)
def load_audit_excel(xlsx_path, drive_master_xlsx_id=None):
    import pandas as pd
    import os
    import tempfile
    from pathlib import Path
    
    cfg_to_use = load_config()
    drive_geral_id = cfg_to_use.get("drive_geral_id", "")
    cache_dir = os.path.join(tempfile.gettempdir(), "aadp_drive_cache")
    drive_geral_path = os.path.join(cache_dir, "geral.csv")
    local_geral_path = os.path.join(str(Path(xlsx_path).parent), "geral.csv")
    
    csv_to_use = None
    if drive_geral_id:
        try:
            os.makedirs(cache_dir, exist_ok=True)
            if not os.path.exists(drive_geral_path) or os.path.getsize(drive_geral_path) == 0:
                _baixar_drive(drive_geral_id, drive_geral_path)
            csv_to_use = drive_geral_path
        except Exception:
            pass
            
    if not csv_to_use and os.path.exists(local_geral_path) and os.path.getsize(local_geral_path) > 0:
        csv_to_use = local_geral_path
        
    if csv_to_use:
        try:
            df = build_audit_data_from_geral(csv_to_use)
            return df, None
        except Exception:
            pass

    # Fallback to master excel
    if drive_master_xlsx_id:
        try:
            if not os.path.exists(xlsx_path) or os.path.getsize(xlsx_path) == 0:
                _baixar_drive(drive_master_xlsx_id, xlsx_path)
        except Exception as e:
            return None, f"Falha ao baixar master Excel do Google Drive (ID: {drive_master_xlsx_id}): {str(e)}"
            
    if not os.path.exists(xlsx_path):
        return None, f"Arquivo Excel consolidado ou geral.csv não encontrado."
        
    try:
        df = pd.read_excel(xlsx_path)
        
        col_media = next((c for c in df.columns if "Aritm" in str(c)), None)
        if col_media:
            import math
            def round_half_up_2(x):
                try:
                    if pd.isna(x) or x is None:
                        return x
                    val = float(str(x).replace(",", "."))
                    return math.floor(val * 100 + 0.5) / 100.0
                except Exception:
                    return x
            df[col_media] = df[col_media].apply(round_half_up_2)
            
        return df, None
    except Exception as e:
        return None, f"Erro ao ler planilha consolidada: {str(e)}"

# gdown: download do Google Drive (opcional — só necessário no modo Drive)


try:


    import gdown


    GDOWN_OK = True


except ImportError:


    GDOWN_OK = False





pd.set_option("styler.render.max_elements", 5_000_000)





st.set_page_config(


    page_title="AADP 2026 — Análise de Avaliações",


    page_icon=None,


    layout="wide",


    initial_sidebar_state="expanded",


)





# ─────────────────────── CSS ──────────────────────────────────────────────────


st.markdown("""
<script>
(function() {
    let changed = false;
    const key = "streamlit:themeConfiguration";
    const expected = '{"themePreset":"dark"}';
    const lightTheme = '{"themePreset":"light"}';
    
    const forceDark = (win) => {
        try {
            const current = win.localStorage.getItem(key);
            if (current !== lightTheme) {
                if (current !== expected) {
                    win.localStorage.setItem(key, expected);
                    changed = true;
                }
                const legacyKeys = ["stActiveTheme-light", "stActiveTheme-dark", "stActiveTheme", "stActiveThemeType", "stTheme", "theme"];
                legacyKeys.forEach(k => {
                    if (win.localStorage.getItem(k)) {
                        win.localStorage.removeItem(k);
                        changed = true;
                    }
                });
            }
        } catch(e) {}
    };

    forceDark(window);
    if (window.parent) {
        forceDark(window.parent);
    }
    if (changed) {
        window.location.reload();
    }
})();
</script>

<style>


@import url('https://fonts.googleapis.com/css2?family=Inter:wght=300;400;500;600;700&display=swap');


html,body,[class*="css"]{font-family:'Inter',sans-serif;}





[data-testid="stSidebar"]{


  background:linear-gradient(180deg,#0c0b07 0%,#1c1c1c 100%);


  border-right:1px solid #9b8a5c;


}


[data-testid="stSidebar"] *{color:#e5dccb!important;}


[data-testid="stSidebar"] h1,[data-testid="stSidebar"] h2,


[data-testid="stSidebar"] h3{color:#9b8a5c!important;}


.main{background:#121212;}





/* Garante que a seta de recolhimento e expansão da barra lateral esteja sempre visível */


[data-testid="collapsedControl"], button[data-testid="stSidebarCollapseButton"] {


  color: #9b8a5c !important;


  display: flex !important;


  visibility: visible !important;


  opacity: 1 !important;


}





.main-title{


  background:linear-gradient(135deg,#0c0b07 0%,#282828 100%);


  color:#9b8a5c;padding:24px 28px;border-radius:12px;margin-bottom:20px;


  display:flex;flex-direction:column;align-items:center;text-align:center;gap:16px;


  box-shadow:0 4px 20px rgba(0,0,0,.15);


  border-bottom:3px solid #9b8a5c;


}


.main-title h1{margin:0;font-size:2.1rem;font-weight:800;color:#9b8a5c;letter-spacing:0.02em;}


.main-title p{margin:0;font-size:.9rem;opacity:.9;color:#e5dccb;margin-top:6px;}





.kpi-card{background:#1c1c1c;border-radius:12px;padding:16px 20px;


  box-shadow:0 4px 15px rgba(0,0,0,0.3);border-left:5px solid;


  transition:transform .2s,box-shadow .2s;text-align:center;}


.kpi-card:hover{transform:translateY(-2px);box-shadow:0 6px 20px rgba(0,0,0,0.4);}


.kpi-card .label{font-size:.7rem;font-weight:700;text-transform:uppercase;


  letter-spacing:.06em;color:#a0a0a0!important;margin-bottom:6px;}


.kpi-card .value{font-size:1.9rem;font-weight:800;line-height:1;}


.kpi-card .sub{font-size:.72rem;color:#777777!important;margin-top:4px;}


.kpi-total    {border-color:#9b8a5c;} .kpi-total    .value{color:#e5dccb!important;}


.kpi-ca       {border-color:#9b8a5c;} .kpi-ca       .value{color:#e5dccb!important;}


.kpi-np       {border-color:#8c6e42;} .kpi-np       .value{color:#9b8a5c!important;}


.kpi-aberta   {border-color:#FF4444;} .kpi-aberta   .value{color:#ff6b6b!important;}


.kpi-parc     {border-color:#FF8C00;} .kpi-parc     .value{color:#ff9f43!important;}


.kpi-hom      {border-color:#FFD966;} .kpi-hom      .value{color:#ffd257!important;}


.kpi-enc      {border-color:#70AD47;} .kpi-enc      .value{color:#7bed9f!important;}
.kpi-active-total  { background: #282828 !important; box-shadow: 0 0 15px rgba(155, 138, 92, 0.45) !important; border: 1.5px solid #9b8a5c !important; border-left: 5px solid #9b8a5c !important; }
.kpi-active-ca     { background: #282828 !important; box-shadow: 0 0 15px rgba(155, 138, 92, 0.45) !important; border: 1.5px solid #9b8a5c !important; border-left: 5px solid #9b8a5c !important; }
.kpi-active-np     { background: #282828 !important; box-shadow: 0 0 15px rgba(140, 110, 66, 0.45) !important; border: 1.5px solid #8c6e42 !important; border-left: 5px solid #8c6e42 !important; }
.kpi-active-enc    { background: #282828 !important; box-shadow: 0 0 15px rgba(112, 173, 71, 0.45) !important; border: 1.5px solid #70AD47 !important; border-left: 5px solid #70AD47 !important; }
.kpi-active-aberta { background: #282828 !important; box-shadow: 0 0 15px rgba(255, 68, 68, 0.45) !important; border: 1.5px solid #FF4444 !important; border-left: 5px solid #FF4444 !important; }
.kpi-active-parc   { background: #282828 !important; box-shadow: 0 0 15px rgba(255, 140, 0, 0.45) !important; border: 1.5px solid #FF8C00 !important; border-left: 5px solid #FF8C00 !important; }
.kpi-active-hom    { background: #282828 !important; box-shadow: 0 0 15px rgba(255, 217, 102, 0.45) !important; border: 1.5px solid #FFD966 !important; border-left: 5px solid #FFD966 !important; }

/* Glassmorphic Crystal Style Button */
button[aria-label="👁️ Mostrar Encerradas"],
button[aria-label="🙈 Ocultar Encerradas"] {
    background: rgba(128, 128, 128, 0.06) !important;
    backdrop-filter: blur(8px) !important;
    -webkit-backdrop-filter: blur(8px) !important;
    border: 1px solid rgba(128, 128, 128, 0.2) !important;
    border-radius: 30px !important;
    color: var(--text-color) !important;
    font-weight: 600 !important;
    box-shadow: 0 4px 15px rgba(0, 0, 0, 0.15), inset 0 1px 2px rgba(255, 255, 255, 0.1) !important;
    text-shadow: 0 1px 2px rgba(0, 0, 0, 0.2) !important;
    transition: all 0.3s ease !important;
}

button[aria-label="👁️ Mostrar Encerradas"]:hover,
button[aria-label="🙈 Ocultar Encerradas"]:hover {
    background: rgba(128, 128, 128, 0.12) !important;
    border-color: rgba(128, 128, 128, 0.3) !important;
    box-shadow: 0 6px 20px rgba(0, 0, 0, 0.2), inset 0 1px 3px rgba(255, 255, 255, 0.2) !important;
    transform: translateY(-1px) !important;
}

button[aria-label="👁️ Mostrar Encerradas"]:active,
button[aria-label="🙈 Ocultar Encerradas"]:active {
    background: rgba(128, 128, 128, 0.02) !important;
    transform: translateY(0px) !important;
}

/* Crystal Liquid styles for interactive legend buttons */
button[aria-label="🟢 Encerrada"],
button[aria-label="🔴 Aberta"],
button[aria-label="🟠 Parcialmente Encerrada"],
button[aria-label="🟡 Homologação"],
button[aria-label="⚪ Encerrada"],
button[aria-label="⚪ Aberta"],
button[aria-label="⚪ Parcialmente Encerrada"],
button[aria-label="⚪ Homologação"] {
    backdrop-filter: blur(8px) !important;
    -webkit-backdrop-filter: blur(8px) !important;
    border-radius: 30px !important;
    color: var(--text-color) !important;
    font-weight: 600 !important;
    text-shadow: 0 1px 2px rgba(0, 0, 0, 0.2) !important;
    transition: all 0.3s ease !important;
    box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15) !important;
}

/* Hover effects */
button[aria-label="🟢 Encerrada"]:hover,
button[aria-label="🔴 Aberta"]:hover,
button[aria-label="🟠 Parcialmente Encerrada"]:hover,
button[aria-label="🟡 Homologação"]:hover,
button[aria-label="⚪ Encerrada"]:hover,
button[aria-label="⚪ Aberta"]:hover,
button[aria-label="⚪ Parcialmente Encerrada"]:hover,
button[aria-label="⚪ Homologação"]:hover {
    transform: translateY(-1px) !important;
}

/* Colored glowing styles for active buttons */
button[aria-label="🟢 Encerrada"] {
    background: rgba(112, 173, 71, 0.15) !important;
    border: 1px solid #70AD47 !important;
    box-shadow: 0 0 10px rgba(112, 173, 71, 0.3) !important;
}
button[aria-label="🔴 Aberta"] {
    background: rgba(255, 68, 68, 0.15) !important;
    border: 1px solid #FF4444 !important;
    box-shadow: 0 0 10px rgba(255, 68, 68, 0.3) !important;
}
button[aria-label="🟠 Parcialmente Encerrada"] {
    background: rgba(255, 140, 0, 0.15) !important;
    border: 1px solid #FF8C00 !important;
    box-shadow: 0 0 10px rgba(255, 140, 0, 0.3) !important;
}
button[aria-label="🟡 Homologação"] {
    background: rgba(255, 217, 102, 0.15) !important;
    border: 1px solid #FFD966 !important;
    box-shadow: 0 0 10px rgba(255, 217, 102, 0.3) !important;
}

/* Muted look for inactive buttons */
button[aria-label="⚪ Encerrada"],
button[aria-label="⚪ Aberta"],
button[aria-label="⚪ Parcialmente Encerrada"],
button[aria-label="⚪ Homologação"] {
    background: rgba(128, 128, 128, 0.03) !important;
    border: 1px solid rgba(128, 128, 128, 0.15) !important;
    opacity: 0.55 !important;
}

/* Glassmorphic Crystal Style for Page Navigation / Horizontal Tab Buttons */
div.element-container:has(.main-nav-marker) + div[data-testid="stHorizontalBlock"] button,
div.element-container:has(.main-nav-marker) + div.element-container button,
div.element-container:has(.main-nav-marker) + div button {
    background: rgba(128, 128, 128, 0.06) !important;
    backdrop-filter: blur(8px) !important;
    -webkit-backdrop-filter: blur(8px) !important;
    border: 1px solid rgba(128, 128, 128, 0.2) !important;
    border-radius: 8px !important;
    color: var(--text-color) !important;
    font-weight: 600 !important;
    padding: 6px 4px !important;
    transition: all 0.25s ease !important;
    box-shadow: 0 2px 8px rgba(0, 0, 0, 0.08) !important;
    
    /* Fixed unified dimensions and vertical centering */
    min-height: 85px !important;
    height: 85px !important;
    display: flex !important;
    flex-direction: column !important;
    justify-content: center !important;
    align-items: center !important;
}

/* Force inner text tags to wrap on pre-line and have uniform small font size */
div.element-container:has(.main-nav-marker) + div[data-testid="stHorizontalBlock"] button *,
div.element-container:has(.main-nav-marker) + div.element-container button *,
div.element-container:has(.main-nav-marker) + div button * {
    color: var(--text-color) !important;
    white-space: pre-line !important;
    text-align: center !important;
    font-size: 0.8rem !important;
    line-height: 1.25 !important;
}

/* Highlight the icon (first line / emoji) by making it significantly larger */
div.element-container:has(.main-nav-marker) + div[data-testid="stHorizontalBlock"] button p::first-line,
div.element-container:has(.main-nav-marker) + div.element-container button p::first-line,
div.element-container:has(.main-nav-marker) + div button p::first-line,
div.element-container:has(.main-nav-marker) + div[data-testid="stHorizontalBlock"] button span::first-line,
div.element-container:has(.main-nav-marker) + div.element-container button span::first-line,
div.element-container:has(.main-nav-marker) + div button span::first-line {
    font-size: 1.65rem !important;
    line-height: 1.45 !important;
    font-weight: normal !important;
}

div.element-container:has(.main-nav-marker) + div[data-testid="stHorizontalBlock"] button:hover,
div.element-container:has(.main-nav-marker) + div.element-container button:hover,
div.element-container:has(.main-nav-marker) + div button:hover {
    background: rgba(128, 128, 128, 0.12) !important;
    border-color: rgba(128, 128, 128, 0.3) !important;
    transform: translateY(-1px) !important;
}

div.element-container:has(.main-nav-marker) + div[data-testid="stHorizontalBlock"] button:hover *,
div.element-container:has(.main-nav-marker) + div.element-container button:hover *,
div.element-container:has(.main-nav-marker) + div button:hover * {
    color: var(--text-color) !important;
}

/* Make active page buttons glow gold */
div.element-container:has(.main-nav-marker) + div[data-testid="stHorizontalBlock"] button[kind="primary"],
div.element-container:has(.main-nav-marker) + div.element-container button[kind="primary"],
div.element-container:has(.main-nav-marker) + div button[kind="primary"] {
    background: rgba(155, 138, 92, 0.2) !important;
    border: 1.5px solid #9b8a5c !important;
    box-shadow: 0 0 12px rgba(155, 138, 92, 0.35) !important;
    color: var(--text-color) !important;
    font-weight: 700 !important;
}

div.element-container:has(.main-nav-marker) + div[data-testid="stHorizontalBlock"] button[kind="primary"] *,
div.element-container:has(.main-nav-marker) + div.element-container button[kind="primary"] *,
div.element-container:has(.main-nav-marker) + div button[kind="primary"] * {
    color: var(--text-color) !important;
}








.stTabs [data-baseweb="tab-list"]{background:#1c1c1c;border-radius:10px;padding:6px;


  box-shadow:0 2px 8px rgba(0,0,0,.3);gap:4px;}


.stTabs [data-baseweb="tab"]{border-radius:8px;font-weight:600;padding:8px 20px;font-size:.85rem;}


.stTabs [aria-selected="true"]{background:#9b8a5c!important;color:#000!important;}





.section-hdr{background:#9b8a5c;color:#000;padding:10px 16px;border-radius:8px;


  font-weight:600;font-size:.9rem;margin:16px 0 8px 0;}


.section-hdr-hom{background:#8c6e42;color:#fff;padding:10px 16px;border-radius:8px;


  font-weight:600;font-size:.9rem;margin:24px 0 8px 0;}


.info-box{background:#1a1a1a;border:1px solid #9b8a5c;border-radius:8px;


  padding:12px 16px;font-size:.85rem;color:#e5dccb;margin-bottom:12px;}


.warn-box{background:#2a1b00;border:1px solid #ffd257;border-radius:8px;


  padding:12px 16px;font-size:.85rem;color:#ffd257;margin-bottom:12px;}


div[data-testid="metric-container"]{background:#1c1c1c;border-radius:10px;


  padding:12px;box-shadow:0 2px 8px rgba(0,0,0,.3);}


/* Botões Gerais - Efeito Cristal Líquido (Glassmorphism Premium) */
/* Botões Gerais - Efeito Cristal Líquido (Glassmorphism Premium) */
.stButton button {
  background: rgba(128, 128, 128, 0.06) !important;
  color: var(--text-color) !important;
  border: 1px solid rgba(128, 128, 128, 0.2) !important;
  border-top: 1px solid rgba(255, 255, 255, 0.3) !important;
  border-radius: 20px !important;
  padding: 8px 24px !important;
  font-weight: 600 !important;
  font-size: 0.9rem !important;
  box-shadow: inset 0 1px 1px rgba(255, 255, 255, 0.1), 0 4px 12px rgba(0, 0, 0, 0.1) !important;
  backdrop-filter: blur(10px) !important;
  -webkit-backdrop-filter: blur(10px) !important;
  transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
  letter-spacing: 0.02em !important;
}

.stButton button * {
  color: var(--text-color) !important;
}

.stButton button:hover {
  background: rgba(128, 128, 128, 0.12) !important;
  color: var(--text-color) !important;
  border-color: rgba(128, 128, 128, 0.3) !important;
  box-shadow: inset 0 1px 2px rgba(255, 255, 255, 0.2), 0 0 15px rgba(155, 138, 92, 0.2) !important;
  transform: translateY(-2px) !important;
}

.stButton button:hover * {
  color: var(--text-color) !important;
}

.stButton button:active {
  transform: translateY(1px) !important;
  box-shadow: inset 0 1px 3px rgba(0, 0, 0, 0.6) !important;
}

/* Botões Primários - Efeito Cristal Ouro (Ativo) */
.stButton button[data-testid="baseButton-primary"] {
  background: linear-gradient(135deg, rgba(155, 138, 92, 0.25) 0%, rgba(0, 0, 0, 0.6) 100%) !important;
  color: var(--text-color) !important;
  border: 1px solid rgba(155, 138, 92, 0.4) !important;
  border-top: 1px solid rgba(255, 255, 255, 0.4) !important;
  box-shadow: inset 0 1px 2px rgba(255, 255, 255, 0.25), 0 4px 15px rgba(155, 138, 92, 0.25) !important;
}

.stButton button[data-testid="baseButton-primary"] * {
  color: var(--text-color) !important;
}

.stButton button[data-testid="baseButton-primary"]:hover {
  background: linear-gradient(135deg, rgba(155, 138, 92, 0.45) 0%, rgba(255, 255, 255, 0.1) 100%) !important;
  border-color: rgba(155, 138, 92, 0.7) !important;
  box-shadow: inset 0 1px 3px rgba(255, 255, 255, 0.4), 0 0 25px rgba(155, 138, 92, 0.6) !important;
  color: var(--text-color) !important;
}

.stButton button[data-testid="baseButton-primary"]:hover * {
  color: var(--text-color) !important;
}

/* Sidebar Botões Secundários */
div[data-testid="stSidebar"] button[data-testid="baseButton-secondary"] {
  background: rgba(128, 128, 128, 0.05) !important;
  color: var(--text-color) !important;
  border: 1px solid rgba(128, 128, 128, 0.2) !important;
  border-top: 1px solid rgba(255, 255, 255, 0.25) !important;
  border-radius: 20px !important;
  box-shadow: inset 0 1px 1px rgba(255, 255, 255, 0.05), 0 4px 10px rgba(0,0,0,0.2) !important;
  font-weight: 500 !important;
  transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
  backdrop-filter: blur(8px) !important;
  -webkit-backdrop-filter: blur(8px) !important;
}

div[data-testid="stSidebar"] button[data-testid="baseButton-secondary"] * {
  color: var(--text-color) !important;
}

div[data-testid="stSidebar"] button[data-testid="baseButton-secondary"]:hover {
  background: rgba(155, 138, 92, 0.15) !important;
  color: var(--text-color) !important;
  border-color: rgba(155, 138, 92, 0.4) !important;
  border-top-color: rgba(155, 138, 92, 0.6) !important;
  box-shadow: inset 0 1px 2px rgba(255, 255, 255, 0.15), 0 0 15px rgba(155, 138, 92, 0.3) !important;
  transform: translateY(-2px) !important;
}

div[data-testid="stSidebar"] button[data-testid="baseButton-secondary"]:hover * {
  color: var(--text-color) !important;
}

/* Sidebar Botões Primários */
div[data-testid="stSidebar"] button[data-testid="baseButton-primary"] {
  background: linear-gradient(135deg, rgba(155, 138, 92, 0.3) 0%, rgba(0, 0, 0, 0.7) 100%) !important;
  color: #ffffff !important;
  border: 1px solid rgba(155, 138, 92, 0.5) !important;
  border-top: 1px solid rgba(255, 255, 255, 0.4) !important;
  border-radius: 20px !important;
  box-shadow: inset 0 1px 3px rgba(255,255,255,0.3), 0 4px 15px rgba(188,163,116,0.3) !important;
  font-weight: 700 !important;
  text-shadow: 0 1px 2px rgba(0,0,0,0.8) !important;
  transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
}

div[data-testid="stSidebar"] button[data-testid="baseButton-primary"]:hover {
  background: linear-gradient(135deg, rgba(255, 255, 255, 0.25) 0%, rgba(155, 138, 92, 0.45) 100%) !important;
  color: #ffffff !important;
  border-color: rgba(255, 255, 255, 0.7) !important;
  box-shadow: inset 0 1px 3px rgba(255,255,255,0.4), 0 0 25px rgba(188,163,116,0.5) !important;
  transform: translateY(-2px) !important;
}


/* Glassmorphic Crystal Style for Report Scope Selection Buttons */
div.element-container:has(.report-scope-marker) + div[data-testid="stHorizontalBlock"] button,
div.element-container:has(.report-scope-marker) + div.element-container button,
div.element-container:has(.report-scope-marker) + div button {
    background: rgba(128, 128, 128, 0.05) !important;
    backdrop-filter: blur(8px) !important;
    -webkit-backdrop-filter: blur(8px) !important;
    border: 1px solid rgba(128, 128, 128, 0.18) !important;
    border-radius: 8px !important;
    color: var(--text-color) !important;
    font-weight: 600 !important;
    padding: 6px 4px !important;
    transition: all 0.25s ease !important;
    box-shadow: 0 2px 8px rgba(0, 0, 0, 0.08) !important;
    
    /* Fixed unified dimensions and vertical centering */
    min-height: 95px !important;
    height: 95px !important;
    display: flex !important;
    flex-direction: column !important;
    justify-content: center !important;
    align-items: center !important;
}

div.element-container:has(.report-scope-marker) + div[data-testid="stHorizontalBlock"] button *,
div.element-container:has(.report-scope-marker) + div.element-container button *,
div.element-container:has(.report-scope-marker) + div button * {
    color: var(--text-color) !important;
    white-space: pre-line !important;
    text-align: center !important;
    font-size: 0.85rem !important;
    line-height: 1.3 !important;
    font-weight: 600 !important;
}

/* Highlight the emoji icon as larger */
div.element-container:has(.report-scope-marker) + div[data-testid="stHorizontalBlock"] button p::first-line,
div.element-container:has(.report-scope-marker) + div.element-container button p::first-line,
div.element-container:has(.report-scope-marker) + div button p::first-line,
div.element-container:has(.report-scope-marker) + div[data-testid="stHorizontalBlock"] button span::first-line,
div.element-container:has(.report-scope-marker) + div.element-container button span::first-line,
div.element-container:has(.report-scope-marker) + div button span::first-line {
    font-size: 1.75rem !important;
    line-height: 1.45 !important;
    font-weight: normal !important;
}

div.element-container:has(.report-scope-marker) + div[data-testid="stHorizontalBlock"] button:hover,
div.element-container:has(.report-scope-marker) + div.element-container button:hover,
div.element-container:has(.report-scope-marker) + div button:hover {
    background: rgba(128, 128, 128, 0.12) !important;
    border-color: rgba(128, 128, 128, 0.3) !important;
    transform: translateY(-1px) !important;
}

div.element-container:has(.report-scope-marker) + div[data-testid="stHorizontalBlock"] button:hover *,
div.element-container:has(.report-scope-marker) + div.element-container button:hover *,
div.element-container:has(.report-scope-marker) + div button:hover * {
    color: var(--text-color) !important;
}

/* Selected active button (primary) */
div.element-container:has(.report-scope-marker) + div[data-testid="stHorizontalBlock"] button[kind="primary"],
div.element-container:has(.report-scope-marker) + div.element-container button[kind="primary"],
div.element-container:has(.report-scope-marker) + div button[kind="primary"] {
    background: rgba(155, 138, 92, 0.22) !important;
    border: 1.5px solid #9b8a5c !important;
    box-shadow: 0 0 14px rgba(155, 138, 92, 0.45) !important;
    color: var(--text-color) !important;
}

div.element-container:has(.report-scope-marker) + div[data-testid="stHorizontalBlock"] button[kind="primary"] *,
div.element-container:has(.report-scope-marker) + div.element-container button[kind="primary"] *,
div.element-container:has(.report-scope-marker) + div button[kind="primary"] * {
    color: var(--text-color) !important;
}

/* Glassmorphic Crystal Style for Excel Scope Selection Buttons */
div.element-container:has(.excel-scope-marker) + div[data-testid="stHorizontalBlock"] button,
div.element-container:has(.excel-scope-marker) + div.element-container button,
div.element-container:has(.excel-scope-marker) + div button {
    background: rgba(128, 128, 128, 0.05) !important;
    backdrop-filter: blur(8px) !important;
    -webkit-backdrop-filter: blur(8px) !important;
    border: 1px solid rgba(128, 128, 128, 0.18) !important;
    border-radius: 8px !important;
    color: var(--text-color) !important;
    font-weight: 600 !important;
    padding: 6px 4px !important;
    transition: all 0.25s ease !important;
    box-shadow: 0 2px 8px rgba(0, 0, 0, 0.08) !important;
    
    /* Fixed unified dimensions and vertical centering */
    min-height: 95px !important;
    height: 95px !important;
    display: flex !important;
    flex-direction: column !important;
    justify-content: center !important;
    align-items: center !important;
}

div.element-container:has(.excel-scope-marker) + div[data-testid="stHorizontalBlock"] button *,
div.element-container:has(.excel-scope-marker) + div.element-container button *,
div.element-container:has(.excel-scope-marker) + div button * {
    color: var(--text-color) !important;
    white-space: pre-line !important;
    text-align: center !important;
    font-size: 0.85rem !important;
    line-height: 1.3 !important;
    font-weight: 600 !important;
}

/* Highlight the emoji icon as larger */
div.element-container:has(.excel-scope-marker) + div[data-testid="stHorizontalBlock"] button p::first-line,
div.element-container:has(.excel-scope-marker) + div.element-container button p::first-line,
div.element-container:has(.excel-scope-marker) + div button p::first-line,
div.element-container:has(.excel-scope-marker) + div[data-testid="stHorizontalBlock"] button span::first-line,
div.element-container:has(.excel-scope-marker) + div.element-container button span::first-line,
div.element-container:has(.excel-scope-marker) + div button span::first-line {
    font-size: 1.75rem !important;
    line-height: 1.45 !important;
    font-weight: normal !important;
}

div.element-container:has(.excel-scope-marker) + div[data-testid="stHorizontalBlock"] button:hover,
div.element-container:has(.excel-scope-marker) + div.element-container button:hover,
div.element-container:has(.excel-scope-marker) + div button:hover {
    background: rgba(128, 128, 128, 0.12) !important;
    border-color: rgba(128, 128, 128, 0.3) !important;
    transform: translateY(-1px) !important;
}

div.element-container:has(.excel-scope-marker) + div[data-testid="stHorizontalBlock"] button:hover *,
div.element-container:has(.excel-scope-marker) + div.element-container button:hover *,
div.element-container:has(.excel-scope-marker) + div button:hover * {
    color: var(--text-color) !important;
}

/* Selected active button (primary) */
div.element-container:has(.excel-scope-marker) + div[data-testid="stHorizontalBlock"] button[kind="primary"],
div.element-container:has(.excel-scope-marker) + div.element-container button[kind="primary"],
div.element-container:has(.excel-scope-marker) + div button[kind="primary"] {
    background: rgba(155, 138, 92, 0.22) !important;
    border: 1.5px solid #9b8a5c !important;
    box-shadow: 0 0 14px rgba(155, 138, 92, 0.45) !important;
    color: var(--text-color) !important;
}

div.element-container:has(.excel-scope-marker) + div[data-testid="stHorizontalBlock"] button[kind="primary"] *,
div.element-container:has(.excel-scope-marker) + div.element-container button[kind="primary"] *,
div.element-container:has(.excel-scope-marker) + div button[kind="primary"] * {
    color: var(--text-color) !important;
}

</style>


""", unsafe_allow_html=True)





# ─────────────────────── CONSTANTES ───────────────────────────────────────────


THIS_DIR  = Path(__file__).parent


DADOS_DIR = THIS_DIR / "dados"


DADOS_DIR.mkdir(exist_ok=True)


CONFIG_FILE = THIS_DIR / "config_aadp.json"





SITUACOES_ALVO = {


    "ATIV. DIRECAO GERAL","ATIV. FIM DESTACADO","ATIV. FIM NA SEDE",


    "ATIV. MEIO","ATIVIDADE MEIO","DISP MED DEFINITIVA",


    "GESTANTE/LAC/ADOTANT","QUADRO ESPECIALISTA",


}


CONCEITO_FAIXA = {


    "nivel superior de desempenho":       (9.00, 10.00),


    "nivel alto de desempenho":           (7.00,  8.99),


    "nivel intermediario de desempenho":  (6.00,  6.99),


    "nivel baixo de desempenho":          (3.00,  5.99),


    "nivel inferior de desempenho":       (0.00,  2.99),


}


STATUS_COLORS = {


    "Encerrada":             "#70AD47",


    "Homologação":           "#FFD966",


    "Parcialmente Encerrada":"#FF8C00",


    "Aberta":                "#FF4444",


}


SIT_COLORS = {"Comissão Atual":"#4472C4","Nota Provisória":"#FFC000"}


# Ordem para empilhamento: Encerradas embaixo, pendentes em cima


STACK_ORDER  = ["Encerrada","Aberta","Parcialmente Encerrada","Homologação"]





# ─────────────────────── CONFIGURAÇÃO ─────────────────────────────────────────


def load_config():


    cfg = {"db_path": str(DADOS_DIR)}


    if CONFIG_FILE.exists():


        try:


            cfg.update(json.loads(CONFIG_FILE.read_text(encoding="utf-8")))


        except Exception:


            pass


    # Carrega do st.secrets do Streamlit para evitar perda de IDs/links após reinicializações


    for key in ["drive_av_id", "drive_si_id", "drive_geral_id", "drive_master_xlsx_id", "sheet_api_url", "fonte_dados", "db_path"]:


        try:


            if key in st.secrets:


                cfg[key] = st.secrets[key]


        except Exception:


            pass


    return cfg





def save_config(cfg):


    CONFIG_FILE.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")





DB_FILE = str(Path(__file__).parent / "aadp_secure.db")





def init_db():


    conn = sqlite3.connect(DB_FILE)


    c = conn.cursor()


    c.execute("""


        CREATE TABLE IF NOT EXISTS users (


            pm TEXT PRIMARY KEY,


            name TEXT,


            rank TEXT,


            rpm TEXT,


            unit TEXT,


            function TEXT,


            role TEXT,


            status TEXT,


            password TEXT,


            created_at TEXT


        )


    """)


    c.execute("""


        CREATE TABLE IF NOT EXISTS logs (


            id INTEGER PRIMARY KEY AUTOINCREMENT,


            timestamp TEXT,


            pm TEXT,


            action TEXT,


            details TEXT


        )


    """)


    c.execute("SELECT * FROM users WHERE pm = 'ADM'")


    if not c.fetchone():


        adm_pass = hashlib.sha256("arquivosDRH2026".encode()).hexdigest()


        c.execute("""


            INSERT INTO users (pm, name, rank, rpm, unit, function, role, status, password, created_at)


            VALUES ('ADM', 'Administrador Geral', 'Desenvolvedor', 'Geral', 'DRH', 'Administrador', 'ADMINISTRADOR', 'Ativo', ?, ?)


        """, (adm_pass, now_br().strftime("%Y-%m-%d %H:%M:%S")))


    conn.commit()


    conn.close()





def log_action(pm: str, action: str, details: str = ""):


    timestamp = now_br().strftime("%Y-%m-%d %H:%M:%S")


    if check_use_cloud():


        run_sheet_api("add_log", {"log": {"timestamp": timestamp, "pm": pm, "action": action, "details": details}})


    else:


        try:


            conn = sqlite3.connect(DB_FILE)


            c = conn.cursor()


            c.execute("INSERT INTO logs (timestamp, pm, action, details) VALUES (?, ?, ?, ?)",


                      (timestamp, pm, action, details))


            conn.commit()


            conn.close()


        except Exception:


            pass





init_db()





cfg = load_config()





# ─────────────────────── DATABASE WRAPPERS (SQLite / Google Sheets Cloud) ───────


def check_use_cloud():


    url = cfg.get("sheet_api_url", "")


    return bool(url and url.strip().lower().startswith("http"))





def run_sheet_api(action, payload=None):


    url = cfg.get("sheet_api_url", "")


    if not url:


        return None


    import requests


    try:


        body = {"action": action}


        if payload:


            body.update(payload)


        r = requests.post(url, json=body, timeout=10)


        if r.status_code == 200:


            res = r.json()


            if res.get("status") == "success":


                return res.get("data")


            else:


                st.error(f"Erro na Planilha: {res.get('message')}")


    except Exception as e:


        st.error(f"Erro ao conectar com Google Sheets: {e}")


    return None





def refresh_db_cache():


    if check_use_cloud():


        users = run_sheet_api("get_users")


        if users is None:


            users = []


    else:


        try:


            conn = sqlite3.connect(DB_FILE)


            c = conn.cursor()


            c.execute("SELECT pm, name, rank, rpm, unit, function, password, role, status, created_at FROM users")


            rows = c.fetchall()


            conn.close()


            users = []


            for r in rows:


                users.append({


                    "pm": r[0], "name": r[1], "rank": r[2], "rpm": r[3], "unit": r[4],


                    "function": r[5], "password": r[6], "role": r[7], "status": r[8],


                    "created_at": r[9]


                })


        except Exception:


            users = []


    st.session_state.db_users = users


    return users





def get_cached_users():


    if "db_users" not in st.session_state:


        refresh_db_cache()


    return st.session_state.db_users





def db_get_user_for_login(pm, password_hash):


    if check_use_cloud():


        users = run_sheet_api("get_users")


        if users:


            for u in users:


                if str(u["pm"]).strip() == str(pm).strip() and str(u["password"]).strip() == str(password_hash).strip():


                    return (u["name"], u["role"], u["rpm"], u["unit"], u["status"])


        return None


    else:


        conn = sqlite3.connect(DB_FILE)


        c = conn.cursor()


        c.execute("SELECT name, role, rpm, unit, status FROM users WHERE pm = ? AND password = ?", (pm, password_hash))


        res = c.fetchone()


        conn.close()


        return res





def db_register_user(pm, name, rank, rpm, unit, function, password_hash):


    created_at = now_br().strftime("%Y-%m-%d %H:%M:%S")


    success = False


    if check_use_cloud():


        user = {


            "pm": pm, "name": name, "rank": rank, "rpm": rpm, "unit": unit,


            "function": function, "role": "AGUARDANDO", "status": "Pendente",


            "password": password_hash, "created_at": created_at


        }


        res = run_sheet_api("add_user", {"user": user})


        success = res is not None


    else:


        try:


            conn = sqlite3.connect(DB_FILE)


            c = conn.cursor()


            c.execute("""


                INSERT INTO users (pm, name, rank, rpm, unit, function, role, status, password, created_at)


                VALUES (?, ?, ?, ?, ?, ?, 'AGUARDANDO', 'Pendente', ?, ?)


            """, (pm, name, rank, rpm, unit, function, password_hash, created_at))


            conn.commit()


            conn.close()


            success = True


        except Exception:


            success = False


    if success:


        refresh_db_cache()


    return success





def db_get_all_pms():


    users = get_cached_users()


    return [str(u["pm"]) for u in users]





def db_update_user_info(pm, name, rank, rpm, unit, sector):


    if check_use_cloud():


        updates = {"name": name, "rank": rank, "rpm": rpm, "unit": unit, "function": sector}


        run_sheet_api("update_user", {"pm": pm, "updates": updates})


    else:


        conn = sqlite3.connect(DB_FILE)


        c = conn.cursor()


        c.execute("""


            UPDATE users


            SET name = ?, rank = ?, rpm = ?, unit = ?, function = ?


            WHERE pm = ?


        """, (name, rank, rpm, unit, sector, pm))


        conn.commit()


        conn.close()


    refresh_db_cache()





def db_get_user_password(pm):


    users = get_cached_users()


    for u in users:


        if str(u["pm"]).strip() == str(pm).strip():


            return u["password"]


    return None





def db_update_password(pm, password_hash):


    if check_use_cloud():


        run_sheet_api("update_user", {"pm": pm, "updates": {"password": password_hash}})


    else:


        conn = sqlite3.connect(DB_FILE)


        c = conn.cursor()


        c.execute("UPDATE users SET password = ? WHERE pm = ?", (password_hash, pm))


        conn.commit()


        conn.close()


    refresh_db_cache()





def db_get_simulator_users():


    users = get_cached_users()


    sim_users = []


    for u in users:


        if u["status"] == "Ativo" and str(u["pm"]) != "ADM":


            sim_users.append((u["pm"], u["name"], u["rank"], u["role"], u["rpm"], u["unit"]))


    sim_users.sort(key=lambda x: x[1])


    return sim_users





def db_get_pending_users():


    users = get_cached_users()


    pend_users = []


    for u in users:


        if u["status"] == "Pendente":


            pend_users.append((u["pm"], u["name"], u["rank"], u["rpm"], u["unit"], u["function"], u["created_at"]))


    return pend_users





def db_get_active_users():
    users = get_cached_users()
    active = []
    for u in users:
        if u["status"] == "Ativo" and str(u["pm"]) != "ADM":
            active.append((
                u["pm"],
                u["rank"],
                u["name"],
                u.get("rpm", ""),
                u.get("unit", ""),
                u.get("function", ""),  # Setor = NOME UNIDADE (col J SIGEF)
                u["role"],
                u["created_at"]
            ))
    return active





def db_approve_user(pm, role, rpm):


    if check_use_cloud():


        run_sheet_api("update_user", {"pm": pm, "updates": {"status": "Ativo", "role": role, "rpm": rpm}})


    else:


        conn = sqlite3.connect(DB_FILE)


        c = conn.cursor()


        c.execute("UPDATE users SET status = 'Ativo', role = ?, rpm = ? WHERE pm = ?", (role, rpm, pm))


        conn.commit()


        conn.close()


    refresh_db_cache()





def db_reject_user(pm):


    if check_use_cloud():


        run_sheet_api("update_user", {"pm": pm, "updates": {"status": "Recusado"}})


    else:


        conn = sqlite3.connect(DB_FILE)


        c = conn.cursor()


        c.execute("UPDATE users SET status = 'Recusado' WHERE pm = ?", (pm,))


        conn.commit()


        conn.close()


    refresh_db_cache()





def db_update_user_role_rpm(pm, role, rpm):


    if check_use_cloud():


        run_sheet_api("update_user", {"pm": pm, "updates": {"role": role, "rpm": rpm}})


    else:


        conn = sqlite3.connect(DB_FILE)


        c = conn.cursor()


        c.execute("UPDATE users SET role = ?, rpm = ? WHERE pm = ?", (role, rpm, pm))


        conn.commit()


        conn.close()


    refresh_db_cache()





def db_revoke_user(pm):


    if check_use_cloud():


        run_sheet_api("update_user", {"pm": pm, "updates": {"status": "Bloqueado"}})


    else:


        conn = sqlite3.connect(DB_FILE)


        c = conn.cursor()


        c.execute("UPDATE users SET status = 'Bloqueado' WHERE pm = ?", (pm,))


        conn.commit()


        conn.close()


    refresh_db_cache()





def db_get_users_df():


    users = get_cached_users()


    rows = []


    for u in users:


        if str(u["pm"]) != "ADM":


            rows.append({


                "pm": u["pm"], "rank": u["rank"], "name": u["name"],


                "role": u["role"], "rpm": u["rpm"], "unit": u["unit"],


                "status": u["status"]


            })


    return pd.DataFrame(rows)





def refresh_logs_cache():


    if check_use_cloud():


        logs = run_sheet_api("get_logs")


        if logs is None:


            logs = []


    else:


        try:


            conn = sqlite3.connect(DB_FILE)


            c = conn.cursor()


            c.execute("SELECT timestamp, pm, action, details FROM logs ORDER BY id DESC")


            rows = c.fetchall()


            conn.close()


            logs = []


            for r in rows:


                logs.append({


                    "timestamp": r[0], "pm": r[1], "action": r[2], "details": r[3]


                })


        except Exception:


            logs = []


    st.session_state.db_logs = logs


    return logs





def get_cached_logs():


    if "db_logs" not in st.session_state:


        refresh_logs_cache()


    return st.session_state.db_logs





def db_get_logs_pms():


    logs = get_cached_logs()


    pms = set(str(l["pm"]) for l in logs)


    return list(pms)





def db_get_user_info(pm):


    users = get_cached_users()


    for u in users:


        if str(u["pm"]).strip() == str(pm).strip():


            return (u["rank"], u["name"])


    return None





def db_get_logs_df(sel_log_user, start_str, end_str):


    logs = get_cached_logs()


    rows = []


    for l in logs:


        t_date = l["timestamp"][:10]


        if start_str <= t_date <= end_str:


            if sel_log_user == "Todos" or str(l["pm"]).strip() == str(sel_log_user).strip():


                rows.append({


                    "timestamp": l["timestamp"],


                    "pm": l["pm"],


                    "action": l["action"],


                    "details": l["details"]


                })


    return pd.DataFrame(rows)





def db_get_pending_count():


    users = get_cached_users()


    return sum(1 for u in users if u["status"] == "Pendente")





def db_check_user_status(pm):


    users = get_cached_users()


    for u in users:


        if str(u["pm"]).strip() == str(pm).strip():


            return u["status"]


    return None





def db_re_request_access(pm, name, rank, rpm, unit, sector, password_hash):


    created_at = now_br().strftime("%Y-%m-%d %H:%M:%S")


    success = False


    if check_use_cloud():


        updates = {


            "name": name, "rank": rank, "rpm": rpm, "unit": unit, "function": sector,


            "role": "PENDENTE", "status": "Pendente", "password": password_hash, "created_at": created_at


        }


        res = run_sheet_api("update_user", {"pm": pm, "updates": updates})


        success = res is not None


    else:


        try:


            conn = sqlite3.connect(DB_FILE)


            c = conn.cursor()


            c.execute("""


                UPDATE users 


                SET name = ?, rank = ?, rpm = ?, unit = ?, function = ?, role = 'PENDENTE', status = 'Pendente', password = ?, created_at = ?


                WHERE pm = ?


            """, (name, rank, rpm, unit, sector, password_hash, created_at, pm))


            conn.commit()


            conn.close()


            success = True


        except Exception:


            success = False


    if success:


        refresh_db_cache()


    return success





def db_create_new_request(pm, name, rank, rpm, unit, sector, password_hash):


    created_at = now_br().strftime("%Y-%m-%d %H:%M:%S")


    success = False


    if check_use_cloud():


        user = {


            "pm": pm, "name": name, "rank": rank, "rpm": rpm, "unit": unit, "function": sector,


            "role": "PENDENTE", "status": "Pendente", "password": password_hash, "created_at": created_at


        }


        res = run_sheet_api("add_user", {"user": user})


        success = res is not None


    else:


        try:


            conn = sqlite3.connect(DB_FILE)


            c = conn.cursor()


            c.execute("""


                INSERT INTO users (pm, name, rank, rpm, unit, function, role, status, password, created_at)


                VALUES (?, ?, ?, ?, ?, ?, 'PENDENTE', 'Pendente', ?, ?)


            """, (pm, name, rank, rpm, unit, sector, password_hash, created_at))


            conn.commit()


            conn.close()


            success = True


        except Exception:


            success = False


    if success:


        refresh_db_cache()


    return success





# ─────────────────────── LÓGICA DADOS ─────────────────────────────────────────


def normaliza(t):


    s = unicodedata.normalize("NFD", t.lower())


    return "".join(c for c in s if unicodedata.category(c) != "Mn")





def is_empty(v):


    if v == 0 or v == 0.0:


        return False


    return not v or str(v).strip() in ("", "-", "nan", "none")





def concordam(j, l):


    if is_empty(j) or is_empty(l): return None


    try: nota = float(str(l).replace(",","."))


    except: return None


    faixa = CONCEITO_FAIXA.get(normaliza(j.strip()))


    if faixa is None: return None


    return faixa[0] <= nota <= faixa[1]





def matches_rpm(reg_rpm, csv_rpm):


    s_reg = str(reg_rpm).strip()


    s_csv = str(csv_rpm).strip()


    if s_reg.lower() == "gestor":


        return True


    if s_reg.lower() == s_csv.lower():


        return True


    


    # Extrair dígitos de forma ultra rápida


    reg_digits = "".join(c for c in s_reg if c.isdigit())


    csv_digits = "".join(c for c in s_csv if c.isdigit())


    


    if reg_digits and csv_digits:


        return int(reg_digits) == int(csv_digits)


    return False





def find_sigef_user(pm_number: str) -> dict:


    """Busca os dados do militar no SIGEF.csv pelo Nº PM (6 dígitos)."""


    pm_clean = pm_number.strip().lstrip("0")


    if not pm_clean:


        return None


    try:


        import os, csv


        si_path = "SIGEF.csv"


        # Se não existe localmente, tenta baixar do Google Drive


        if not os.path.exists(si_path):


            sigef_drive_id = "10Ld_4XEz9b4kI_T6TC9W19tQdtBJCz5F"


            try:


                _baixar_drive(sigef_drive_id, si_path)


            except Exception as e:


                # Tenta olhar na subpasta dados/


                si_path = os.path.join("dados", "SIGEF.csv")


                if not os.path.exists(si_path):


                    st.error(f"Erro ao baixar base SIGEF do Google Drive: {e}")


                    return None


        with open(si_path, encoding="cp1252", errors="replace") as f:


            reader = csv.reader(f, delimiter=";")


            header = next(reader)


            for row in reader:


                if len(row) > 24:


                    curr_pm = row[0].strip().lstrip("0")


                    if curr_pm == pm_clean:


                        return {


                            "pm": row[0].strip(),


                            "rank": row[2].strip().title(),


                            "name": row[3].strip().title(),


                            "rpm": row[5].strip(),      # UDI/UDG (NOME RPM)


                            "unit": row[7].strip(),     # Unidade Principal (NOME UNIDADE PRINCIPAL)


                            "sector": row[9].strip(),    # Setor (NOME UNIDADE)


                            "birthdate": row[16].strip(), # DATA NASCIMENTO (Q)


                            "cpf": row[24].strip()       # NUMERO CPF (Y)


                        }


    except Exception as e:


        st.error(f"Erro ao ler banco SIGEF: {e}")


    return None





def sync_users_with_sigef():


    """Sincroniza os dados cadastrais de todos os usuários com o SIGEF.csv de forma otimizada."""


    try:


        import os, csv


        si_path = "SIGEF.csv"


        if not os.path.exists(si_path):


            si_path = os.path.join("dados", "SIGEF.csv")


            if not os.path.exists(si_path):


                # Tenta baixar


                sigef_drive_id = "10Ld_4XEz9b4kI_T6TC9W19tQdtBJCz5F"


                try:


                    _baixar_drive(sigef_drive_id, "SIGEF.csv")


                    si_path = "SIGEF.csv"


                except Exception:


                    return





        # Carrega todo o SIGEF para um dicionário em memória (uma única leitura de disco)


        sigef_dict = {}


        with open(si_path, encoding="cp1252", errors="replace") as f:


            reader = csv.reader(f, delimiter=";")


            next(reader)  # pula cabeçalho


            for row in reader:


                if len(row) > 24:


                    pm_clean = row[0].strip().lstrip("0")


                    if pm_clean:


                        sigef_dict[pm_clean] = {
                            "name": row[3].strip(),
                            "rank": row[2].strip(),
                            "rpm": row[5].strip(),
                            "unit": row[7].strip(),
                            "sector": row[9].strip()
                        }


        


        pms = db_get_all_pms()


        for pm in pms:


            if pm == "ADM":


                continue


            pm_clean = pm.strip().lstrip("0")


            if pm_clean in sigef_dict:


                info = sigef_dict[pm_clean]


                db_update_user_info(pm, info["name"], info["rank"], info["rpm"], info["unit"], info["sector"])


    except Exception:


        pass





def calc_cert(j, l):


    if is_empty(j) or is_empty(l): return "-"


    c = concordam(j, l)


    return "NÃO" if c is True else ("SIM" if c is False else "-")





def calc_status(j, l, n):


    if is_empty(j):


        return "Aberta"


    if is_empty(l):


        return "Parcialmente Encerrada"


    c = concordam(j, l)


    if c is True:


        return "Encerrada"


    elif c is False:


        return "Encerrada" if not is_empty(n) else "Homologação"


    return "Parcialmente Encerrada"





def rpm_sort_key(name):


    m = re.match(r'^(\d+)\s+RPM', str(name))


    if m: return (0, int(m.group(1)), "")


    return (1, 0, str(name))





def _baixar_drive(file_id: str, destino: str):
    """Baixa um arquivo do Google Drive para destino local.
    Compatível com todas as versões do gdown (com e sem parâmetro fuzzy).
    """
    if not GDOWN_OK:
        raise ImportError("Biblioteca 'gdown' não instalada. Execute: pip install gdown")
    import inspect
    url = f"https://drive.google.com/uc?id={file_id}&export=download"
    
    # gdown >= 4.6 suporta fuzzy; versões mais antigas não suportam
    sig = inspect.signature(gdown.download)
    if "fuzzy" in sig.parameters:
        gdown.download(url, destino, quiet=True, fuzzy=True)
    else:
        gdown.download(url, destino, quiet=True)
        
    if not os.path.exists(destino) or os.path.getsize(destino) == 0:
        raise FileNotFoundError(f"Falha ao baixar arquivo do Drive (ID: {file_id})")
        
    try:
        with open(destino, "rb") as f:
            head = f.read(200)
        if b"<!DOCTYPE" in head.upper() or b"<HTML" in head.upper():
            raise ValueError("O Google Drive retornou uma página HTML em vez do arquivo binário. Isso ocorre se o arquivo não estiver compartilhado como público ('Qualquer pessoa com o link pode ler'), se o ID estiver incorreto, ou se você estiver tentando baixar uma Planilha Google (Google Sheets) em vez de um arquivo Excel (.xlsx) carregado no Drive.")
    except ValueError as ve:
        if os.path.exists(destino):
            os.remove(destino)
        raise ve
    except Exception:
        pass


def _parse_csv(av_f: str, si_f: str) -> pd.DataFrame:


    """Processa os dois CSVs e retorna o DataFrame final."""


    sigef = {}


    with open(si_f, encoding="cp1252", errors="replace") as f:


        for row in csv.reader(f, delimiter=";"):


            if len(row) > 9:


                sigef[row[0].strip().lstrip("0") or "0"] = row[9].strip()


    rows = []


    with open(av_f, encoding="cp1252", errors="replace") as f:


        reader = csv.reader(f, delimiter=";")


        next(reader)


        for row in reader:


            while len(row) < 50: row.append("")  # CSV tem 50 colunas (colégio até Homologador)


            sit = row[7].strip()


            if sit not in SITUACOES_ALVO: continue


            nrpm = row[0].strip(); local = row[5].strip()


            j = row[9].strip(); l = row[11].strip(); n = row[13].strip()


            sc = "Comissão Atual" if local.upper().strip() == sigef.get(nrpm.lstrip("0") or "0","").upper().strip() else "Nota Provisória"


            rows.append({


                "nrPM (Avaliado)":          nrpm,


                "Nome (Avaliado)":           row[1].strip(),


                "Posto/Grad. (Avaliado)":    row[2].strip(),


                "Unidade RPM (Avaliado)":    row[3].strip(),


                "Unidade Principal (Avaliado)": row[4].strip(),


                "Local/Unidade (Avaliado)":  local,


                "Quadro Atual (Avaliado)":   row[6].strip(),


                "Situação Funcional":        sit,


                "Data AV1":                  row[8].strip(),


                "Conceito Geral":            j,


                "Data AV2":                  row[10].strip(),


                "Nota Geral":                l,


                "Certificação Homologador":  calc_cert(j, l),


                "Data HOM":                  row[12].strip(),


                "Nota Homologação":          n,


                "Competência 1":             row[14].strip(),


                "Conceito Comp.1":           row[15].strip(), "Nota Comp.1": row[16].strip(),


                "Competência 2":             row[17].strip(),


                "Conceito Comp.2":           row[18].strip(), "Nota Comp.2": row[19].strip(),


                "Competência 3":             row[20].strip(),


                "Conceito Comp.3":           row[21].strip(), "Nota Comp.3": row[22].strip(),


                "Competência 4":             row[23].strip(),


                "Conceito Comp.4":           row[24].strip(), "Nota Comp.4": row[25].strip(),


                # Avaliador 1


                "nrPM (Av1)":    row[26].strip(), "Nome (Av1)":  row[27].strip(),


                "Posto (Av1)":   row[28].strip(), "RPM (Av1)":   row[29].strip(),


                "Unid. Principal (Av1)": row[30].strip(), "Local (Av1)": row[31].strip(),


                "Quadro (Av1)":  row[32].strip(), "Situação (Av1)": row[33].strip(),


                # Avaliador 2


                "nrPM (Av2)":    row[34].strip(), "Nome (Av2)":  row[35].strip(),


                "Posto (Av2)":   row[36].strip(), "RPM (Av2)":   row[37].strip(),


                "Unid. Principal (Av2)": row[38].strip(), "Local (Av2)": row[39].strip(),


                "Quadro (Av2)":  row[40].strip(), "Situação (Av2)": row[41].strip(),


                # Homologador (colunas 42–49)


                "nrPM (Hom)":    row[42].strip(), "Nome (Hom)":  row[43].strip(),


                "Posto (Hom)":   row[44].strip(), "RPM (Hom)":   row[45].strip(),


                "Unid. Principal (Hom)": row[46].strip(), "Local (Hom)": row[47].strip(),


                "Quadro (Hom)":  row[48].strip(), "Situação (Hom)": row[49].strip(),


                "Situação Comissão": sc,


                "Status Avaliação":  calc_status(j, l, n),


            })


    return pd.DataFrame(rows)





@st.cache_resource(show_spinner="⏳ Carregando e processando dados...")
def load_data(db_path: str, drive_av_id: str = "", drive_si_id: str = ""):
    """Carrega dados de pasta local ou Google Drive e gera o Geral.xlsx automaticamente."""
    if drive_av_id and drive_si_id:
        # ── Modo Google Drive ──────────────────────────────────────────────
        cache_dir = os.path.join(tempfile.gettempdir(), "aadp_drive_cache")
        os.makedirs(cache_dir, exist_ok=True)
        av_f = os.path.join(cache_dir, "avaliacoes.csv")
        si_f = os.path.join(cache_dir, "SIGEF.csv")
        
        if not os.path.exists(av_f) or os.path.getsize(av_f) == 0:
            _baixar_drive(drive_av_id, av_f)
        if not os.path.exists(si_f) or os.path.getsize(si_f) == 0:
            _baixar_drive(drive_si_id, si_f)


    else:


        # ── Modo pasta local ───────────────────────────────────────────────


        av_f = os.path.join(db_path, "avaliacoes.csv")


        si_f = os.path.join(db_path, "SIGEF.csv")


        if not os.path.exists(av_f): raise FileNotFoundError(f"Não encontrado: {av_f}")


        if not os.path.exists(si_f): raise FileNotFoundError(f"Não encontrado: {si_f}")


        


    df = _parse_csv(av_f, si_f)


    


    # Gera e substitui o Geral.xlsx na pasta local/servidor apenas se necessário


    try:


        local_dir = db_path if db_path else str(DADOS_DIR)


        geral_out = os.path.join(local_dir, "Geral.xlsx")


        


        build_needed = True


        if os.path.exists(geral_out) and os.path.exists(av_f):


            if os.path.getmtime(geral_out) >= os.path.getmtime(av_f):


                build_needed = False


                


        if build_needed:


            xlsx_bytes = _build_workbook(df, "GERAL — AADP 2026", df)


            with open(geral_out, "wb") as f_out:


                f_out.write(xlsx_bytes)


    except Exception:


        pass


        


    return df





def apply_filters(df, rpm_f, unid_f, sc_f, st_f, cert_f):


    if rpm_f:  df = df[df["Unidade RPM (Avaliado)"].isin(rpm_f)]


    if unid_f: df = df[df["Unidade Principal (Avaliado)"].isin(unid_f)]


    if sc_f:   df = df[df["Situação Comissão"].isin(sc_f)]


    if st_f:   df = df[df["Status Avaliação"].isin(st_f)]


    if cert_f: df = df[df["Certificação Homologador"].isin(cert_f)]


    return df





def fmt_num(n): return f"{n:,}".replace(",",".")





def df_to_xlsx(df: pd.DataFrame) -> bytes:


    import io


    output = io.BytesIO()


    with pd.ExcelWriter(output, engine='openpyxl') as writer:


        df.to_excel(writer, index=False, sheet_name='Planilha')


    output.seek(0)


    return output.read()





MAX_STYLE = 4_000_000


def clean_none_values(df):
    if df is None or not hasattr(df, "columns"):
        return df
    import pandas as pd
    df = df.copy()
    df = df.fillna("-")
    for col in df.columns:
        if df[col].dtype == object or str(df[col].dtype) == "string":
            df[col] = df[col].apply(lambda x: "-" if str(x).strip().lower() in ("none", "nan", "<na>", "nat") else x)
            df[col] = df[col].replace({None: "-", "None": "-"})
    return df


def style_audit_dataframe(df):
    if df is None or not hasattr(df, "columns") or df.empty:
        return df
    col_media = next((c for c in df.columns if "Aritm" in str(c)), None)
    cols_av1 = [c for c in df.columns if str(c).endswith(" 1")]
    cols_av2 = [c for c in df.columns if str(c).endswith(" 2")]
    cols_av3 = [c for c in df.columns if str(c).endswith(" 3")]
    cols_av4 = [c for c in df.columns if str(c).endswith(" 4")]
    
    styles = {}
    if col_media:
        styles[col_media] = "background-color: #ffe599; color: black; font-weight: 500;" # Soft yellow
    for c in cols_av1:
        styles[c] = "background-color: #e2f0d9; color: #2d6a0f; font-weight: 500;" # Soft green
    for c in cols_av2:
        styles[c] = "background-color: #fce4d6; color: #7a3d00; font-weight: 500;" # Soft orange/peach
    for c in cols_av3:
        styles[c] = "background-color: #ebd9eb; color: #4a148c; font-weight: 500;" # Soft purple
    for c in cols_av4:
        styles[c] = "background-color: #e8f0fe; color: #1a0dab; font-weight: 500;" # Soft blue

    def get_column_styles(row):
        return [styles.get(col, "") for col in row.index]

    styler = df.style.apply(get_column_styles, axis=1)
    if col_media:
        def format_media(val):
            try:
                import math
                if val is None or val == "-":
                    return "-"
                f_val = float(str(val).replace(",", "."))
                if math.isnan(f_val):
                    return "-"
                return f"{f_val:.2f}"
            except Exception:
                return str(val)
        styler = styler.format(formatter={col_media: format_media})
    return styler


def safe_df(styled_or_df, height=520, key_prefix=None):
    """Exibe um DataFrame com st.dataframe nativo.
    - Ordenacao crescente/decrescente: clique no cabecalho de qualquer coluna.
    - Filtro rapido: campo de busca global acima da tabela.
    """
    import pandas as pd

    # Extrair DataFrame subjacente
    if hasattr(styled_or_df, "data"):
        styled_or_df.data = clean_none_values(styled_or_df.data)
        raw_df = styled_or_df.data.copy()
        is_styled = True
        is_large = raw_df.size > MAX_STYLE
    elif isinstance(styled_or_df, pd.DataFrame):
        raw_df = clean_none_values(styled_or_df)
        is_styled = False
        is_large = False
    else:
        raw_df = clean_none_values(styled_or_df)
        is_styled = False
        is_large = False

    if not isinstance(raw_df, pd.DataFrame) or raw_df.empty:
        st.info("\u2139\ufe0f Nenhum dado dispon\u00edvel para exibir.")
        return

    # Gerar chave unica para nao colidir entre telas
    if key_prefix is None:
        import hashlib
        key_prefix = "sdf_" + hashlib.md5(
            "".join(str(c) for c in raw_df.columns).encode()
        ).hexdigest()[:8]

    # Campo de busca rapida global
    busca = st.text_input(
        "\U0001f50d Busca r\u00e1pida (filtra qualquer coluna)",
        key=f"{key_prefix}_busca",
        placeholder="Digite para filtrar\u2026",
        label_visibility="collapsed",
    )

    df_filtered = raw_df.copy()
    if busca and busca.strip():
        termo = busca.strip().lower()
        mask = df_filtered.apply(
            lambda col: col.astype(str).str.lower().str.contains(termo, na=False)
        ).any(axis=1)
        df_filtered = df_filtered[mask]

    total = len(raw_df)
    shown = len(df_filtered)
    limit = 500
    
    if shown > limit:
        st.caption(f"📊 Exibindo as primeiras **{limit}** de **{shown:,}** linhas encontradas (Total: {total:,} registros). Use o campo **Busca rápida** acima para filtrar.")
        df_to_show = df_filtered.head(limit)
    else:
        st.caption(f"📊 Exibindo **{shown:,}** de **{total:,}** registros.")
        df_to_show = df_filtered

    # Exibir com estilo quando possivel
    is_audit_df = any("Aritm" in str(c) for c in df_to_show.columns) and any(str(c).endswith(" 1") for c in df_to_show.columns)
    
    if is_audit_df:
        try:
            styled_df = style_audit_dataframe(df_to_show)
            st.dataframe(styled_df, use_container_width=True, height=height)
        except Exception:
            st.dataframe(df_to_show, use_container_width=True, height=height)
    elif is_styled and not is_large:
        try:
            st.dataframe(styled_or_df.data.loc[df_to_show.index], use_container_width=True, height=height)
        except Exception:
            st.dataframe(df_to_show, use_container_width=True, height=height)
    else:
        st.dataframe(df_to_show, use_container_width=True, height=height)


def color_status(val):


    m = {"Encerrada":"background-color:#e8f5e1;color:#2d6a0f;font-weight:600",


         "Homologação":"background-color:#fff8db;color:#7a5c00;font-weight:600",


         "Parcialmente Encerrada":"background-color:#fff0db;color:#7a3d00;font-weight:600",


         "Aberta":"background-color:#fde8e8;color:#8b0000;font-weight:600"}


    return m.get(val, "")





def color_sit(val):


    if val == "Comissão Atual":   return "background-color:#dce8f5;color:#1a3a6a;font-weight:600"


    if val == "Nota Provisória":  return "background-color:#fff9e6;color:#7a5c00;font-weight:600"


    return ""





# ─────────────────────── SEGURANÇA E AUTENTICAÇÃO ──────────────────────────────


if "authenticated" not in st.session_state:


    st.session_state.authenticated = False


    st.session_state.user_pm = ""


    st.session_state.user_name = ""


    st.session_state.user_role = ""


    st.session_state.user_rpm = ""


    st.session_state.user_unit = ""





if not st.session_state.authenticated:


    c1, c2, c3 = st.columns([1, 2, 1])


    with c2:


        if os.path.exists("logo_drh.png"):
            st.image("logo_drh.png", use_container_width=True)
        else:
            st.markdown("<div style='text-align: center; font-size: 4.5rem; margin-bottom: 25px;'>👮</div>", unsafe_allow_html=True)


        st.markdown("<h2 style='text-align: center; color: #9b8a5c; margin-top: -15px;'>Painel de Controle AADP</h2>", unsafe_allow_html=True)


        st.markdown("<h4 style='text-align: center; color: #a0a0a0; font-size: 0.95rem;'>Polícia Militar de Minas Gerais · Resolução 5458/2025</h4>", unsafe_allow_html=True)


        st.markdown("---")


        


        if "auth_mode" not in st.session_state:


            st.session_state.auth_mode = "🔑 Acessar Conta"


            


        auth_mode = st.radio("Selecione uma opção:", ["🔑 Acessar Conta", "📝 Solicitar Cadastro"], horizontal=True, key="auth_mode_radio")


        st.markdown("---")


        


        if auth_mode == "🔑 Acessar Conta":
            if st.session_state.get("forgot_password_mode", False):
                st.markdown("##### ❓ Recuperar Acesso (Esqueci minha senha)")
                
                if st.session_state.get("forgot_step2", False) and st.session_state.get("forgot_target_pm"):
                    target_pm = st.session_state.forgot_target_pm
                    
                    st.warning(
                        f"⚠️ **Atenção (PM: {target_pm})**:\n\n"
                        "Para recuperar seu acesso você precisa revogar o acesso e criar novo acesso.\n\n"
                        "Deseja prosseguir com a revogação do seu cadastro atual?"
                    )
                    
                    col_c1, col_c2 = st.columns(2)
                    with col_c1:
                        if st.button("Cancelar", use_container_width=True, key="btn_forgot_cancel"):
                            st.session_state.forgot_password_mode = False
                            st.session_state.forgot_step2 = False
                            st.session_state.forgot_target_pm = None
                            st.rerun()
                    with col_c2:
                        if st.button("Revogar", use_container_width=True, type="primary", key="btn_forgot_revoke"):
                            db_revoke_user(target_pm)
                            log_action(target_pm, "REVOGAR_AUTOCADASTRO", "Usuario solicitou revogacao por esquecimento de senha")
                            st.success("✅ Seu cadastro foi revogado com sucesso! Agora você pode solicitar um novo cadastro.")
                            st.session_state.forgot_password_mode = False
                            st.session_state.forgot_step2 = False
                            st.session_state.forgot_target_pm = None
                            st.session_state.auth_mode_radio = "📝 Solicitar Cadastro"
                            st.rerun()
                else:
                    st.write("Informe o seu Nº PM para verificar seu cadastro:")
                    forgot_pm = st.text_input("Nº PM:", key="forgot_pm_input", placeholder="Ex: 123456")
                    
                    col_f1, col_f2 = st.columns(2)
                    with col_f1:
                        if st.button("Voltar ao Login", use_container_width=True, key="btn_forgot_back"):
                            st.session_state.forgot_password_mode = False
                            st.rerun()
                    with col_f2:
                        if st.button("Verificar Cadastro", use_container_width=True, type="primary", key="btn_forgot_verify"):
                            if not forgot_pm or not forgot_pm.isdigit():
                                st.error("Por favor, informe um Nº PM válido (apenas números).")
                            else:
                                f_pm = forgot_pm.strip()
                                status = db_check_user_status(f_pm)
                                if not status:
                                    st.error("❌ Nº PM não encontrado no sistema!")
                                elif status == "Bloqueado":
                                    st.warning("⚠️ Seu acesso já está revogado. Você pode ir para a opção '📝 Solicitar Cadastro' e realizar seu novo cadastro.")
                                else:
                                    st.session_state.forgot_target_pm = f_pm
                                    st.session_state.forgot_step2 = True
                                    st.rerun()
            else:
                with st.form("form_login", clear_on_submit=False):
                    login_pm = st.text_input("Nº PM:", key="login_pm_val", placeholder="Ex: 123456 ou ADM")
                    login_pass = st.text_input("Senha:", type="password", key="login_pass_val")
                    submitted_login = st.form_submit_button("Entrar", use_container_width=True, type="primary")

                if submitted_login:
                    spm = login_pm.strip()
                    spass = login_pass
                    
                    if not spm or not spass:
                        st.error("Por favor, preencha todos os campos.")
                    else:
                        h_pass = hashlib.sha256(spass.encode()).hexdigest()
                        res = db_get_user_for_login(spm, h_pass)
                        if res:
                            name, role, rpm, unit, status = res
                            if status == "Ativo":
                                st.session_state.authenticated = True
                                st.session_state.user_pm = spm
                                st.session_state.user_name = name
                                if role in ("P1/SADM", "P1"):
                                    role = "P1"
                                elif role in ("DRH6", "Gestor"):
                                    role = "Gestor"
                                st.session_state.user_role = role
                                st.session_state.user_rpm = rpm
                                st.session_state.user_unit = unit
                                refresh_db_cache()
                                log_action(spm, "LOGIN", "Acesso realizado com sucesso")
                                st.success(f"Bem-vindo, {name}!")
                                st.rerun()
                            elif status == "Pendente":
                                st.warning("⚠️ Sua conta está aguardando liberação do Administrador.")
                            else:
                                st.error("❌ Acesso revogado/bloqueado. Entre em contato com a DRH.")
                        else:
                            st.error("❌ Nº PM ou Senha incorretos.")

                st.markdown("<div style='margin-top: 10px;'></div>", unsafe_allow_html=True)
                if st.button("❓ Esqueci minha senha", use_container_width=True, key="btn_forgot_password"):
                    st.session_state.forgot_password_mode = True
                    st.session_state.forgot_step2 = False
                    st.session_state.forgot_target_pm = None
                    st.rerun()


                        


        else:


            st.markdown("##### 📝 Solicitação de acesso - Painel de Controle AADP")


            st.info("⚠️ Informe apenas os **6 primeiros dígitos** do seu Nº PM (sem o dígito verificador).")


            


            # Inicializa variáveis de estado


            if "sigef_data" not in st.session_state:


                st.session_state.sigef_data = None


            if "sigef_verified" not in st.session_state:


                st.session_state.sigef_verified = False


                


            # Entrada de Nº PM


            reg_pm = st.text_input("Nº PM (Apenas os 6 primeiros dígitos):", max_chars=6, key="reg_pm", placeholder="Ex: 053108")


            


            col_cons, col_clear = st.columns([3, 1])


            with col_cons:


                if st.button("🔍 Consultar", use_container_width=True, type="secondary"):


                    if not reg_pm or not reg_pm.isdigit():


                        st.error("Por favor, informe um Nº PM válido (apenas números, máximo 6 dígitos).")


                        st.session_state.sigef_data = None


                        st.session_state.sigef_verified = False


                    else:


                        with st.spinner("⏳ Consultando banco de dados do SIGEF..."):


                            res = find_sigef_user(reg_pm)


                            if res:


                                st.session_state.sigef_data = res


                                st.session_state.sigef_verified = False # Reseta a verificação para nova busca


                                st.success("✅ Militar encontrado no SIGEF! Prossiga com a verificação de segurança abaixo.")


                            else:


                                st.error("❌ Nº PM não encontrado no banco SIGEF. Verifique se digitou os 6 primeiros dígitos corretamente.")


                                st.session_state.sigef_data = None


                                st.session_state.sigef_verified = False


            with col_clear:


                if st.button("🧹 Limpar", use_container_width=True):


                    st.session_state.sigef_data = None


                    st.session_state.sigef_verified = False


                    st.rerun()





            if st.session_state.sigef_data:


                data = st.session_state.sigef_data


                


                # Etapa 2: Verificação de Segurança (CPF e Nascimento)


                if not st.session_state.sigef_verified:


                    st.markdown("---")


                    st.markdown("##### 🔒 Verificação de Segurança")


                    st.write("Para confirmar se realmente é você, confirme as duas informações abaixo:")


                    


                    v_cpf = st.text_input("Digite seu CPF (apenas números):", max_chars=11, key="v_cpf", placeholder="Ex: 12345678901")


                    v_birth = st.text_input("Digite sua Data de Nascimento (DD/MM/AAAA):", max_chars=10, key="v_birth", placeholder="Ex: 25/09/1957")


                    


                    if st.button("Confirmar Dados", use_container_width=True, type="primary"):


                        # Normaliza CPF: remove pontuação e também todos os zeros à esquerda


                        clean_input_cpf = re.sub(r'\D', '', v_cpf).lstrip('0')


                        clean_sigef_cpf = re.sub(r'\D', '', data["cpf"]).lstrip('0')


                        


                        # Normaliza Data de Nascimento: remove qualquer caractere que não seja número (ex: barras)


                        clean_input_birth = re.sub(r'\D', '', v_birth)


                        clean_sigef_birth = re.sub(r'\D', '', data["birthdate"])


                        


                        if clean_input_cpf == clean_sigef_cpf and clean_input_birth == clean_sigef_birth:


                            st.session_state.sigef_verified = True


                            st.success("✅ Identidade confirmada com sucesso!")


                            st.rerun()


                        else:


                            st.error("❌ CPF ou Data de Nascimento incorretos. Verifique suas informações e tente novamente.")


                


                # Etapa 3: Liberação do Formulário de Senha


                if st.session_state.sigef_verified:


                    st.markdown("---")


                    st.markdown("##### 👤 Dados Funcionais Confirmados:")


                    st.text_input("Posto/Graduação:", value=data["rank"], disabled=True, key="disp_rank")


                    st.text_input("Nome Completo:", value=data["name"], disabled=True, key="disp_name")


                    st.text_input("UDI/UDG:", value=data["rpm"], disabled=True, key="disp_rpm")


                    st.text_input("Unidade Principal:", value=data["unit"], disabled=True, key="disp_unit")


                    st.text_input("Setor:", value=data["sector"], disabled=True, key="disp_sector")


                    


                    st.markdown("##### 🔑 Configuração de Senha de Acesso:")


                    with st.form("form_cadastro_final", clear_on_submit=False):
                        st.warning("⚠️ **Segurança:** Por motivos de segurança, a senha de acesso cadastrada **NÃO** deve ser igual à sua senha da **IntranetPM** ou do **SIRH**.")
                        reg_pass = st.text_input("Escolha uma Senha:", type="password", key="reg_pass")


                        reg_pass_conf = st.text_input("Confirme a Senha:", type="password", key="reg_pass_conf")


                        


                        submitted = st.form_submit_button("Enviar Solicitação", use_container_width=True, type="primary")


                        


                    if submitted:


                        spass = reg_pass


                        if not spass:


                            st.error("Por favor, preencha o campo de senha.")


                        elif spass != reg_pass_conf:


                            st.error("As senhas não coincidem!")


                        elif len(spass) < 6:


                            st.error("A senha deve ter pelo menos 6 caracteres.")


                        else:


                            try:


                                current_status = db_check_user_status(data["pm"])


                                if current_status:


                                    if current_status in ("Pendente", "Ativo"):


                                        st.error("❌ Este Nº PM já possui solicitação de acesso ativa ou pendente no sistema!")


                                    else:


                                        # Usuário Bloqueado ou Recusado - pode solicitar novamente


                                        h_pass = hashlib.sha256(spass.encode()).hexdigest()


                                        db_re_request_access(data["pm"], data["name"], data["rank"], data["rpm"], data["unit"], data["sector"], h_pass)


                                        log_action(data["pm"], "RE_CADASTRO_SOLICITADO", f"Nome: {data['name']}, Posto: {data['rank']}")


                                        st.success("✅ Nova solicitação enviada com sucesso! Aguarde a liberação do Administrador.")


                                        st.session_state.sigef_data = None


                                        st.session_state.sigef_verified = False


                                else:


                                    h_pass = hashlib.sha256(spass.encode()).hexdigest()


                                    db_create_new_request(data["pm"], data["name"], data["rank"], data["rpm"], data["unit"], data["sector"], h_pass)


                                    log_action(data["pm"], "CADASTRO_SOLICITADO", f"Nome: {data['name']}, Posto: {data['rank']}, UDI/UDG: {data['rpm']}")


                                    st.success("✅ Solicitação enviada com sucesso! Aguarde a liberação do Administrador.")


                                    st.session_state.sigef_data = None


                                    st.session_state.sigef_verified = False


                                    st.session_state.sigef_verified = False


                            except Exception as e:


                                st.error(f"Erro ao salvar cadastro: {str(e)}")

    st.stop()



# ─────────────────────── SIDEBAR ──────────────────────────────────────────────


with st.sidebar:


    st.image("logo_drh.png", use_container_width=True)


    st.markdown("### AADP 2026")


    st.markdown("**Sistema de Análise de Avaliações**")


    


    # Determina o perfil ativo (real vs simulado) para ajustar as opções da barra lateral


    sidebar_active_role = st.session_state.get("simulated_role", st.session_state.user_role) if st.session_state.get("simulation_active", False) else st.session_state.user_role


    


    # Exibe informações do militar


    st.markdown(f"<small>👤 <b>Militar:</b> {st.session_state.user_name} ({st.session_state.user_pm})</small>", unsafe_allow_html=True)


    if st.session_state.get("simulation_active", False) and st.session_state.user_role == "ADMINISTRADOR":


        st.markdown(f"<small>🔑 <b>Perfil Real:</b> ADMINISTRADOR</small>", unsafe_allow_html=True)


        st.markdown(f"<small>🕵️ <b>Simulado:</b> <span style='color:#ff9f43;'>{st.session_state.simulated_role}</span></small>", unsafe_allow_html=True)


        if st.button("Voltar simulação", key="sidebar_stop_sim", type="primary", use_container_width=True):
            st.session_state.simulation_active = False
            st.session_state.simulated_pm = ""
            st.session_state.simulated_name = ""
            st.session_state.simulated_role = ""
            st.session_state.simulated_rpm = ""
            st.session_state.simulated_unit = ""
            st.session_state.active_page = "Painel Administrador"
            log_action("ADM", "ENCERRAR_SIMULACAO", "Simulacao desativada")
            st.rerun()


    else:


        st.markdown(f"<small>🔑 <b>Perfil:</b> <span style='color:#9b8a5c;'>{sidebar_active_role}</span></small>", unsafe_allow_html=True)


        


    st.markdown("---")





    # 1. Mostrar Filtros (Primeiro)


    if "show_filtros" not in st.session_state:


        st.session_state.show_filtros = False


        


    btn_filtros_label = "🔍 Ocultar Filtros" if st.session_state.show_filtros else "🔍 Mostrar Filtros"


    btn_filtros_type = "primary" if st.session_state.show_filtros else "secondary"


    if st.button(btn_filtros_label, use_container_width=True, key="btn_toggle_filtros", type=btn_filtros_type):


        st.session_state.show_filtros = not st.session_state.show_filtros


        st.rerun()





    container_filtros = st.container()


    st.markdown("---")





    # 2. Páginas / Navegação (Segundo)


    st.markdown("#### 🧭 Páginas")


    pages = [


        ("📊 Análise Gráfica", "Análise Gráfica"),


        ("📋 Dados Gerais", "Dados Gerais"),


        ("⏳ Avaliações Pendentes", "Avaliações Pendentes"),


        ("👥 Avaliadores Pendentes", "Avaliadores Pendentes"),


    ]


    # P1 e SADM não possuem acesso às opções de exportação/relatórios


    if sidebar_active_role not in ("P1", "SADM"):
        pages.append(("📥 Gerar Relatório", "Gerar Relatório"))
        pages.append(("📄 Relatório Word", "Relatório Word"))

    # Auditoria de Notas: visível para ADMINISTRADOR, GESTOR, P1 e SADM
    if sidebar_active_role.upper() in ("ADMINISTRADOR", "GESTOR", "P1", "SADM"):
        pages.append(("📊 Auditoria de Notas", "Auditoria de Notas"))

    if sidebar_active_role.upper() in ("ADMINISTRADOR", "GESTOR"):
        pages.append(("📊 Dados Consolidados", "Dados Consolidados"))

    # O administrador real sempre vê o painel administrador


    if st.session_state.user_role == "ADMINISTRADOR":


        pending_count = db_get_pending_count()


            


        if pending_count > 0:


            pages.append((f"⚙️ Painel Administrador (🔴 :red[{pending_count}])", "Painel Administrador"))


        else:


            pages.append(("⚙️ Painel Administrador", "Painel Administrador"))





    if "active_page" not in st.session_state:


        st.session_state.active_page = "Análise Gráfica"


        


    if st.session_state.active_page == "Painel Administrador" and st.session_state.user_role != "ADMINISTRADOR":


        st.session_state.active_page = "Análise Gráfica"


        


    if st.session_state.active_page in ("Gerar Relatório", "Relatório Word") and sidebar_active_role in ("P1", "SADM"):


        st.session_state.active_page = "Análise Gráfica"





    for label, page_name in pages:


        is_active = st.session_state.active_page == page_name


        btn_type = "primary" if is_active else "secondary"


        if st.button(label, key=f"nav_{page_name}", use_container_width=True, type=btn_type):
            st.session_state.active_page = page_name
            st.rerun()





    # Inicializa variáveis para não dar NameError


    drive_av_id = drive_si_id = drive_geral_id = ""


    db_path = ""


    fonte = cfg.get("fonte_dados", "📁 Pasta local / Servidor")


    reload = False





    # Recarregar Dados para Administrador


    if st.session_state.user_role == "ADMINISTRADOR":


        st.markdown("---")


        if st.button("🔄 Recarregar Dados", use_container_width=True, type="primary", key="btn_reload"):
            st.cache_data.clear()
            st.cache_resource.clear()
            
            # Limpar arquivos baixados para forçar download novo
            import tempfile
            cache_dir = os.path.join(tempfile.gettempdir(), "aadp_drive_cache")
            for f in ["avaliacoes.csv", "SIGEF.csv", "geral.csv"]:
                p = os.path.join(cache_dir, f)
                if os.path.exists(p):
                    try: os.remove(p)
                    except: pass
            
            # Limpar planilha mestre
            local_master = os.path.join(str(DADOS_DIR), "Analise avaliacoes completa.xlsx")
            if os.path.exists(local_master):
                try: os.remove(local_master)
                except: pass
                
            local_master_parent = os.path.join(str(Path(DADOS_DIR).parent), "Analise avaliacoes completa.xlsx")
            if os.path.exists(local_master_parent):
                try: os.remove(local_master_parent)
                except: pass

            if "db_users" in st.session_state:
                del st.session_state.db_users
            if "db_logs" in st.session_state:
                del st.session_state.db_logs
            st.success("Dados recarregados com sucesso!")
            st.rerun()





    # Botões de Alterar Senha e Sair/Logoff no final da barra lateral


    st.markdown("---")


    if st.button("🔑 Alterar Senha", use_container_width=True, key="btn_toggle_change_password"):


        st.session_state.show_change_password = not st.session_state.get("show_change_password", False)


        st.rerun()





    if st.button("🚪 Sair / Logoff", use_container_width=True, key="btn_logoff"):


        log_action(st.session_state.user_pm, "LOGOFF", "Saída voluntária")


        st.session_state.authenticated = False


        st.session_state.user_pm = ""


        st.session_state.user_name = ""


        st.session_state.user_role = ""


        st.session_state.user_rpm = ""


        st.session_state.user_unit = ""


        st.session_state.simulation_active = False


        st.session_state.simulated_pm = ""


        st.session_state.simulated_name = ""


        st.session_state.simulated_role = ""


        st.session_state.simulated_rpm = ""


        st.session_state.simulated_unit = ""


        st.rerun()

# ─────────────────────── CARREGAR DADOS ───────────────────────────────────────


# Calcula as variáveis ativas considerando simulação


if st.session_state.get("simulation_active", False):


    active_role = st.session_state.get("simulated_role", st.session_state.user_role)


    active_rpm = st.session_state.get("simulated_rpm", st.session_state.user_rpm)


    active_unit = st.session_state.get("simulated_unit", st.session_state.user_unit)


    active_pm = st.session_state.get("simulated_pm", st.session_state.user_pm)


    active_name = st.session_state.get("simulated_name", st.session_state.user_name)


else:


    active_role = st.session_state.user_role


    active_rpm = st.session_state.user_rpm


    active_unit = st.session_state.user_unit


    active_pm = st.session_state.user_pm


    active_name = st.session_state.user_name





try:


    if reload: st.cache_data.clear()


    av_csv_path = os.path.join(db_path or cfg.get("db_path", str(DADOS_DIR)), "avaliacoes.csv")


    last_mod_str = get_last_updated_time(av_csv_path, drive_av_id or cfg.get("drive_av_id", ""))


    


    df_full = load_data(


        db_path   = db_path or cfg.get("db_path", str(DADOS_DIR)),


        drive_av_id = drive_av_id or cfg.get("drive_av_id", ""),


        drive_si_id = drive_si_id or cfg.get("drive_si_id", ""),


    )


    if active_role == "P1":


        df_full = df_full[df_full["Unidade RPM (Avaliado)"].apply(lambda x: matches_rpm(active_rpm, x))]


    elif active_role == "SADM":


        df_full = df_full[df_full["Unidade Principal (Avaliado)"].str.strip().str.upper() == active_unit.strip().upper()]


        


    data_ok = True; ts = now_br().strftime("%d/%m/%Y %H:%M")


except Exception as e:


    data_ok = False; err_msg = str(e); last_mod_str = "Data/Hora indisponível"





# ─────────────────────── FILTROS ──────────────────────────────────────────────


rpm_filter = unid_filter = sit_com_filter = status_filter = cert_filter = []


if data_ok:


    all_rpm   = sorted(df_full["Unidade RPM (Avaliado)"].dropna().unique(), key=rpm_sort_key)


    all_status= ["Aberta","Parcialmente Encerrada","Homologação","Encerrada"]


    all_sit   = ["Comissão Atual","Nota Provisória"]


    all_cert  = ["SIM","NÃO","-"]


    with container_filtros:


        if st.session_state.show_filtros:


            st.markdown("#### 🔍 Filtros de Visualização")


            if active_role not in ("P1", "SADM"):


                rpm_filter = st.multiselect("🏢 Unidade RPM", all_rpm, placeholder="Todas")


            else:


                rpm_filter = []


            df_tmp = df_full[df_full["Unidade RPM (Avaliado)"].isin(rpm_filter)] if rpm_filter else df_full


            all_unid = sorted(df_tmp["Unidade Principal (Avaliado)"].dropna().unique())


            


            if active_role != "SADM":


                unid_filter = st.multiselect("🏛️ Subunidade", all_unid, placeholder="Todas")


            else:


                unid_filter = []


                


            st.markdown("")


            sit_com_filter = st.multiselect("🔵 Situação Comissão", all_sit, placeholder="Todas")


            status_filter  = st.multiselect("📊 Status",            all_status, placeholder="Todos")


            cert_filter    = st.multiselect("✅ Cert. Homologador",  all_cert,   placeholder="Todos")


            st.markdown("---")


            st.markdown(f"<small>🕐 Carregado: {ts}</small>", unsafe_allow_html=True)


            st.markdown(f"<small>📊 {fmt_num(len(df_full))} registros</small>", unsafe_allow_html=True)


    df = apply_filters(df_full, rpm_filter, unid_filter, sit_com_filter, status_filter, cert_filter)


else:


    df = pd.DataFrame()





# ─────────────────────── CABEÇALHO ────────────────────────────────────────────


logo_base64 = ""


if os.path.exists("logo_drh.png"):


    import base64


    with open("logo_drh.png", "rb") as f:


        logo_base64 = base64.b64encode(f.read()).decode("utf-8")





logo_html = f'<img src="data:image/png;base64,{logo_base64}" style="width: 100%; max-width: 480px; height: auto; max-height: 120px; object-fit: contain; border-radius: 8px; align-self: center;" />' if logo_base64 else ""





st.markdown(f'<div class="main-title">{logo_html}<div style="margin-top: 10px;"><h1 style="font-size: 2.3rem; margin: 0; font-weight: 800; color: #9b8a5c; text-transform: uppercase;">Painel de Controle AADP</h1><p style="font-size: 1.05rem; margin: 5px 0 0 0; color: #e5dccb; font-weight: 500;">Polícia Militar de Minas Gerais · Resolução 5458/2025</p><p style="font-size: 0.9rem; margin-top: 8px; color: #a0a0a0; font-style: italic;">Dados consolidados em {last_mod_str}</p></div></div>', unsafe_allow_html=True)





# --- FORMULÁRIO DE ALTERAÇÃO DE SENHA ---
if st.session_state.get("show_change_password", False):
    st.markdown('<div class="info-box" style="border-left: 5px solid #9b8a5c;">🛡️ <b>Alterar Senha do Usuário</b></div>', unsafe_allow_html=True)
    with st.form("form_change_password", clear_on_submit=True):
        curr_pw = st.text_input("Senha Atual:", type="password", key="chg_curr_pw")
        new_pw = st.text_input("Nova Senha:", type="password", key="chg_new_pw")
        conf_pw = st.text_input("Confirmar Nova Senha:", type="password", key="chg_conf_pw")
        submit_chg = st.form_submit_button("💾 Atualizar Senha", use_container_width=True, type="primary")
            
    if submit_chg:
        if not curr_pw or not new_pw or not conf_pw:
            st.error("❌ Por favor, preencha todos os campos.")
        elif new_pw != conf_pw:
            st.error("❌ A nova senha e a confirmação não coincidem.")
        else:
            h_curr = hashlib.sha256(curr_pw.encode()).hexdigest()
            row_pw = db_get_user_password(st.session_state.user_pm)
            if not row_pw or row_pw != h_curr:
                st.error("❌ Senha atual incorreta.")
            else:
                h_new = hashlib.sha256(new_pw.encode()).hexdigest()
                db_update_password(st.session_state.user_pm, h_new)
                log_action(st.session_state.user_pm, "ALTERAR_SENHA", "Senha alterada com sucesso pelo proprio usuario")
                st.success("✅ Senha alterada com sucesso!")
                st.session_state.show_change_password = False
                st.rerun()
                
    if st.button("❌ Cancelar / Fechar", use_container_width=True, key="btn_cancel_change_pw"):
        st.session_state.show_change_password = False
        st.rerun()
    st.markdown("---")


if not data_ok:
    st.error(f"❌ {err_msg}")
    st.markdown(f"""<div class="info-box">
    👈 Configure a pasta dos CSVs na barra lateral.<br>
    📂 Pasta padrão criada: <code>{DADOS_DIR}</code><br>
    Coloque os arquivos <code>avaliacoes.csv</code> e <code>SIGEF.csv</code> nessa pasta.
    </div>""", unsafe_allow_html=True)
    st.stop()


# ─────────────────────── KPI CARDS ────────────────────────────────────────────
n_total  = len(df)
n_enc    = (df["Status Avaliação"]=="Encerrada").sum()
n_hom    = (df["Status Avaliação"]=="Homologação").sum()
n_parc   = (df["Status Avaliação"]=="Parcialmente Encerrada").sum()
n_aberta = (df["Status Avaliação"]=="Aberta").sum()
n_ca     = (df["Situação Comissão"]=="Comissão Atual").sum()
n_np     = (df["Situação Comissão"]=="Nota Provisória").sum()



col_block1, col_block2 = st.columns([1, 1.25], gap="large")

with col_block1:
    st.markdown('<div class="kpi-card kpi-total">'
                '<div class="label">TOTAL AVALIAÇÕES</div>'
                f'<div class="value">{fmt_num(n_total)}</div>'
                '<div class="sub">avaliações</div>'
                '</div>', unsafe_allow_html=True)

    st.markdown("<div style='margin-bottom: 12px;'></div>", unsafe_allow_html=True)

    cb1_1, cb1_2 = st.columns(2)
    with cb1_1:
        st.markdown('<div class="kpi-card kpi-ca">'
                    '<div class="label">COMISSÃO ATUAL</div>'
                    f'<div class="value">{fmt_num(n_ca)}</div>'
                    f'<div class="sub">{n_ca/max(n_total,1)*100:.1f}%</div>'
                    '</div>', unsafe_allow_html=True)
    with cb1_2:
        st.markdown('<div class="kpi-card kpi-np">'
                    '<div class="label">NOTA PROVISÓRIA</div>'
                    f'<div class="value">{fmt_num(n_np)}</div>'
                    f'<div class="sub">{n_np/max(n_total,1)*100:.1f}%</div>'
                    '</div>', unsafe_allow_html=True)

with col_block2:
    st.markdown('<div class="kpi-card kpi-enc">'
                '<div class="label">ENCERRADAS</div>'
                f'<div class="value">{fmt_num(n_enc)}</div>'
                f'<div class="sub">{n_enc/max(n_total,1)*100:.1f}%</div>'
                '</div>', unsafe_allow_html=True)

    st.markdown("<div style='margin-bottom: 12px;'></div>", unsafe_allow_html=True)

    cb2_1, cb2_2, cb2_3 = st.columns(3)
    with cb2_1:
        st.markdown('<div class="kpi-card kpi-aberta">'
                    '<div class="label">ABERTAS</div>'
                    f'<div class="value">{fmt_num(n_aberta)}</div>'
                    '<div class="sub">AV1 pendente</div>'
                    '</div>', unsafe_allow_html=True)
    with cb2_2:
        st.markdown('<div class="kpi-card kpi-parc">'
                    '<div class="label">PARC. ENCERRADA</div>'
                    f'<div class="value">{fmt_num(n_parc)}</div>'
                    '<div class="sub">AV2 pendente</div>'
                    '</div>', unsafe_allow_html=True)
    with cb2_3:
        st.markdown('<div class="kpi-card kpi-hom">'
                    '<div class="label">HOMOLOGAÇÃO</div>'
                    f'<div class="value">{fmt_num(n_hom)}</div>'
                    '<div class="sub">HOM pendente</div>'
                    '</div>', unsafe_allow_html=True)







st.markdown("<div class='main-nav-marker' style='margin-bottom: 15px;'></div>", unsafe_allow_html=True)

# ─────────────────────── HORIZONTAL NAVIGATION TABS ──────────────────────────
main_active_role = st.session_state.get("simulated_role", st.session_state.user_role) if st.session_state.get("simulation_active", False) else st.session_state.user_role

main_nav_pages = [
    ("📊\nAnálise Gráfica", "Análise Gráfica"),
    ("📋\nDados Gerais", "Dados Gerais"),
    ("⏳\nAvaliações Pendentes", "Avaliações Pendentes"),
    ("👥\nAvaliadores Pendentes", "Avaliadores Pendentes"),
]

if main_active_role not in ("P1", "SADM"):
    main_nav_pages.append(("📥\nGerar Relatório", "Gerar Relatório"))
    main_nav_pages.append(("📄\nRelatório Word", "Relatório Word"))

if main_active_role.upper() in ("ADMINISTRADOR", "GESTOR", "P1", "SADM"):
    main_nav_pages.append(("📊\nAuditoria de Notas", "Auditoria de Notas"))

if main_active_role.upper() in ("ADMINISTRADOR", "GESTOR"):
    main_nav_pages.append(("📊\nDados Consolidados", "Dados Consolidados"))

if st.session_state.user_role == "ADMINISTRADOR":
    p_count = db_get_pending_count()
    if p_count > 0:
        main_nav_pages.append((f"⚙️\nPainel Administrador ({p_count} 🔴)", "Painel Administrador"))
    else:
        main_nav_pages.append(("⚙️\nPainel Administrador", "Painel Administrador"))

# Render as horizontal buttons in columns
num_tabs = len(main_nav_pages)
tab_cols = st.columns(num_tabs)
for idx, (label, page_name) in enumerate(main_nav_pages):
    with tab_cols[idx]:
        is_active = (st.session_state.active_page == page_name)
        btn_type = "primary" if is_active else "secondary"
        if st.button(label, key=f"main_tab_{page_name}", use_container_width=True, type=btn_type):
            st.session_state.active_page = page_name
            st.rerun()

st.markdown("<div style='margin-bottom: 25px; border-bottom: 1px solid rgba(255,255,255,0.08);'></div>", unsafe_allow_html=True)





# ─────────────────────── SELEÇÃO DE ABAS VIA SESSION STATE ───────────────────


if "active_page" not in st.session_state:


    st.session_state.active_page = "Análise Gráfica"





active_page = st.session_state.active_page




# TAB 1 — ANÁLISE GRÁFICA


# ══════════════════════════════════════════════════════════════════════════════


if active_page == "Análise Gráfica":


    st.markdown("### 📊 Análise Gráfica das Avaliações")





    # ── LINHA 1: Pizza de Status (destaque, full width) ──────────────────────


    st.markdown("---")


    ordered_labels = ["Aberta","Parcialmente Encerrada","Homologação","Encerrada"]


    sd = df.groupby("Status Avaliação").size()


    vals_pizza  = [int(sd.get(s, 0)) for s in ordered_labels]


    cols_pizza  = [STATUS_COLORS[s] for s in ordered_labels]





    fig_status = go.Figure()





    # Sombra para efeito 3D (circle com proporção assimétrica = elipse)


    fig_status.add_shape(type="circle", xref="paper", yref="paper",


        x0=0.12, y0=0.01, x1=0.88, y1=0.10,


        fillcolor="rgba(0,0,0,0.18)", line_color="rgba(0,0,0,0)", layer="below")





    fig_status.add_trace(go.Pie(


        labels=ordered_labels, values=vals_pizza,


        hole=0.54,


        pull=[0.09, 0.06, 0.04, 0],


        texttemplate="<b>%{label}</b><br>%{value:,} (%{percent})",


        textposition="outside",


        rotation=135,  # Inclinado (tilted starting angle) para mover valores do topo mais para o lado!


        textfont=dict(size=13, family="Inter, sans-serif"),


        insidetextorientation="radial",


        marker=dict(colors=cols_pizza, line=dict(color="#121212", width=3)),


        hovertemplate="<b>%{label}</b><br>Avaliações: <b>%{value:,}</b><br>%{percent}<extra></extra>",


        sort=False,


    ))


    fig_status.add_annotation(


        text=f"<b>{fmt_num(n_total)}</b><br><span style='font-size:11px;color:#a0a0a0'>avaliações</span>",


        x=0.5, y=0.5, showarrow=False,


        font=dict(size=22, color="#9b8a5c", family="Inter"),


        align="center",


    )


    fig_status.update_layout(


        template="plotly_dark",


        title=dict(text="<b>Status das Avaliações — AADP 2026</b>",


                   font=dict(size=20, color="#9b8a5c"), x=0.5, y=0.96),


        height=500, showlegend=False,


        paper_bgcolor="rgba(0,0,0,0)",


        plot_bgcolor="rgba(0,0,0,0)",


        margin=dict(t=60, b=80, l=120, r=120),


    )


    st.plotly_chart(fig_status, use_container_width=True)





    # ── LINHA 2: Situação Comissão + Status × Comissão ────────────────────────


    st.markdown("---")


    c1, c2 = st.columns([1, 1.6])





    with c1:


        sit_d = df.groupby("Situação Comissão").size().reset_index(name="Qtd")


        fig_sit = go.Figure(go.Pie(


            labels=sit_d["Situação Comissão"], values=sit_d["Qtd"],


            hole=0.50, pull=[0.05,0],


            texttemplate="<b>%{label}</b><br>%{value:,} (%{percent})", textposition="outside",


            textfont=dict(size=12), sort=False,


            marker=dict(colors=[SIT_COLORS.get(s,"#aaa") for s in sit_d["Situação Comissão"]],


                        line=dict(color="#121212", width=3)),


            hovertemplate="<b>%{label}</b><br>%{value:,} avaliações (%{percent})<extra></extra>",


        ))


        fig_sit.update_layout(


            template="plotly_dark",


            title=dict(text="<b>Situação da Comissão</b>", font_size=15, x=0.5, font=dict(color="#9b8a5c")),


            height=380, showlegend=False,


            paper_bgcolor="rgba(0,0,0,0)",


            plot_bgcolor="rgba(0,0,0,0)",


            margin=dict(t=50, b=70, l=50, r=50),


        )


        st.plotly_chart(fig_sit, use_container_width=True)

    with c2:
        if "hide_enc_chart" not in st.session_state:
            st.session_state.hide_enc_chart = False
            
        btn_label = "👁️ Mostrar Encerradas" if st.session_state.hide_enc_chart else "🙈 Ocultar Encerradas"
        if st.button(btn_label, key="btn_toggle_enc_chart", use_container_width=True):
            st.session_state.hide_enc_chart = not st.session_state.hide_enc_chart
            st.rerun()
            
        excluir_encerradas = st.session_state.hide_enc_chart
        st.markdown("<p style='font-size: 0.78rem; color: #a0a0a0; margin-top: -8px; margin-bottom: 12px; font-style: italic;'>ℹ️ Oculta avaliações encerradas para ampliar e detalhar a escala dos status pendentes.</p>", unsafe_allow_html=True)
        
        cross = df.groupby(["Status Avaliação","Situação Comissão"]).size().reset_index(name="Qtd")
        
        if excluir_encerradas:
            cross = cross[cross["Status Avaliação"] != "Encerrada"]
            
        current_labels = [l for l in ordered_labels if l != "Encerrada"] if excluir_encerradas else ordered_labels
        cross["Status Avaliação"] = pd.Categorical(cross["Status Avaliação"],
                                                    categories=current_labels, ordered=True)
        cross = cross.sort_values("Status Avaliação")
        
        fig_bar = px.bar(cross, x="Status Avaliação", y="Qtd", color="Situação Comissão",
                         color_discrete_map=SIT_COLORS, barmode="group", text="Qtd",
                         template="plotly_dark",
                         title="<b>Status × Situação Comissão</b>")
        
        fig_bar.update_traces(textposition="outside", textfont_size=11)
        
        fig_bar.update_layout(height=380, title_font_size=15, title_x=0.5,
                               paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                               xaxis_title="", yaxis_title="Qtd",
                               showlegend=False,
                               title_font=dict(color="#9b8a5c"))
        
        fig_bar.update_xaxes(showgrid=False)
        fig_bar.update_yaxes(showgrid=True, gridcolor="#2a2a2a")
        st.plotly_chart(fig_bar, use_container_width=True)





    # ── LINHA 3: Barras por RPM ────────────────────────────────────────────────


    st.markdown("---")


    col_s1, col_s2 = st.columns([1, 1])
    with col_s1:
        sort_option = st.selectbox(
            "Ordenação das Unidades (RPM):",
            [
                "Crescente por Unidade",
                "Decrescente por Unidade",
                "Crescente por Quantidade",
                "Decrescente por Quantidade"
            ],
            key="dist_chart_sort_opt"
        )
    with col_s2:
        st.write("") # spacing

    # Initialize legend button states if not in session state
    if "dist_legend_enc" not in st.session_state:
        st.session_state.dist_legend_enc = True
    if "dist_legend_abe" not in st.session_state:
        st.session_state.dist_legend_abe = True
    if "dist_legend_par" not in st.session_state:
        st.session_state.dist_legend_par = True
    if "dist_legend_hom" not in st.session_state:
        st.session_state.dist_legend_hom = True

    # Legenda Interativa - Checkboxes para selecionar os status
    st.markdown("<p style='font-size: 0.95rem; font-weight: bold; margin-bottom: 6px; color: #9b8a5c;'>Legenda Interativa — Selecione os Status para Exibição e Ordenação:</p>", unsafe_allow_html=True)
    l1, l2, l3, l4 = st.columns(4)
    with l1:
        enc_label = "🟢 Encerrada" if st.session_state.dist_legend_enc else "⚪ Encerrada"
        if st.button(enc_label, key="btn_dist_enc", use_container_width=True):
            st.session_state.dist_legend_enc = not st.session_state.dist_legend_enc
            st.rerun()
        show_enc = st.session_state.dist_legend_enc
    with l2:
        abe_label = "🔴 Aberta" if st.session_state.dist_legend_abe else "⚪ Aberta"
        if st.button(abe_label, key="btn_dist_abe", use_container_width=True):
            st.session_state.dist_legend_abe = not st.session_state.dist_legend_abe
            st.rerun()
        show_abe = st.session_state.dist_legend_abe
    with l3:
        par_label = "🟠 Parcialmente Encerrada" if st.session_state.dist_legend_par else "⚪ Parcialmente Encerrada"
        if st.button(par_label, key="btn_dist_par", use_container_width=True):
            st.session_state.dist_legend_par = not st.session_state.dist_legend_par
            st.rerun()
        show_par = st.session_state.dist_legend_par
    with l4:
        hom_label = "🟡 Homologação" if st.session_state.dist_legend_hom else "⚪ Homologação"
        if st.button(hom_label, key="btn_dist_hom", use_container_width=True):
            st.session_state.dist_legend_hom = not st.session_state.dist_legend_hom
            st.rerun()
        show_hom = st.session_state.dist_legend_hom

    # Mapear status selecionados
    active_statuses = []
    if show_enc: active_statuses.append("Encerrada")
    if show_abe: active_statuses.append("Aberta")
    if show_par: active_statuses.append("Parcialmente Encerrada")
    if show_hom: active_statuses.append("Homologação")

    if not active_statuses:
        active_statuses = ["Encerrada", "Aberta", "Parcialmente Encerrada", "Homologação"]

    # Filtrar o DataFrame pelos status selecionados
    df_filtered = df[df["Status Avaliação"].isin(active_statuses)]

    # Obter lista de unidades presentes
    all_units = df_filtered["Unidade RPM (Avaliado)"].dropna().unique()
    if len(all_units) == 0:
        all_units = df["Unidade RPM (Avaliado)"].dropna().unique()

    # Ordenar as unidades com base na opção selecionada e nos status ativos
    if sort_option == "Crescente por Unidade":
        all_units_sorted = sorted(all_units, key=rpm_sort_key)
    elif sort_option == "Decrescente por Unidade":
        all_units_sorted = sorted(all_units, key=rpm_sort_key, reverse=True)
    elif sort_option == "Crescente por Quantidade":
        unit_totals = df_filtered.groupby("Unidade RPM (Avaliado)").size().reset_index(name="Total")
        unit_totals_sorted = unit_totals.sort_values("Total", ascending=True)
        all_units_sorted = list(unit_totals_sorted["Unidade RPM (Avaliado)"])
        for u in all_units:
            if u not in all_units_sorted:
                all_units_sorted.append(u)
    else: # Decrescente por Quantidade
        unit_totals = df_filtered.groupby("Unidade RPM (Avaliado)").size().reset_index(name="Total")
        unit_totals_sorted = unit_totals.sort_values("Total", ascending=False)
        all_units_sorted = list(unit_totals_sorted["Unidade RPM (Avaliado)"])
        for u in all_units:
            if u not in all_units_sorted:
                all_units_sorted.append(u)

    rpm_cross = df_filtered.groupby(["Unidade RPM (Avaliado)","Status Avaliação"]).size().reset_index(name="Qtd")
    
    fig_rpm = px.bar(
        rpm_cross, x="Unidade RPM (Avaliado)", y="Qtd",
        color="Status Avaliação", color_discrete_map=STATUS_COLORS,
        barmode="stack", text_auto=True,
        template="plotly_dark",
        title="<b>Distribuição por Unidade RPM e Status</b>",
        category_orders={
            "Unidade RPM (Avaliado)": all_units_sorted,
            "Status Avaliação": STACK_ORDER,
        },
    )
    
    fig_rpm.update_traces(textposition="auto")
    
    fig_rpm.update_layout(
        uirevision="constant_value",
        height=480, title_font_size=15, title_x=0.5,
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        xaxis_title="", yaxis_title="Avaliações",
        showlegend=False,
        title_font=dict(color="#9b8a5c")
    )
    
    fig_rpm.update_xaxes(tickangle=45, showgrid=False)
    fig_rpm.update_yaxes(showgrid=True, gridcolor="#2a2a2a")
    st.plotly_chart(fig_rpm, use_container_width=True)





    # ── LINHA 4: Certificação + Timeline AV1/AV2/HOM ──────────────────────────


    st.markdown("---")


    c3, c4 = st.columns([1, 1.8])





    with c3:


        cert_d = df["Certificação Homologador"].value_counts().reset_index()


        cert_d.columns = ["Cert","Qtd"]


        cert_map = {"SIM":"#FF6B6B","NÃO":"#70AD47","-":"#AAAAAA"}


        fig_cert = px.bar(cert_d, x="Cert", y="Qtd", color="Cert",


                          color_discrete_map=cert_map,


                          template="plotly_dark",


                          title="<b>Certificação Homologador</b>", text="Qtd")


        fig_cert.update_traces(textposition="outside", textfont_size=12)


        fig_cert.update_layout(height=360, title_x=0.5, showlegend=False,


                                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",


                                xaxis_title="", yaxis_title="Qtd",


                                title_font=dict(color="#9b8a5c"))


        fig_cert.update_xaxes(showgrid=False)


        fig_cert.update_yaxes(showgrid=True, gridcolor="#2a2a2a")


        st.plotly_chart(fig_cert, use_container_width=True)





    with c4:


        # Timeline com AV1, AV2 e HOM diferenciados por cor


        timeline_frames = []


        for col, label in [("Data AV1","AV1 — Avaliador 1"),


                            ("Data AV2","AV2 — Avaliador 2"),


                            ("Data HOM","HOM — Homologador")]:


            sub = df[df[col] != "-"].copy()


            if sub.empty: continue


            try:


                sub["Data"] = pd.to_datetime(sub[col], dayfirst=True, errors="coerce")


                sub = sub.dropna(subset=["Data"])


                cnt = sub.groupby("Data").size().reset_index(name="Qtd")


                cnt["Função"] = label


                timeline_frames.append(cnt)


            except Exception:


                pass


        if timeline_frames:


            df_time = pd.concat(timeline_frames, ignore_index=True)


            fig_time = px.line(


                df_time, x="Data", y="Qtd", color="Função",


                title="<b>Avaliações por Data e Função</b>",


                template="plotly_dark",


                markers=True,


                color_discrete_map={


                    "AV1 — Avaliador 1":"#4472C4",


                    "AV2 — Avaliador 2":"#ED7D31",


                    "HOM — Homologador":"#70AD47",


                },


            )


            fig_time.update_traces(


                hovertemplate="<b>%{fullData.name}</b><br>"


                              "📅 <b>%{x|%d/%m/%Y}</b><br>"


                              "Avaliações: <b>%{y}</b><extra></extra>",


                line=dict(width=2.5), marker=dict(size=7),


            )


            fig_time.update_layout(
                height=380, title_x=0.5, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                xaxis_title="", yaxis_title="Avaliações encerradas",
                showlegend=True,
                legend=dict(
                    orientation="h",
                    yanchor="top",
                    y=-0.45,
                    xanchor="center",
                    x=0.5,
                    font=dict(color="#e5dccb"),
                    bgcolor="rgba(0,0,0,0)"
                ),
                hovermode="x unified",
                title_font=dict(color="#9b8a5c")
            )


            fig_time.update_xaxes(showgrid=False, tickformat="%d/%m/%Y",


                                   rangeslider_visible=True)


            fig_time.update_yaxes(showgrid=True, gridcolor="#2a2a2a")


            st.plotly_chart(fig_time, use_container_width=True)


        else:


            st.info("Sem dados de datas disponíveis para o gráfico de linha do tempo.")





# ══════════════════════════════════════════════════════════════════════════════


# TAB 2 — DADOS GERAIS


# ══════════════════════════════════════════════════════════════════════════════


if active_page == "Dados Gerais":


    st.markdown(f"### 📋 Dados Gerais — {fmt_num(len(df))} avaliações")


    cols_d = [


        # Avaliado


        "nrPM (Avaliado)", "Posto/Grad. (Avaliado)", "Nome (Avaliado)",


        "Unidade RPM (Avaliado)", "Unidade Principal (Avaliado)", "Local/Unidade (Avaliado)",


        "Situação Funcional",


        # Datas e Certificação (sem notas/conceitos)


        "Data AV1", "Data AV2", "Data HOM", "Certificação Homologador",


        # Avaliador 1


        "nrPM (Av1)", "Posto (Av1)", "Nome (Av1)", "RPM (Av1)", "Unid. Principal (Av1)",


        # Avaliador 2


        "nrPM (Av2)", "Posto (Av2)", "Nome (Av2)", "RPM (Av2)", "Unid. Principal (Av2)",


        # Homologador


        "nrPM (Hom)", "Posto (Hom)", "Nome (Hom)", "RPM (Hom)", "Unid. Principal (Hom)",


        # Status


        "Situação Comissão", "Status Avaliação",


    ]


    cols_d = [c for c in cols_d if c in df.columns]


    safe_df(df[cols_d].style.map(color_status,subset=["Status Avaliação"])


                            .map(color_sit,   subset=["Situação Comissão"]), height=540)


    csv_d = df[cols_d].to_csv(index=False, sep=";", encoding="utf-8-sig")


    st.download_button("⬇️ Baixar dados filtrados (CSV)", csv_d.encode("utf-8-sig"),


                        f"avaliacoes_filtradas_{now_br().strftime('%Y%m%d_%H%M')}.csv",


                        mime="text/csv")





# ══════════════════════════════════════════════════════════════════════════════


# TAB 3 — AVALIAÇÕES PENDENTES


# ══════════════════════════════════════════════════════════════════════════════


if active_page == "Avaliações Pendentes":


    st.markdown("### ⏳ Avaliações Pendentes")


    STATUS_PEND = {"Homologação","Parcialmente Encerrada","Aberta"}


    df_pend = df[df["Status Avaliação"].isin(STATUS_PEND)].copy()





    c1, c2 = st.columns(2)


    with c1:


        tipo_pend = st.multiselect("Status:", ["Aberta","Parcialmente Encerrada","Homologação"],


                                   default=["Aberta","Parcialmente Encerrada","Homologação"],


                                   key="tp")


    with c2:


        sc_pend = st.multiselect("Situação:", ["Comissão Atual","Nota Provisória"],


                                  default=["Comissão Atual","Nota Provisória"], key="sp")





    df_pv = df_pend[df_pend["Status Avaliação"].isin(tipo_pend) &


                    df_pend["Situação Comissão"].isin(sc_pend)] if tipo_pend and sc_pend else df_pend





    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.markdown(f'<div class="kpi-card kpi-aberta">'
                    '<div class="label">🔴 Abertas</div>'
                    f'<div class="value">{fmt_num((df_pv["Status Avaliação"]=="Aberta").sum())}</div>'
                    '<div class="sub">AV1 pendente</div>'
                    '</div>', unsafe_allow_html=True)
    with k2:
        st.markdown(f'<div class="kpi-card kpi-parc">'
                    '<div class="label">🟠 Parc. Encerradas</div>'
                    f'<div class="value">{fmt_num((df_pv["Status Avaliação"]=="Parcialmente Encerrada").sum())}</div>'
                    '<div class="sub">AV2 pendente</div>'
                    '</div>', unsafe_allow_html=True)
    with k3:
        st.markdown(f'<div class="kpi-card kpi-hom">'
                    '<div class="label">🟡 Homologação</div>'
                    f'<div class="value">{fmt_num((df_pv["Status Avaliação"]=="Homologação").sum())}</div>'
                    '<div class="sub">Homologação pendente</div>'
                    '</div>', unsafe_allow_html=True)
    with k4:
        st.markdown(f'<div class="kpi-card kpi-total">'
                    '<div class="label">📊 Total</div>'
                    f'<div class="value">{fmt_num(len(df_pv))}</div>'
                    '<div class="sub">Registros filtrados</div>'
                    '</div>', unsafe_allow_html=True)





    cols_pend = [


        # ── Avaliado ──────────────────────────────────────────────────────────


        "nrPM (Avaliado)", "Posto/Grad. (Avaliado)", "Nome (Avaliado)",


        "Unidade RPM (Avaliado)", "Unidade Principal (Avaliado)",


        "Situação Funcional",


        # ── Status e Datas (sem notas/conceitos) ─────────────────────────────


        "Status Avaliação", "Situação Comissão",


        "Data AV1", "Data AV2", "Data HOM", "Certificação Homologador",


        # ── Avaliador 1 ───────────────────────────────────────────────────────


        "nrPM (Av1)", "Posto (Av1)", "Nome (Av1)",


        "RPM (Av1)", "Unid. Principal (Av1)",


        # ── Avaliador 2 ───────────────────────────────────────────────────────


        "nrPM (Av2)", "Posto (Av2)", "Nome (Av2)",


        "RPM (Av2)", "Unid. Principal (Av2)",


        # ── Homologador ───────────────────────────────────────────────────────


        "nrPM (Hom)", "Posto (Hom)", "Nome (Hom)",


        "RPM (Hom)", "Unid. Principal (Hom)",


    ]


    # Manter apenas colunas que existem no DataFrame


    cols_pend = [c for c in cols_pend if c in df_pv.columns]





    safe_df(df_pv[cols_pend]


            .sort_values(["Unidade RPM (Avaliado)", "Status Avaliação", "Nome (Avaliado)"])


            .reset_index(drop=True)


            .style.map(color_status, subset=["Status Avaliação"])


            .map(color_sit,          subset=["Situação Comissão"]),


            height=520)


    csv_p = df_pv[cols_pend].sort_values(


        ["Unidade RPM (Avaliado)", "Status Avaliação", "Nome (Avaliado)"]


    ).to_csv(index=False, sep=";", encoding="utf-8-sig")


    st.download_button("⬇️ Baixar pendentes (CSV)", csv_p.encode("utf-8-sig"),


                        f"avaliacoes_pendentes_{now_br().strftime('%Y%m%d_%H%M')}.csv",


                        mime="text/csv")








# ══════════════════════════════════════════════════════════════════════════════


# TAB 4 — AVALIADORES PENDENTES


# ══════════════════════════════════════════════════════════════════════════════


if active_page == "Avaliadores Pendentes":


    st.markdown("### 👥 Avaliadores Pendentes")





    if rpm_filter:


        df_ab = df_full[(df_full["Status Avaliação"] == "Aberta") & (df_full["RPM (Av1)"].isin(rpm_filter))]


        df_pe = df_full[(df_full["Status Avaliação"].isin(["Aberta", "Parcialmente Encerrada"])) & (df_full["RPM (Av2)"].isin(rpm_filter))]


        df_hom = df_full[(df_full["Status Avaliação"] == "Homologação") & (df_full["RPM (Hom)"].isin(rpm_filter))]


    else:


        df_ab = df[df["Status Avaliação"] == "Aberta"]


        df_pe = df[df["Status Avaliação"].isin(["Aberta", "Parcialmente Encerrada"])]


        df_hom = df[df["Status Avaliação"] == "Homologação"]

    # Pre-calcular quantitativo de avaliadores/homologadores únicos pendentes (função)
    cnt_av1 = df_ab["nrPM (Av1)"].dropna().astype(str).str.strip().replace("", pd.NA).dropna().nunique()
    cnt_av2 = df_pe["nrPM (Av2)"].dropna().astype(str).str.strip().replace("", pd.NA).dropna().nunique()
    cnt_hom = df_hom["nrPM (Hom)"].dropna().astype(str).str.strip().replace("", pd.NA).dropna().nunique()

    # Renderizar cards de quantitativos funcionais
    col_av1, col_av2, col_av3 = st.columns(3)
    with col_av1:
        st.markdown(f'<div class="kpi-card kpi-total">'
                    '<div class="label">👤 Avaliador 1 (AV1)</div>'
                    f'<div class="value">{fmt_num(cnt_av1)}</div>'
                    '<div class="sub">Avaliadores com pendência de AV1</div>'
                    '</div>', unsafe_allow_html=True)
    with col_av2:
        st.markdown(f'<div class="kpi-card kpi-parc">'
                    '<div class="label">👥 Avaliador 2 (AV2)</div>'
                    f'<div class="value">{fmt_num(cnt_av2)}</div>'
                    '<div class="sub">Avaliadores com pendência de AV2</div>'
                    '</div>', unsafe_allow_html=True)
    with col_av3:
        st.markdown(f'<div class="kpi-card kpi-hom">'
                    '<div class="label">⚖️ Homologador (HOM)</div>'
                    f'<div class="value">{fmt_num(cnt_hom)}</div>'
                    '<div class="sub">Homologadores com pendência de HOM</div>'
                    '</div>', unsafe_allow_html=True)

    st.markdown("<div style='margin-bottom: 20px;'></div>", unsafe_allow_html=True)





    av1 = defaultdict(lambda:{"nome":"","posto":"","rpm":"","unid":"","CA":0,"NP":0})


    for _, r in df_ab.iterrows():


        k = r["nrPM (Av1)"]


        if not k: continue


        av1[k].update(nome=r["Nome (Av1)"],posto=r["Posto (Av1)"],rpm=r["RPM (Av1)"],unid=r["Unid. Principal (Av1)"])


        if r["Situação Comissão"]=="Comissão Atual": av1[k]["CA"]+=1


        else: av1[k]["NP"]+=1





    av2 = defaultdict(lambda:{"nome":"","posto":"","rpm":"","unid":"","CA_ab":0,"CA_pe":0,"NP_ab":0,"NP_pe":0})


    for _, r in df_pe.iterrows():


        k = r["nrPM (Av2)"]


        if not k: continue


        av2[k].update(nome=r["Nome (Av2)"],posto=r["Posto (Av2)"],rpm=r["RPM (Av2)"],unid=r["Unid. Principal (Av2)"])


        is_ca = r["Situação Comissão"]=="Comissão Atual"


        if r["Status Avaliação"]=="Aberta":


            if is_ca: av2[k]["CA_ab"]+=1


            else: av2[k]["NP_ab"]+=1


        else:


            if is_ca: av2[k]["CA_pe"]+=1


            else: av2[k]["NP_pe"]+=1





    # AV1


    st.markdown('<div class="section-hdr">👤 AVALIADOR 1 — Avaliações Em Aberto</div>', unsafe_allow_html=True)


    tb1_rows = [{"Nº PM":k,"Nome":d["nome"],"Posto":d["posto"],"RPM":d["rpm"],


        "Unid. Principal":d["unid"],"CA—Aberta":d["CA"],"NP—Aberta":d["NP"],


        "Total AV1":d["CA"]+d["NP"]} for k,d in av1.items() if d["CA"]+d["NP"]>0]


    tb1 = (pd.DataFrame(tb1_rows).sort_values("Total AV1", ascending=False)


           if tb1_rows else pd.DataFrame())


    k1,k2 = st.columns(2)


    k1.metric("Avaliadores pendentes (AV1)", len(tb1))


    k2.metric("Total avaliações Em Aberto", tb1["Total AV1"].sum() if not tb1.empty else 0)


    if not tb1.empty:


        tb1_disp = tb1.reset_index(drop=True)
        tb1_disp.index = range(1, len(tb1_disp) + 1)
        tb1_disp = clean_none_values(tb1_disp)


        


        import inspect


        sig = inspect.signature(st.dataframe)


        has_select = "on_select" in sig.parameters


        


        selected_pm = None


        selected_nome = None


        


        if has_select:


            st.write("💡 *Dica: Clique em uma linha da tabela abaixo para abrir as avaliações deste avaliador.*")


            event1 = st.dataframe(


                tb1_disp,


                use_container_width=True,


                height=260,


                on_select="rerun",


                selection_mode="single-row",


                key="select_tb1"


            )


            rows1 = event1.get("selection", {}).get("rows", [])


            if rows1:


                idx = rows1[0]


                selected_pm = tb1.iloc[idx]["Nº PM"]


                selected_nome = tb1.iloc[idx]["Nome"]


        else:


            st.dataframe(tb1_disp, use_container_width=True, height=260)


            sel_nome = st.selectbox("🔎 Selecione um Avaliador 1 para ver as avaliações:", ["-- Selecione --"] + list(tb1["Nome"].unique()), key="sel_tb1")


            if sel_nome != "-- Selecione --":


                row_sel = tb1[tb1["Nome"] == sel_nome].iloc[0]


                selected_pm = row_sel["Nº PM"]


                selected_nome = row_sel["Nome"]


                


        if selected_pm:


            df_det1 = df_ab[df_ab["nrPM (Av1)"] == selected_pm].copy()


            st.markdown(f"#### 📋 Avaliações pendentes de AV1: **{selected_nome}** ({selected_pm})")


            cols_det1 = ["nrPM (Avaliado)", "Posto/Grad. (Avaliado)", "Nome (Avaliado)",


                         "Unidade RPM (Avaliado)", "Unidade Principal (Avaliado)",


                         "Status Avaliação", "Situação Comissão", "Data AV1"]


            cols_ok1 = [c for c in cols_det1 if c in df_det1.columns]


            safe_df(df_det1[cols_ok1].reset_index(drop=True).style.map(color_status, subset=["Status Avaliação"]), height=180)


            


            # Botões de download lado a lado para esta lista filtrada específica


            dl1, dl2 = st.columns(2)


            with dl1:


                st.download_button(


                    "⬇️ Baixar esta lista (Excel .xlsx)",


                    df_to_xlsx(df_det1[cols_ok1]),


                    f"pendencias_AV1_{selected_pm}_{now_br().strftime('%Y%m%d_%H%M')}.xlsx",


                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",


                    key="dl_av1_xlsx"


                )


            with dl2:


                st.download_button(


                    "⬇️ Baixar esta lista (CSV)",


                    df_det1[cols_ok1].to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig"),


                    f"pendencias_AV1_{selected_pm}_{now_br().strftime('%Y%m%d_%H%M')}.csv",


                    mime="text/csv",


                    key="dl_av1_csv"


                )


            


        st.download_button("⬇️ AV1 (CSV)", tb1.to_csv(index=False,sep=";",encoding="utf-8-sig").encode("utf-8-sig"),


                            f"av1_{now_br().strftime('%Y%m%d_%H%M')}.csv", mime="text/csv")


    else: st.success("✅ Nenhum AV1 com pendências!")





    # AV2


    st.markdown('<div class="section-hdr">👤 AVALIADOR 2 — Em Aberto + Parcialmente Encerrada</div>', unsafe_allow_html=True)


    tb2_rows = [{"Nº PM":k,"Nome":d["nome"],"Posto":d["posto"],"RPM":d["rpm"],


        "Unid. Principal":d["unid"],"CA—Aberta":d["CA_ab"],"CA—Parc.Enc.":d["CA_pe"],


        "NP—Aberta":d["NP_ab"],"NP—Parc.Enc.":d["NP_pe"],


        "Total AV2":d["CA_ab"]+d["CA_pe"]+d["NP_ab"]+d["NP_pe"]} for k,d in av2.items()


        if d["CA_ab"]+d["CA_pe"]+d["NP_ab"]+d["NP_pe"]>0]


    tb2 = (pd.DataFrame(tb2_rows).sort_values("Total AV2", ascending=False)


           if tb2_rows else pd.DataFrame())


    k1,k2 = st.columns(2)


    k1.metric("Avaliadores pendentes (AV2)", len(tb2))


    k2.metric("Total pendências AV2", tb2["Total AV2"].sum() if not tb2.empty else 0)


    if not tb2.empty:


        tb2_disp = tb2.reset_index(drop=True)
        tb2_disp.index = range(1, len(tb2_disp) + 1)
        tb2_disp = clean_none_values(tb2_disp)


        


        import inspect


        sig = inspect.signature(st.dataframe)


        has_select = "on_select" in sig.parameters


        


        selected_pm2 = None


        selected_nome2 = None


        


        if has_select:


            st.write("💡 *Dica: Clique em uma linha da tabela abaixo para abrir as avaliações deste avaliador.*")


            event2 = st.dataframe(


                tb2_disp,


                use_container_width=True,


                height=260,


                on_select="rerun",


                selection_mode="single-row",


                key="select_tb2"


            )


            rows2 = event2.get("selection", {}).get("rows", [])


            if rows2:


                idx = rows2[0]


                selected_pm2 = tb2.iloc[idx]["Nº PM"]


                selected_nome2 = tb2.iloc[idx]["Nome"]


        else:


            st.dataframe(tb2_disp, use_container_width=True, height=260)


            sel_nome2 = st.selectbox("🔎 Selecione um Avaliador 2 para ver as avaliações:", ["-- Selecione --"] + list(tb2["Nome"].unique()), key="sel_tb2")


            if sel_nome2 != "-- Selecione --":


                row_sel2 = tb2[tb2["Nome"] == sel_nome2].iloc[0]


                selected_pm2 = row_sel2["Nº PM"]


                selected_nome2 = row_sel2["Nome"]


                


        if selected_pm2:


            df_det2 = df_pe[df_pe["nrPM (Av2)"] == selected_pm2].copy()


            st.markdown(f"#### 📋 Avaliações pendentes de AV2: **{selected_nome2}** ({selected_pm2})")


            cols_det2 = ["nrPM (Avaliado)", "Posto/Grad. (Avaliado)", "Nome (Avaliado)",


                         "Unidade RPM (Avaliado)", "Unidade Principal (Avaliado)",


                         "Status Avaliação", "Situação Comissão", "Data AV1", "Data AV2"]


            cols_ok2 = [c for c in cols_det2 if c in df_det2.columns]


            safe_df(df_det2[cols_ok2].reset_index(drop=True).style.map(color_status, subset=["Status Avaliação"]), height=180)


            


            # Botões de download lado a lado para esta lista filtrada específica


            dl1, dl2 = st.columns(2)


            with dl1:


                st.download_button(


                    "⬇️ Baixar esta lista (Excel .xlsx)",


                    df_to_xlsx(df_det2[cols_ok2]),


                    f"pendencias_AV2_{selected_pm2}_{now_br().strftime('%Y%m%d_%H%M')}.xlsx",


                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",


                    key="dl_av2_xlsx"


                )


            with dl2:


                st.download_button(


                    "⬇️ Baixar esta lista (CSV)",


                    df_det2[cols_ok2].to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig"),


                    f"pendencias_AV2_{selected_pm2}_{now_br().strftime('%Y%m%d_%H%M')}.csv",


                    mime="text/csv",


                    key="dl_av2_csv"


                )


            


        st.download_button("⬇️ AV2 (CSV)", tb2.to_csv(index=False,sep=";",encoding="utf-8-sig").encode("utf-8-sig"),


                            f"av2_{now_br().strftime('%Y%m%d_%H%M')}.csv", mime="text/csv")


    else: st.success("✅ Nenhum AV2 com pendências!")





    # HOMOLOGADOR


    st.markdown(


        '<div class="section-hdr-hom">🏛️ HOMOLOGADOR — Avaliações com Divergência Aguardando Nota de Homologação</div>',


        unsafe_allow_html=True)








    # ── Tabela agregada por Homologador (mesmo modelo AV1/AV2) ──────────────────


    hom_map = defaultdict(lambda: {"nome":"","posto":"","rpm":"","unid":"","CA":0,"NP":0})


    for _, r in df_hom.iterrows():


        k = str(r.get("nrPM (Hom)", "")).strip() or "Não identificado"


        hom_map[k]["nome"]  = r.get("Nome (Hom)", "")


        hom_map[k]["posto"] = r.get("Posto (Hom)", "")


        hom_map[k]["rpm"]   = r.get("RPM (Hom)", "")


        hom_map[k]["unid"]  = r.get("Unid. Principal (Hom)", "")


        if r["Situação Comissão"] == "Comissão Atual":


            hom_map[k]["CA"] += 1


        else:


            hom_map[k]["NP"] += 1





    tb3_rows = [{"Nº PM": k, "Nome": d["nome"], "Posto": d["posto"],


                 "RPM": d["rpm"], "Unid. Principal": d["unid"],


                 "CA—Hom.Pend.": d["CA"], "NP—Hom.Pend.": d["NP"],


                 "Total HOM": d["CA"] + d["NP"]}


                for k, d in hom_map.items() if d["CA"] + d["NP"] > 0]


    tb3 = (pd.DataFrame(tb3_rows).sort_values("Total HOM", ascending=False)


           if tb3_rows else pd.DataFrame())





    k1, k2, k3 = st.columns(3)


    k1.metric("Homologadores com pendência", len(tb3))


    k2.metric("Total aguardando HOM", len(df_hom))


    k3.metric("CA / NP pendentes",


              f"{(df_hom['Situação Comissão']=='Comissão Atual').sum()} / "


              f"{(df_hom['Situação Comissão']=='Nota Provisória').sum()}")





    if not tb3.empty:


        tb3_disp = tb3.reset_index(drop=True)
        tb3_disp.index = range(1, len(tb3_disp) + 1)
        tb3_disp = clean_none_values(tb3_disp)


        


        import inspect


        sig = inspect.signature(st.dataframe)


        has_select = "on_select" in sig.parameters


        


        selected_pm3 = None


        selected_nome3 = None


        


        if has_select:


            st.write("💡 *Dica: Clique em uma linha da tabela abaixo para abrir as avaliações deste homologador.*")


            event3 = st.dataframe(


                tb3_disp,


                use_container_width=True,


                height=280,


                on_select="rerun",


                selection_mode="single-row",


                key="select_tb3"


            )


            rows3 = event3.get("selection", {}).get("rows", [])


            if rows3:


                idx = rows3[0]


                selected_pm3 = tb3.iloc[idx]["Nº PM"]


                selected_nome3 = tb3.iloc[idx]["Nome"]


        else:


            st.dataframe(tb3_disp, use_container_width=True, height=280)


            sel_nome3 = st.selectbox("🔎 Selecione um Homologador para ver as avaliações:", ["-- Selecione --"] + list(tb3["Nome"].unique()), key="sel_tb3")


            if sel_nome3 != "-- Selecione --":


                row_sel3 = tb3[tb3["Nome"] == sel_nome3].iloc[0]


                selected_pm3 = row_sel3["Nº PM"]


                selected_nome3 = row_sel3["Nome"]


                


        if selected_pm3:


            df_det3 = df_hom[df_hom["nrPM (Hom)"] == selected_pm3].copy()


            st.markdown(f"#### 📋 Avaliações pendentes do Homologador: **{selected_nome3}** ({selected_pm3})")


            cols_det3 = ["nrPM (Avaliado)", "Posto/Grad. (Avaliado)", "Nome (Avaliado)",


                         "Unidade RPM (Avaliado)", "Unidade Principal (Avaliado)",


                         "Status Avaliação", "Situação Comissão", "Data AV1", "Data AV2", "Data HOM"]


            cols_ok3 = [c for c in cols_det3 if c in df_det3.columns]


            safe_df(df_det3[cols_ok3].reset_index(drop=True).style.map(color_status, subset=["Status Avaliação"]), height=180)


            


            # Botões de download lado a lado para esta lista filtrada específica


            dl1, dl2 = st.columns(2)


            with dl1:


                st.download_button(


                    "⬇️ Baixar esta lista (Excel .xlsx)",


                    df_to_xlsx(df_det3[cols_ok3]),


                    f"pendencias_HOM_{selected_pm3}_{now_br().strftime('%Y%m%d_%H%M')}.xlsx",


                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",


                    key="dl_hom_xlsx"


                )


            with dl2:


                st.download_button(


                    "⬇️ Baixar esta lista (CSV)",


                    df_det3[cols_ok3].to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig"),


                    f"pendencias_HOM_{selected_pm3}_{now_br().strftime('%Y%m%d_%H%M')}.csv",


                    mime="text/csv",


                    key="dl_hom_csv"


                )


            


        st.download_button(


            "⬇️ Homologadores pendentes (CSV)",


            tb3.to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig"),


            f"hom_{now_br().strftime('%Y%m%d_%H%M')}.csv", mime="text/csv")


    else:


        st.success("✅ Nenhum Homologador com pendências!")





    # ── Lista detalhada das avaliações em Homologação ──────────────────────────


    if not df_hom.empty:


        with st.expander(f"📋 Ver {len(df_hom)} avaliação(ões) pendentes de homologação", expanded=False):


            cols_det = ["nrPM (Avaliado)", "Posto/Grad. (Avaliado)", "Nome (Avaliado)",


                        "Unidade RPM (Avaliado)", "Unidade Principal (Avaliado)",


                        "Status Avaliação", "Situação Comissão",


                        "Data AV1", "Data AV2", "Data HOM", "Certificação Homologador",


                        "nrPM (Hom)", "Posto (Hom)", "Nome (Hom)", "RPM (Hom)",


                        "Unid. Principal (Hom)"]


            cols_ok = [c for c in cols_det if c in df_hom.columns]


            safe_df(


                df_hom[cols_ok]


                .sort_values(["Unidade RPM (Avaliado)", "Nome (Avaliado)"])


                .reset_index(drop=True)


                .style.map(color_sit, subset=["Situação Comissão"]),


                height=340)


            st.download_button(


                "⬇️ Lista detalhada (CSV)",


                df_hom[cols_ok].to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig"),


                f"hom_detalhe_{now_br().strftime('%Y%m%d_%H%M')}.csv", mime="text/csv")





# ══════════════════════════════════════════════════════════════════════════════


# TAB 5 — GERAR RELATÓRIO (geração 100% em memória — funciona local e na nuvem)


# ══════════════════════════════════════════════════════════════════════════════





# ── Motor de geração de Excel em memória ─────────────────────────────────────


STATUS_BG_XL = {


    "Encerrada":             "70AD47",


    "Homologação":           "FFD966",


    "Parcialmente Encerrada":"FF8C00",


    "Aberta":                "FF4444",


}


SIT_BG_XL = {"Comissão Atual": "4472C4", "Nota Provisória": "FFC000"}





# Colunas exportadas para Excel (SEM Conceito Geral, Nota Geral, Nota Homologação)


COLS_XLS = [


    "nrPM (Avaliado)", "Posto/Grad. (Avaliado)", "Nome (Avaliado)",


    "Unidade RPM (Avaliado)", "Unidade Principal (Avaliado)", "Local/Unidade (Avaliado)",


    "Situação Funcional",


    "Data AV1", "Data AV2", "Data HOM", "Certificação Homologador",


    "nrPM (Av1)", "Posto (Av1)", "Nome (Av1)", "RPM (Av1)", "Unid. Principal (Av1)",


    "nrPM (Av2)", "Posto (Av2)", "Nome (Av2)", "RPM (Av2)", "Unid. Principal (Av2)",


    "nrPM (Hom)", "Posto (Hom)", "Nome (Hom)", "RPM (Hom)", "Unid. Principal (Hom)",


    "Situação Comissão", "Status Avaliação",


]








def _xl_styles():


    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


    thin = Side(border_style="thin", color="CCCCCC")


    return {


        "hdr_fill":  PatternFill("solid", fgColor="1F3864"),


        "hdr_font":  Font(bold=True, color="FFFFFF", name="Calibri", size=10),


        "hdr_al":    Alignment(horizontal="center", vertical="center", wrap_text=True),


        "brd":       Border(left=thin, right=thin, top=thin, bottom=thin),


        "data_font": Font(name="Calibri", size=9),


        "center":    Alignment(horizontal="center", vertical="center"),


        "left":      Alignment(vertical="center"),


        "title_font":Font(bold=True, color="FFFFFF", name="Calibri", size=13),


        "title_al":  Alignment(horizontal="center", vertical="center"),


    }








def _write_title(ws, titulo, n_cols, s):


    from openpyxl.styles import PatternFill


    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(n_cols, 1))


    c = ws.cell(1, 1, titulo)


    c.fill = PatternFill("solid", fgColor="1F3864")


    c.font = s["title_font"]; c.alignment = s["title_al"]


    ws.row_dimensions[1].height = 22








def _write_headers(ws, cols, s):


    from openpyxl.utils import get_column_letter


    for i, col in enumerate(cols, 1):


        c = ws.cell(2, i, col)


        c.fill = s["hdr_fill"]; c.font = s["hdr_font"]


        c.alignment = s["hdr_al"]; c.border = s["brd"]


    ws.row_dimensions[2].height = 28


    ws.freeze_panes = "A3"


    return len(cols)








def _write_data_rows(ws, df, cols, s):


    from openpyxl.styles import PatternFill, Font


    for r, (_, row) in enumerate(df.iterrows(), 3):


        for ci, col in enumerate(cols, 1):


            val = row.get(col, "")


            cell = ws.cell(r, ci, str(val) if pd.notna(val) and val != "" else "")


            cell.font = s["data_font"]; cell.border = s["brd"]


            cell.alignment = s["center"] if ci > 7 else s["left"]


            if col == "Status Avaliação" and val in STATUS_BG_XL:


                cell.fill = PatternFill("solid", fgColor=STATUS_BG_XL[val])


                cell.font = Font(bold=True, name="Calibri", size=9,


                                  color="FFFFFF" if val == "Aberta" else "000000")


            elif col == "Situação Comissão" and val in SIT_BG_XL:


                cell.fill = PatternFill("solid", fgColor=SIT_BG_XL[val])


                cell.font = Font(bold=True, name="Calibri", size=9,


                                  color="FFFFFF" if val == "Comissão Atual" else "000000")








def _auto_widths(ws, df, cols):


    from openpyxl.utils import get_column_letter


    for ci, col in enumerate(cols, 1):


        max_len = 0


        if col in df.columns and not df.empty:


            max_val = df[col].astype(str).str.len().max()


            if pd.notna(max_val):


                max_len = int(max_val)


        max_len = max(len(str(col)), max_len)


        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len * 0.92, 8), 40)








def _write_data_sheet(ws, df, titulo, cols, s):


    """Escreve título + cabeçalhos + dados em uma aba."""


    df_c = df[[c for c in cols if c in df.columns]].reset_index(drop=True)


    actual_cols = list(df_c.columns)


    _write_title(ws, titulo, len(actual_cols), s)


    _write_headers(ws, actual_cols, s)


    _write_data_rows(ws, df_c, actual_cols, s)


    _auto_widths(ws, df_c, actual_cols)








def _write_avaliadores_sheet(ws, df_unit, s, df_global=None, titulo_unidade=""):


    """Aba Avaliadores Pendentes: avaliadores lotados na unidade com pendências."""


    from openpyxl.styles import PatternFill, Font, Alignment


    _write_title(ws, "AVALIADORES PENDENTES — LOTADOS NA UNIDADE", 14, s)





    is_geral = "GERAL" in str(titulo_unidade).upper()





    if df_global is not None and not is_geral:


        df_ab = df_global[(df_global["Status Avaliação"] == "Aberta") & (df_global["RPM (Av1)"] == titulo_unidade)]


        df_pe = df_global[(df_global["Status Avaliação"].isin(["Aberta", "Parcialmente Encerrada"])) & (df_global["RPM (Av2)"] == titulo_unidade)]


        df_hom = df_global[(df_global["Status Avaliação"] == "Homologação") & (df_global["RPM (Hom)"] == titulo_unidade)]


    else:


        df_ab = df_unit[df_unit["Status Avaliação"] == "Aberta"]


        df_pe = df_unit[df_unit["Status Avaliação"].isin(["Aberta", "Parcialmente Encerrada"])]


        df_hom = df_unit[df_unit["Status Avaliação"] == "Homologação"]





    av1 = defaultdict(lambda: {"nome":"","posto":"","rpm":"","unid":"","CA":0,"NP":0})


    for _, r in df_ab.iterrows():


        k = str(r.get("nrPM (Av1)","")).strip()


        if not k: continue


        av1[k].update(nome=r.get("Nome (Av1)",""), posto=r.get("Posto (Av1)",""),


                      rpm=r.get("RPM (Av1)",""), unid=r.get("Unid. Principal (Av1)",""))


        if r["Situação Comissão"] == "Comissão Atual": av1[k]["CA"] += 1


        else: av1[k]["NP"] += 1





    av2 = defaultdict(lambda: {"nome":"","posto":"","rpm":"","unid":"","CA_ab":0,"CA_pe":0,"NP_ab":0,"NP_pe":0})


    for _, r in df_pe.iterrows():


        k = str(r.get("nrPM (Av2)","")).strip()


        if not k: continue


        av2[k].update(nome=r.get("Nome (Av2)",""), posto=r.get("Posto (Av2)",""),


                      rpm=r.get("RPM (Av2)",""), unid=r.get("Unid. Principal (Av2)",""))


        is_ca = r["Situação Comissão"] == "Comissão Atual"


        if r["Status Avaliação"] == "Aberta":


            if is_ca: av2[k]["CA_ab"] += 1


            else: av2[k]["NP_ab"] += 1


        else:


            if is_ca: av2[k]["CA_pe"] += 1


            else: av2[k]["NP_pe"] += 1





    hom = defaultdict(lambda: {"nome":"","posto":"","rpm":"","unid":"","CA":0,"NP":0})


    for _, r in df_hom.iterrows():


        k = str(r.get("nrPM (Hom)","")).strip() or "N/I"


        hom[k].update(nome=r.get("Nome (Hom)",""), posto=r.get("Posto (Hom)",""),


                      rpm=r.get("RPM (Hom)",""), unid=r.get("Unid. Principal (Hom)",""))


        if r["Situação Comissão"] == "Comissão Atual": hom[k]["CA"] += 1


        else: hom[k]["NP"] += 1





    row_num = 3


    # Cabeçalho seção AV1


    titles_av1 = ["Nº PM","Posto","Nome","RPM","Unidade","CA—Aberta","NP—Aberta","Total AV1"]


    fill_av1 = PatternFill("solid", fgColor="1F3864")


    for ci, h in enumerate(titles_av1, 1):


        c = ws.cell(row_num, ci, h)


        c.fill = fill_av1; c.font = s["hdr_font"]; c.alignment = s["hdr_al"]; c.border = s["brd"]


    ws.merge_cells(start_row=row_num-1, start_column=1, end_row=row_num-1, end_column=8)


    lbl = ws.cell(row_num-1, 1, "AVALIADOR 1 — Avaliações Em Aberto")


    lbl.fill = PatternFill("solid", fgColor="2E5090"); lbl.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)


    lbl.alignment = Alignment(horizontal="center", vertical="center")


    row_num += 1


    for k, d in sorted(av1.items(), key=lambda x: -(x[1]["CA"]+x[1]["NP"])):


        tot = d["CA"]+d["NP"]


        if tot == 0: continue


        for ci, val in enumerate([k, d["posto"], d["nome"], d["rpm"], d["unid"],


                                    d["CA"], d["NP"], tot], 1):


            c = ws.cell(row_num, ci, val)


            c.font = s["data_font"]; c.border = s["brd"]; c.alignment = s["center"]


        row_num += 1





    row_num += 2


    # Cabeçalho seção AV2


    titles_av2 = ["Nº PM","Posto","Nome","RPM","Unidade","CA—Ab.","CA—PE","NP—Ab.","NP—PE","Total AV2"]


    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)


    lbl2 = ws.cell(row_num, 1, "AVALIADOR 2 — Em Aberto + Parcialmente Encerrada")


    lbl2.fill = PatternFill("solid", fgColor="2E5090"); lbl2.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)


    lbl2.alignment = Alignment(horizontal="center", vertical="center")


    row_num += 1


    for ci, h in enumerate(titles_av2, 1):


        c = ws.cell(row_num, ci, h)


        c.fill = fill_av1; c.font = s["hdr_font"]; c.alignment = s["hdr_al"]; c.border = s["brd"]


    row_num += 1


    for k, d in sorted(av2.items(), key=lambda x: -(x[1]["CA_ab"]+x[1]["CA_pe"]+x[1]["NP_ab"]+x[1]["NP_pe"])):


        tot = d["CA_ab"]+d["CA_pe"]+d["NP_ab"]+d["NP_pe"]


        if tot == 0: continue


        for ci, val in enumerate([k, d["posto"], d["nome"], d["rpm"], d["unid"],


                                    d["CA_ab"], d["CA_pe"], d["NP_ab"], d["NP_pe"], tot], 1):


            c = ws.cell(row_num, ci, val)


            c.font = s["data_font"]; c.border = s["brd"]; c.alignment = s["center"]


        row_num += 1





    row_num += 2


    # Cabeçalho seção HOM


    titles_hom = ["Nº PM","Posto","Nome","RPM","Unidade","CA—Pend.","NP—Pend.","Total HOM"]


    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)


    lbl3 = ws.cell(row_num, 1, "HOMOLOGADOR — Avaliações com Divergência Aguardando Nota")


    lbl3.fill = PatternFill("solid", fgColor="7B3F00"); lbl3.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)


    lbl3.alignment = Alignment(horizontal="center", vertical="center")


    row_num += 1


    for ci, h in enumerate(titles_hom, 1):


        c = ws.cell(row_num, ci, h)


        c.fill = PatternFill("solid", fgColor="7B3F00"); c.font = s["hdr_font"]


        c.alignment = s["hdr_al"]; c.border = s["brd"]


    row_num += 1


    for k, d in sorted(hom.items(), key=lambda x: -(x[1]["CA"]+x[1]["NP"])):


        tot = d["CA"]+d["NP"]


        if tot == 0: continue


        for ci, val in enumerate([k, d["posto"], d["nome"], d["rpm"], d["unid"],


                                    d["CA"], d["NP"], tot], 1):


            c = ws.cell(row_num, ci, val)


            c.font = s["data_font"]; c.border = s["brd"]; c.alignment = s["center"]


        row_num += 1





    from openpyxl.utils import get_column_letter


    for ci in range(1, 15):


        ws.column_dimensions[get_column_letter(ci)].width = 18








def _write_resumo_sheet(ws, df, titulo, s):


    """Aba Resumo: grade CA × Status e NP × Status com cores."""


    from openpyxl.styles import PatternFill, Font, Alignment


    _write_title(ws, f"RESUMO — {titulo}", 5, s)





    STATUS_ORD = ["Aberta", "Parcialmente Encerrada", "Homologação", "Encerrada"]


    ca = df[df["Situação Comissão"] == "Comissão Atual"]


    np_ = df[df["Situação Comissão"] == "Nota Provisória"]





    fill_ca  = PatternFill("solid", fgColor="4472C4")   # azul  — Comissão Atual


    fill_np  = PatternFill("solid", fgColor="FFC000")   # amarelo — Nota Provisória


    fill_tot = PatternFill("solid", fgColor="1F3864")   # azul escuro — Total


    fill_st  = {


        "Aberta":                 PatternFill("solid", fgColor="FF4444"),


        "Parcialmente Encerrada": PatternFill("solid", fgColor="FF8C00"),


        "Homologação":            PatternFill("solid", fgColor="FFD966"),


        "Encerrada":              PatternFill("solid", fgColor="70AD47"),


    }


    white_bold  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)


    black_bold  = Font(bold=True, color="000000", name="Calibri", size=11)


    center_bold = Alignment(horizontal="center", vertical="center")


    thin = s["brd"]





    r = 3


    # Cabeçalho da tabela


    headers = ["STATUS / SITUAÇÃO", "COMISSÃO ATUAL ✅", "NOTA PROVISÓRIA ⚠️",


               "TOTAL STATUS", "% do Total"]


    for ci, h in enumerate(headers, 1):


        cell = ws.cell(r, ci, h)


        cell.fill = fill_tot; cell.font = white_bold


        cell.alignment = center_bold; cell.border = thin


        ws.column_dimensions[ws.cell(r, ci).column_letter].width = 26 if ci == 1 else 18


    ws.row_dimensions[r].height = 30


    r += 1





    total = len(df)


    for st in STATUS_ORD:


        ca_n  = (ca["Status Avaliação"] == st).sum()


        np_n  = (np_["Status Avaliação"] == st).sum()


        tot_n = ca_n + np_n


        pct   = f"{tot_n/total*100:.1f}%" if total > 0 else "0%"





        # Coluna Status


        c0 = ws.cell(r, 1, st)


        c0.fill = fill_st.get(st, PatternFill()); c0.border = thin


        c0.font = black_bold if st in ("Homologação","Parcialmente Encerrada","Encerrada") else white_bold


        c0.alignment = center_bold





        # CA


        c1 = ws.cell(r, 2, ca_n)


        c1.fill = fill_ca; c1.font = white_bold; c1.alignment = center_bold; c1.border = thin





        # NP


        c2 = ws.cell(r, 3, np_n)


        c2.fill = fill_np; c2.font = black_bold; c2.alignment = center_bold; c2.border = thin





        # Total


        c3 = ws.cell(r, 4, tot_n)


        c3.fill = fill_tot; c3.font = white_bold; c3.alignment = center_bold; c3.border = thin





        # %


        c4 = ws.cell(r, 5, pct)


        c4.font = Font(name="Calibri", size=11); c4.alignment = center_bold; c4.border = thin





        ws.row_dimensions[r].height = 24


        r += 1





    # Linha TOTAL GERAL


    ca_total  = len(ca); np_total  = len(np_)


    for ci, val in enumerate(["TOTAL GERAL", ca_total, np_total, total, "100%"], 1):


        c = ws.cell(r, ci, val)


        c.fill = fill_tot; c.font = white_bold; c.alignment = center_bold; c.border = thin


    ws.row_dimensions[r].height = 28





    # Legenda


    r += 2


    for ci, (txt, fill, fnt) in enumerate([


        ("COMISSÃO ATUAL — Policial está lotado na unidade avaliadora (CA)", fill_ca, white_bold),


        ("NOTA PROVISÓRIA — Policial transferido; nota pode mudar (NP)",     fill_np, black_bold),


    ], 1):


        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)


        lc = ws.cell(r, 1, txt)


        lc.fill = fill; lc.font = fnt


        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)


        ws.row_dimensions[r].height = 20


        r += 1








def _write_analise_sheet(ws, df, titulo, s):


    from openpyxl.styles import PatternFill, Font, Alignment


    from openpyxl.chart import PieChart3D, Reference


    from openpyxl.chart.series import DataPoint





    _write_title(ws, f"ANÁLISE — {titulo}", 6, s)





    STATUS_ORD = ["Aberta", "Parcialmente Encerrada", "Homologação", "Encerrada"]


    ca  = df[df["Situação Comissão"] == "Comissão Atual"]


    np_ = df[df["Situação Comissão"] == "Nota Provisória"]


    total = len(df)





    fill_ca  = PatternFill("solid", fgColor="4472C4")


    fill_np  = PatternFill("solid", fgColor="FFC000")


    fill_tot = PatternFill("solid", fgColor="1F3864")


    fill_st  = {


        "Aberta":                 PatternFill("solid", fgColor="FF4444"),


        "Parcialmente Encerrada": PatternFill("solid", fgColor="FF8C00"),


        "Homologação":            PatternFill("solid", fgColor="FFD966"),


        "Encerrada":              PatternFill("solid", fgColor="70AD47"),


    }


    white_bold  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)


    black_bold  = Font(bold=True, color="000000", name="Calibri", size=11)


    center_al   = Alignment(horizontal="center", vertical="center")


    thin = s["brd"]





    # ── Cabeçalho da tabela ────────────────────────────────────────────────────


    r = 3


    headers = ["STATUS", "COMISSÃO ATUAL", "NOTA PROVISÓRIA", "TOTAL", "%"]


    col_fills = [fill_tot, fill_ca, fill_np, fill_tot, fill_tot]


    col_fonts = [white_bold, white_bold, black_bold, white_bold, white_bold]


    for ci, (h, fll, fnt) in enumerate(zip(headers, col_fills, col_fonts), 1):


        c = ws.cell(r, ci, h)


        c.fill = fll; c.font = fnt; c.alignment = center_al; c.border = thin


        ws.column_dimensions[ws.cell(r, ci).column_letter].width = 26 if ci == 1 else 18


    ws.row_dimensions[r].height = 28


    r += 1





    # ── Dados por Status ───────────────────────────────────────────────────────


    data_start_row = r  # usado pelo gráfico


    for st in STATUS_ORD:


        ca_n  = int((ca["Status Avaliação"] == st).sum())


        np_n  = int((np_["Status Avaliação"] == st).sum())


        tot_n = ca_n + np_n


        pct   = f"{tot_n/total*100:.1f}%" if total > 0 else "0%"





        fll_st = fill_st.get(st, PatternFill())


        use_white = st == "Aberta"





        c0 = ws.cell(r, 1, st)


        c0.fill = fll_st; c0.border = thin


        c0.font = white_bold if use_white else black_bold


        c0.alignment = center_al





        c1 = ws.cell(r, 2, ca_n)


        c1.fill = fill_ca; c1.font = white_bold; c1.alignment = center_al; c1.border = thin





        c2 = ws.cell(r, 3, np_n)


        c2.fill = fill_np; c2.font = black_bold; c2.alignment = center_al; c2.border = thin





        c3 = ws.cell(r, 4, tot_n)


        c3.fill = fill_tot; c3.font = white_bold; c3.alignment = center_al; c3.border = thin





        c4 = ws.cell(r, 5, pct)


        c4.font = Font(name="Calibri", size=11); c4.alignment = center_al; c4.border = thin





        ws.row_dimensions[r].height = 22


        r += 1


    data_end_row = r - 1





    # Linha TOTAL GERAL


    ca_tot = len(ca); np_tot = len(np_)


    for ci, val in enumerate(["TOTAL GERAL", ca_tot, np_tot, total, "100%"], 1):


        c = ws.cell(r, ci, val)


        c.fill = fill_tot; c.font = white_bold; c.alignment = center_al; c.border = thin


    ws.row_dimensions[r].height = 26


    total_row = r


    r += 2





    # Legenda


    for txt, fll, fnt in [


        ("COMISSÃO ATUAL — Policial lotado na unidade avaliadora", fill_ca, white_bold),


        ("NOTA PROVISÓRIA — Policial transferido; nota pode mudar", fill_np, black_bold),


    ]:


        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)


        lc = ws.cell(r, 1, txt)


        lc.fill = fll; lc.font = fnt


        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)


        ws.row_dimensions[r].height = 18


        r += 1





    # ── Gráfico 1: Pizza por Status (Total por status - 3D) ─────────────────────────


    pie1 = PieChart3D()


    pie1.title  = "Status das Avaliações"


    pie1.style  = 10


    pie1.width  = 14


    pie1.height = 10





    # Dados: coluna 4 (TOTAL) linhas data_start_row até data_end_row


    data1 = Reference(ws, min_col=4, min_row=data_start_row, max_row=data_end_row)


    cats1 = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)


    pie1.add_data(data1)


    pie1.set_categories(cats1)


    pie1.series[0].title = None





    # Cores manuais para cada fatia (ordem: Aberta, Parc.Enc, Hom, Encerrada)


    SLICE_COLORS = ["FF4444", "FF8C00", "FFD966", "70AD47"]


    for idx, hex_color in enumerate(SLICE_COLORS):


        pt = DataPoint(idx=idx)


        pt.graphicalProperties.solidFill = hex_color


        pie1.series[0].dPt.append(pt)





    # Legenda na lateral direita


    pie1.legend.position = "r"





    from openpyxl.chart.label import DataLabelList


    pie1.dataLabels = DataLabelList()


    pie1.dataLabels.showPercent     = False


    pie1.dataLabels.showCatName     = False


    pie1.dataLabels.showVal         = True


    pie1.dataLabels.showLeaderLines = True





    # Posiciona o gráfico na coluna G linha 3


    ws.add_chart(pie1, "G3")








def _build_workbook(df_unit: pd.DataFrame, titulo: str, df_global: pd.DataFrame = None) -> bytes:


    """Monta workbook completo com 4 abas para uma unidade (sem Resumo duplicado)."""


    from openpyxl import Workbook


    s = _xl_styles()


    wb = Workbook()





    is_geral = "GERAL" in str(titulo).upper()


    cols = [c for c in COLS_XLS if c in df_unit.columns]


    


    if is_geral:


        sensitive_cols = [c for c in [


            "Conceito Geral", "Nota Geral", "Nota Homologação",


            "Competência 1", "Conceito Comp.1", "Nota Comp.1",


            "Competência 2", "Conceito Comp.2", "Nota Comp.2",


            "Competência 3", "Conceito Comp.3", "Nota Comp.3",


            "Competência 4", "Conceito Comp.4", "Nota Comp.4"


        ] if c in df_unit.columns]


        


        if "Data HOM" in cols:


            idx = cols.index("Data HOM") + 1


            cols = cols[:idx] + sensitive_cols + cols[idx:]


        else:


            cols.extend(sensitive_cols)





    # Aba 1 — Geral


    ws1 = wb.active; ws1.title = "Geral"


    _write_data_sheet(ws1, df_unit, f"AVALIAÇÕES — {titulo}", cols, s)





    # Aba 2 — Avaliações Pendentes (Aberta, Parcialmente Encerrada e Homologação)


    ws2 = wb.create_sheet("Avaliações Pendentes")


    df_pend = df_unit[df_unit["Status Avaliação"].isin(["Aberta", "Parcialmente Encerrada", "Homologação"])]


    _write_data_sheet(ws2, df_pend, f"AVALIAÇÕES PENDENTES — {titulo}", cols, s)





    # Aba 3 — Avaliadores Pendentes


    ws3 = wb.create_sheet("Avaliadores Pendentes")


    _write_avaliadores_sheet(ws3, df_unit, s, df_global=df_global, titulo_unidade=titulo)





    # Aba 4 — Análise (tabela + gráficos de pizza)


    ws4 = wb.create_sheet("Análise")


    _write_analise_sheet(ws4, df_unit, titulo, s)





    buf = io.BytesIO()


    wb.save(buf); buf.seek(0)


    return buf.read()








def _gerar_zip_bytes(df_full: pd.DataFrame, modo: str, units_sel: list) -> tuple:
    """Gera ZIP com planilhas Excel em memória e retorna (bytes, filename)."""
    zip_buf  = io.BytesIO()
    zip_name = f"AADP_2026_{now_br().strftime('%Y%m%d_%H%M%S')}.zip"

    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        if modo in ("all", "geral"):
            zf.writestr("Analise_Avaliacoes_Geral.xlsx",
                        _build_workbook(df_full, "GERAL — AADP 2026", df_full))
        if modo in ("all", "units"):
            targets = units_sel if units_sel else sorted(
                df_full["Unidade RPM (Avaliado)"].dropna().unique(), key=rpm_sort_key)
            for rpm in targets:
                mask   = df_full["Unidade RPM (Avaliado)"] == rpm
                df_rpm = df_full[mask].copy()
                safe   = re.sub(r'[^\w]', '_', str(rpm))
                zf.writestr(f"Analise_Avaliacoes_{safe}.xlsx",
                            _build_workbook(df_rpm, rpm, df_full))

    zip_buf.seek(0)
    return zip_buf.read(), zip_name


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — GERAR RELATÓRIO EXCEL
# ══════════════════════════════════════════════════════════════════════════════
if active_page == "Gerar Relatório":
    st.markdown("### 📥 Gerar Relatório Excel")
    
    if active_role in ("P1", "SADM"):
        st.warning("⚠️ Você não possui permissão para acessar esta funcionalidade.")
    else:
        if "excel_modo_rel" not in st.session_state:
            st.session_state.excel_modo_rel = "Completo"
            
        st.markdown("<p style='font-size: 1.1rem; font-weight: bold; margin-bottom: 8px; color: #9b8a5c;'>Tipo de relatório:</p>", unsafe_allow_html=True)
        st.markdown("<div class='excel-scope-marker'></div>", unsafe_allow_html=True)
        
        col_ex1, col_ex2, col_ex3 = st.columns(3)
        with col_ex1:
            is_ex1 = (st.session_state.excel_modo_rel == "Completo")
            btn_ex1_type = "primary" if is_ex1 else "secondary"
            if st.button("🌐\nCompleto\n(Geral + RPMs)", key="btn_excel_scope_completo", use_container_width=True, type=btn_ex1_type):
                st.session_state.excel_modo_rel = "Completo"
                st.rerun()
        with col_ex2:
            is_ex2 = (st.session_state.excel_modo_rel == "Somente Geral")
            btn_ex2_type = "primary" if is_ex2 else "secondary"
            if st.button("📋\nGeral\n(Somente Geral)", key="btn_excel_scope_geral", use_container_width=True, type=btn_ex2_type):
                st.session_state.excel_modo_rel = "Somente Geral"
                st.rerun()
        with col_ex3:
            is_ex3 = (st.session_state.excel_modo_rel == "Unidades RPM específicas")
            btn_ex3_type = "primary" if is_ex3 else "secondary"
            if st.button("🎯\nEspecíficas\n(Filtrar Unidades)", key="btn_excel_scope_especifica", use_container_width=True, type=btn_ex3_type):
                st.session_state.excel_modo_rel = "Unidades RPM específicas"
                st.rerun()
                
        modo_rel = st.session_state.excel_modo_rel
    
        units_sel = []
        if "específicas" in modo_rel:
            all_rpms_sorted = sorted(df_full["Unidade RPM (Avaliado)"].dropna().unique(),
                                      key=rpm_sort_key)
            units_sel = st.multiselect("Selecione as Unidades RPM:", all_rpms_sorted,
                                        placeholder="Escolha uma ou mais unidades...")
            if units_sel:
                n_prev = sum((df_full["Unidade RPM (Avaliado)"]==u).sum() for u in units_sel)
                st.markdown(f"<small>📊 {len(units_sel)} unidade(s) · {fmt_num(n_prev)} registros</small>",
                            unsafe_allow_html=True)
    
        st.markdown("---")
    
        if "específicas" in modo_rel and not units_sel:
            st.warning("⚠️ Selecione ao menos uma Unidade RPM.")
        else:
            if st.button("🚀 Gerar e Baixar Relatório", type="primary", use_container_width=True):
                modo_code = ("geral" if "Somente" in modo_rel
                             else "units" if "específicas" in modo_rel else "all")
                with st.spinner("⏳ Gerando planilhas Excel... aguarde."):
                    try:
                        zip_bytes, zip_name = _gerar_zip_bytes(df_full, modo_code, units_sel)
                        st.success(f"✅ ZIP gerado com sucesso! ({len(zip_bytes)//1024:,} KB)")
                        log_action(st.session_state.user_pm, "EXPORTAR_EXCEL", f"Modo: {modo_code}, Unidades: {units_sel}")
                        st.download_button(
                            label=f"⬇️ Baixar {zip_name}",
                            data=zip_bytes,
                            file_name=zip_name,
                            mime="application/zip",
                            use_container_width=True,
                        )
                    except Exception as ex:
                        st.error(f"❌ Erro na geração: {ex}")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 6 — HOMOLOGAÇÃO RELATÓRIO WORD (.DOCX)
# ══════════════════════════════════════════════════════════════════════════════
if active_page == "Relatório Word":
    st.markdown("### 📄 Relatório Word (.docx)")
    df_word = df_full
                
    if df_word is not None:
        st.markdown("---")
        
        if active_role in ("P1", "SADM"):
            st.warning("⚠️ Você não possui permissão para acessar esta funcionalidade.")
        else:
            st.markdown("<h4 style='font-size: 1.35rem; font-weight: bold; margin-bottom: 12px; color: #9b8a5c;'>Configurações do Relatório</h4>", unsafe_allow_html=True)
            
            if "rel_scope" not in st.session_state:
                st.session_state.rel_scope = "Geral RPM"
                
            st.markdown("<p style='font-size: 1.1rem; font-weight: bold; margin-bottom: 8px; color: #9b8a5c;'>Escopo do Relatório:</p>", unsafe_allow_html=True)
            st.markdown("<div class='report-scope-marker'></div>", unsafe_allow_html=True)
            
            col_sc1, col_sc2, col_sc3 = st.columns(3)
            with col_sc1:
                is_sc1 = (st.session_state.rel_scope == "Geral RPM")
                btn_sc1_type = "primary" if is_sc1 else "secondary"
                if st.button("🏢\nGeral RPM\n(UDI/UDG Principais)", key="btn_scope_geral_rpm", use_container_width=True, type=btn_sc1_type):
                    st.session_state.rel_scope = "Geral RPM"
                    st.rerun()
            with col_sc2:
                is_sc2 = (st.session_state.rel_scope == "Geral Subordinadas")
                btn_sc2_type = "primary" if is_sc2 else "secondary"
                if st.button("🌐\nGeral Subordinadas\n(UDI/UDG + Subordinadas)", key="btn_scope_geral_sub", use_container_width=True, type=btn_sc2_type):
                    st.session_state.rel_scope = "Geral Subordinadas"
                    st.rerun()
            with col_sc3:
                is_sc3 = (st.session_state.rel_scope == "Por RPM específica")
                btn_sc3_type = "primary" if is_sc3 else "secondary"
                if st.button("🎯\nPor RPM específica\n(Filtrar por Unidades)", key="btn_scope_especifica", use_container_width=True, type=btn_sc3_type):
                    st.session_state.rel_scope = "Por RPM específica"
                    st.rerun()
                    
            rel_scope = st.session_state.rel_scope
            
            selected_rpms = []
            if "específica" in rel_scope:
                unique_rpms = sorted(df_word["Unidade RPM (Avaliado)"].dropna().unique().tolist(), key=rpm_sort_key)
                selected_rpms = st.multiselect("Selecione as Unidades UDI/UDG para o relatório:", unique_rpms)
                
            st.markdown("---")
            
            # Validação de botão de geração
            if "específica" in rel_scope and not selected_rpms:
                st.warning("⚠️ Selecione ao menos uma unidade para gerar o relatório.")
            else:
                if st.button("🚀 Gerar e Baixar Relatório Word", key="btn_word_gen"):
                    with st.spinner("⏳ Gerando relatório executivo Word com gráficos... (Isso pode levar alguns instantes)"):
                        try:
                            from gerar_relatorio_word import generate_word_report
                            
                            # Mapear escopo para código
                            mode_code = "geral_rpm" if "Geral RPM" in rel_scope else "geral_subordinadas" if "Geral Subordinadas" in rel_scope else "especifica"
                            
                            doc_bytes = generate_word_report(df_word, mode_code, selected_rpms, user_role=active_role)
                            
                            st.success("✅ Relatório Word gerado com sucesso!")
                            log_action(active_pm, "EXPORTAR_WORD", f"Modo: {mode_code}, RPMs: {selected_rpms}")
                            
                            doc_name = f"Relatorio_Executivo_AADP2026_{now_br().strftime('%Y%m%d_%H%M%S')}.docx"
                            
                            st.download_button(
                                label=f"⬇️ Baixar {doc_name}",
                                data=doc_bytes,
                                file_name=doc_name,
                                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                                use_container_width=True
                            )
                        except Exception as ex:
                            st.error(f"❌ Erro ao gerar o relatório: {ex}")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 7 — PAINEL ADMINISTRADOR
# ══════════════════════════════════════════════════════════════════════════════
if active_page == "Auditoria de Notas" and sidebar_active_role.upper() in ("ADMINISTRADOR", "GESTOR", "P1", "SADM"):
    st.markdown("### 📊 Auditoria de Notas")
    
    _role_audit = sidebar_active_role
    _user_rpm   = st.session_state.get("simulated_rpm", st.session_state.get("user_rpm", "")) if st.session_state.get("simulation_active", False) else st.session_state.get("user_rpm", "")
    _user_unit  = st.session_state.get("simulated_unit", st.session_state.get("user_unit", "")) if st.session_state.get("simulation_active", False) else st.session_state.get("user_unit", "")

    # Obter caminhos dos arquivos locais e Drive
    fonte = cfg.get("fonte_dados", "📁 Pasta local / Servidor")
    drive_master_xlsx_id = cfg.get("drive_master_xlsx_id", "")
    
    if fonte == "☁️ Google Drive":
        master_xlsx_path = os.path.join(str(DADOS_DIR), "Analise avaliacoes completa.xlsx")
        if not drive_master_xlsx_id:
            st.error("❌ ID da Planilha Mestre no Google Drive não configurado!")
            st.warning("⚠️ Configure a chave `drive_master_xlsx_id` nas configurações (st.secrets ou config_aadp.json) para habilitar o download automático da auditoria online.")
            st.stop()
    else:
        # Modo pasta local
        master_xlsx_path = os.path.join(str(Path(DADOS_DIR).parent), "Analise avaliacoes completa.xlsx")
        if not os.path.exists(master_xlsx_path):
            st.error("❌ Arquivo consolidado não encontrado localmente!")
            st.warning(f"Certifique-se de que o arquivo `Analise avaliacoes completa.xlsx` está na pasta raiz do projeto: `{str(Path(DADOS_DIR).parent)}`")
            st.stop()
            
    with st.spinner("Carregando dados da Planilha Mestre de Auditoria..."):
        df_audit, err = load_audit_excel(master_xlsx_path, drive_master_xlsx_id)
        
    if err:
        st.error(f"Erro ao carregar auditoria: {err}")
        st.stop()
        
    # ── Escopo por perfil ─────────────────────────────────────────────────────
    # ADMINISTRADOR / GESTOR → acesso integral (visão de toda a PMMG)
    # P1                     → filtrado pelo RPM do usuário (UDI / UDG)
    # SADM                   → filtrado pela Unidade Principal do usuário
    df_audit_disp = df_audit.copy()

    if _role_audit == "P1":
        # P1 enxerga apenas os avaliados da sua UDI/UDG (RPM)
        if _user_rpm and "Nome RPM" in df_audit_disp.columns:
            df_audit_disp = df_audit_disp[
                df_audit_disp["Nome RPM"].astype(str).str.upper() == str(_user_rpm).upper()
            ]
            st.info(f"🔒 Exibindo apenas registros da sua UDI/UDG: **{_user_rpm}**")
        else:
            st.warning("⚠️ RPM do usuário não identificado. Contate o administrador.")
            st.stop()

    elif _role_audit == "SADM":
        # SADM enxerga apenas os avaliados da sua Unidade Principal
        if _user_unit and "Nome Unidade Principal" in df_audit_disp.columns:
            df_audit_disp = df_audit_disp[
                df_audit_disp["Nome Unidade Principal"].astype(str).str.upper() == str(_user_unit).upper()
            ]
            st.info(f"🔒 Exibindo apenas registros da sua Unidade Principal: **{_user_unit}**")
        else:
            st.warning("⚠️ Unidade Principal do usuário não identificada. Contate o administrador.")
            st.stop()

    else:
        # ADMINISTRADOR / GESTOR: aplicar filtros opcionais da barra lateral
        if rpm_filter:
            df_audit_disp = df_audit_disp[df_audit_disp["Nome RPM"].isin(rpm_filter)]
        if unid_filter:
            df_audit_disp = df_audit_disp[df_audit_disp["Nome Unidade Principal"].isin(unid_filter)]



    # ── CARDS DE RESUMO DA AUDITORIA ──────────────────────────────────────────
    # 1. Total militares com avaliações abertas e sem nota final encerrada
    c1_mask = (df_audit_disp['Todas Avaliações Foram Encerradas?'].astype(str).str.upper() == 'NAO')
    card1_count = len(df_audit_disp[c1_mask])
    
    # 2. Total militares com todas avaliações encerradas e com nota final
    def is_numeric_grade(val):
        try:
            if val is None or str(val).strip() in ("", "-", "None", "nan"):
                return False
            float(str(val).replace(",", "."))
            return True
        except ValueError:
            return False

    c2_mask = (df_audit_disp['Todas Avaliações Foram Encerradas?'].astype(str).str.upper() == 'SIM') & \
               (df_audit_disp['Nota Final - Média Aritmética'].apply(is_numeric_grade))
    card2_count = len(df_audit_disp[c2_mask])
    
    # 3. Média da Nota da Unidade (média aritmética de todas as notas finais já encerradas)
    def get_numeric_value(val):
        try:
            if val is None or str(val).strip() in ("", "-", "None", "nan"):
                return None
            return float(str(val).replace(",", "."))
        except ValueError:
            return None

    grades = df_audit_disp['Nota Final - Média Aritmética'].apply(get_numeric_value).dropna()
    if len(grades) > 0:
        card3_avg = grades.mean()
        import math
        card3_avg_rounded = math.floor(card3_avg * 100 + 0.5) / 100.0
        card3_val = f"{card3_avg_rounded:.2f}".replace(".", ",")
    else:
        card3_val = "0,00"

    # Renderizar os 3 cards lado a lado usando CSS kpi-card
    col_k1, col_k2, col_k3 = st.columns(3)
    with col_k1:
        st.markdown('<div class="kpi-card kpi-aberta">'
                    '<div class="label">AVALIAÇÕES ABERTAS / SEM NOTA FINAL</div>'
                    f'<div class="value">{fmt_num(card1_count)}</div>'
                    '<div class="sub">Militares com pendências de encerramento</div>'
                    '</div>', unsafe_allow_html=True)
    with col_k2:
        st.markdown('<div class="kpi-card kpi-enc">'
                    '<div class="label">AVALIAÇÕES ENCERRADAS COM NOTA</div>'
                    f'<div class="value">{fmt_num(card2_count)}</div>'
                    '<div class="sub">Militares com nota final encerrada</div>'
                    '</div>', unsafe_allow_html=True)
    with col_k3:
        st.markdown('<div class="kpi-card kpi-ca">'
                    '<div class="label">MÉDIA DAS NOTAS FINAIS</div>'
                    f'<div class="value">{card3_val}</div>'
                    '<div class="sub">Média aritmética das notas encerradas</div>'
                    '</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown(f"##### Conteúdo da Planilha Mestre Consolidada ({fmt_num(len(df_audit_disp))} registros)")
    
    # Exibir a planilha de auditoria diretamente!
    safe_df(df_audit_disp, height=540)
    
    # Baixar relatório Excel
    dl_xlsx = df_to_xlsx(df_audit_disp)
    st.download_button(
        "📥 Baixar Resultados Filtrados (Excel .xlsx)",
        dl_xlsx,
        f"Auditoria_Notas_Consolidado_{now_br().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_audit_notes_xlsx",
        use_container_width=True
    )



# ══════════════════════════════════════════════════════════════════════════════
# TAB 7 — DADOS CONSOLIDADOS
# ══════════════════════════════════════════════════════════════════════════════
if active_page == "Dados Consolidados" and sidebar_active_role.upper() in ("ADMINISTRADOR", "GESTOR"):
    st.markdown("### 📊 Dados Consolidados")
    st.markdown("---")
    
    st.markdown("""
        <style>
            div[data-testid="stExpander"] div.kpi-card {
                min-height: 145px !important;
                height: 145px !important;
                display: flex !important;
                flex-direction: column !important;
                justify-content: space-between !important;
                padding: 12px 16px !important;
            }
            div[data-testid="stExpander"] div.kpi-card .value {
                margin: auto 0 !important;
            }
        </style>
    """, unsafe_allow_html=True)
    
    drive_master_xlsx_id = cfg.get("drive_master_xlsx_id", "")
    fonte = cfg.get("fonte_dados", "📁 Pasta local / Servidor")
    
    if fonte == "☁️ Google Drive":
        master_xlsx_path = os.path.join(str(DADOS_DIR), "Analise avaliacoes completa.xlsx")
    else:
        master_xlsx_path = os.path.join(str(Path(DADOS_DIR).parent), "Analise avaliacoes completa.xlsx")
        
    with st.spinner("Consolidando dados do sistema..."):
        df_audit, err = load_audit_excel(master_xlsx_path, drive_master_xlsx_id)
        if err or df_audit is None:
            st.error(f"Erro ao carregar auditoria: {err}")
            st.stop()
            
    df_evals_source = df_full
    df_audit_source = df_audit
    
    def compute_metrics(df_sub_evals, df_sub_audit):
        def get_numeric_value(val):
            try:
                if val is None or str(val).strip() in ("", "-", "None", "nan"):
                    return None
                return float(str(val).replace(",", "."))
            except ValueError:
                return None

        grades = df_sub_audit["Nota Final - Média Aritmética"].apply(get_numeric_value).dropna()
        mean_val = grades.mean() if len(grades) > 0 else 0.0
        
        total_evals = len(df_sub_evals)
        n_ca = (df_sub_evals["Situação Comissão"] == "Comissão Atual").sum()
        n_np = (df_sub_evals["Situação Comissão"] == "Nota Provisória").sum()
        n_enc = (df_sub_evals["Status Avaliação"] == "Encerrada").sum()
        n_aberta = (df_sub_evals["Status Avaliação"] == "Aberta").sum()
        n_parc = (df_sub_evals["Status Avaliação"] == "Parcialmente Encerrada").sum()
        n_hom = (df_sub_evals["Status Avaliação"] == "Homologação").sum()
        
        mil_sim = (df_sub_audit["Todas Avaliações Foram Encerradas?"].astype(str).str.upper() == "SIM").sum()
        mil_nao = (df_sub_audit["Todas Avaliações Foram Encerradas?"].astype(str).str.upper() == "NAO").sum()
        
        return {
            "Avaliações Realizadas": total_evals,
            "Comissão Atual": n_ca,
            "Nota Provisória": n_np,
            "Encerradas": n_enc,
            "Abertas": n_aberta,
            "Parc. Encerradas": n_parc,
            "Homologação": n_hom,
            "AV1 Pendente": n_aberta,
            "AV2 Pendente": n_parc,
            "HOM Pendente": n_hom,
            "Militares Encerrados": mil_sim,
            "Militares Pendentes": mil_nao,
            "Média Notas": mean_val
        }
        
    all_rpms = sorted(df_evals_source["Unidade RPM (Avaliado)"].dropna().unique(), key=rpm_sort_key)
    
    # ── PAINEL DE EXPORTAÇÃO CONSOLIDADA ─────────────────────────────────────
    st.markdown("<h4 style='font-size: 1.2rem; color: #9b8a5c; margin-bottom: 8px;'>📥 Exportar Relatório Consolidado</h4>", unsafe_allow_html=True)
    escopo_rel = st.radio("Selecione o escopo da exportação:", ["🌐 Geral (Todas as RPMs)", "🏢 Por RPM específica"], horizontal=True, key="escopo_rel_consolidado")
    
    selected_rpms_rel = all_rpms
    if "específica" in escopo_rel:
        selected_rpms_rel = st.multiselect("Selecione as Unidades RPM/UDG para exportar:", all_rpms, default=[])
        
    if "específica" in escopo_rel and not selected_rpms_rel:
        st.warning("⚠️ Selecione ao menos uma unidade para habilitar a exportação.")
    else:
        # Gerar planilha consolidada em memória
        def export_consolidated_xlsx(df_evals, df_audit, rpms_to_export):
            import io
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            def get_numeric_value(val):
                try:
                    if val is None or str(val).strip() in ("", "-", "None", "nan"):
                        return None
                    return float(str(val).replace(",", "."))
                except ValueError:
                    return None
                    
            ws1 = wb.create_sheet("Consolidado RPM")
            headers = [
                "Unidade Principal (RPM/UDG)", "Total Avaliações", "Comissão Atual", "Nota Provisória",
                "Encerradas", "Abertas", "Parcialmente Encerradas", "Homologação",
                "AV1 Pendentes", "AV2 Pendentes", "HOM Pendentes",
                "Militares Encerrados", "Militares Pendentes", "Média Notas"
            ]
            ws1.append(headers)
            
            ws2 = wb.create_sheet("Detalhamento Subordinadas")
            headers_sub = ["RPM/UDG", "Unidade Subordinada"] + headers[1:]
            ws2.append(headers_sub)
            
            font_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            fill_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for ws in [ws1, ws2]:
                ws.row_dimensions[1].height = 28
                for cell in ws[1]:
                    cell.font = font_bold
                    cell.fill = fill_header
                    cell.alignment = align_center
                    
            for rpm in rpms_to_export:
                df_rpm_evals = df_evals[df_evals["Unidade RPM (Avaliado)"] == rpm]
                df_rpm_audit = df_audit[df_audit["Nome RPM"] == rpm]
                
                if len(df_rpm_evals) == 0:
                    continue
                    
                grades = df_rpm_audit["Nota Final - Média Aritmética"].apply(get_numeric_value).dropna()
                mean_val = grades.mean() if len(grades) > 0 else 0.0
                
                mil_sim = (df_rpm_audit["Todas Avaliações Foram Encerradas?"].astype(str).str.upper() == "SIM").sum()
                mil_nao = (df_rpm_audit["Todas Avaliações Foram Encerradas?"].astype(str).str.upper() == "NAO").sum()
                
                rpm_row = [
                    rpm,
                    len(df_rpm_evals),
                    (df_rpm_evals["Situação Comissão"] == "Comissão Atual").sum(),
                    (df_rpm_evals["Situação Comissão"] == "Nota Provisória").sum(),
                    (df_rpm_evals["Status Avaliação"] == "Encerrada").sum(),
                    (df_rpm_evals["Status Avaliação"] == "Aberta").sum(),
                    (df_rpm_evals["Status Avaliação"] == "Parcialmente Encerrada").sum(),
                    (df_rpm_evals["Status Avaliação"] == "Homologação").sum(),
                    (df_rpm_evals["Status Avaliação"] == "Aberta").sum(),
                    (df_rpm_evals["Status Avaliação"] == "Parcialmente Encerrada").sum(),
                    (df_rpm_evals["Status Avaliação"] == "Homologação").sum(),
                    mil_sim,
                    mil_nao,
                    round(mean_val, 2)
                ]
                ws1.append(rpm_row)
                
                unique_subs = sorted(df_rpm_evals["Unidade Principal (Avaliado)"].dropna().unique())
                for sub in unique_subs:
                    sub_evals = df_rpm_evals[df_rpm_evals["Unidade Principal (Avaliado)"] == sub]
                    sub_audit = df_rpm_audit[df_rpm_audit["Nome Unidade Principal"] == sub]
                    
                    sub_grades = sub_audit["Nota Final - Média Aritmética"].apply(get_numeric_value).dropna()
                    sub_mean_val = sub_grades.mean() if len(sub_grades) > 0 else 0.0
                    
                    sub_mil_sim = (sub_audit["Todas Avaliações Foram Encerradas?"].astype(str).str.upper() == "SIM").sum()
                    sub_mil_nao = (sub_audit["Todas Avaliações Foram Encerradas?"].astype(str).str.upper() == "NAO").sum()
                    
                    sub_row = [
                        rpm,
                        sub,
                        len(sub_evals),
                        (sub_evals["Situação Comissão"] == "Comissão Atual").sum(),
                        (sub_evals["Situação Comissão"] == "Nota Provisória").sum(),
                        (sub_evals["Status Avaliação"] == "Encerrada").sum(),
                        (sub_evals["Status Avaliação"] == "Aberta").sum(),
                        (sub_evals["Status Avaliação"] == "Parcialmente Encerrada").sum(),
                        (sub_evals["Status Avaliação"] == "Homologação").sum(),
                        (sub_evals["Status Avaliação"] == "Aberta").sum(),
                        (sub_evals["Status Avaliação"] == "Parcialmente Encerrada").sum(),
                        (sub_evals["Status Avaliação"] == "Homologação").sum(),
                        sub_mil_sim,
                        sub_mil_nao,
                        round(sub_mean_val, 2)
                    ]
                    ws2.append(sub_row)
                    
            for ws in [ws1, ws2]:
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                    
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.read()
            
        xlsx_consolidated = export_consolidated_xlsx(df_evals_source, df_audit_source, selected_rpms_rel)
        st.download_button(
            label="📥 Baixar Planilha Consolidada (Excel .xlsx)",
            data=xlsx_consolidated,
            file_name=f"Relatorio_Consolidado_AADP_{now_br().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="btn_download_consolidado"
        )
        
    st.markdown("---")
    st.markdown("Clique em uma Unidade Principal (RPM/UDG) para visualizar seu resumo consolidado e expandir suas Unidades Subordinadas.")
    
    for rpm in all_rpms:
        df_rpm_evals = df_evals_source[df_evals_source["Unidade RPM (Avaliado)"] == rpm]
        df_rpm_audit = df_audit_source[df_audit_source["Nome RPM"] == rpm]
        
        rpm_metrics = compute_metrics(df_rpm_evals, df_rpm_audit)
        
        total_evals = rpm_metrics["Avaliações Realizadas"]
        enc_evals = rpm_metrics["Encerradas"]
        avg_grade = rpm_metrics["Média Notas"]
        
        exp_header = f"🏢 {rpm} | Total Avaliações: {fmt_num(total_evals)} | Encerradas: {fmt_num(enc_evals)} | Média: {avg_grade:.2f}".replace(".", ",")
        
        with st.expander(exp_header):
            # Linha 1: 4 Cards principais
            k_c1, k_c2, k_c3, k_c4 = st.columns(4)
            with k_c1:
                st.markdown(f'<div class="kpi-card kpi-total">'
                            f'<div class="label">Total Avaliações</div>'
                            f'<div class="value">{fmt_num(total_evals)}</div>'
                            f'<div class="sub">Comissão: {fmt_num(rpm_metrics["Comissão Atual"])} CA | {fmt_num(rpm_metrics["Nota Provisória"])} NP</div>'
                            f'</div>', unsafe_allow_html=True)
            with k_c2:
                st.markdown(f'<div class="kpi-card kpi-aberta">'
                            f'<div class="label">Pendências Funcionais</div>'
                            f'<div class="value" style="font-size: 1.05rem; line-height: 1.35; font-weight: 700; margin-top: 6px; margin-bottom: 6px;">'
                            f'<span style="color: #ff6b6b !important;">AVALIADOR 1: {fmt_num(rpm_metrics["AV1 Pendente"])}</span><br>'
                            f'<span style="color: #ff9f43 !important;">AVALIADOR 2: {fmt_num(rpm_metrics["AV2 Pendente"])}</span><br>'
                            f'<span style="color: #ffd257 !important;">HOMOLOGADOR: {fmt_num(rpm_metrics["HOM Pendente"])}</span>'
                            f'</div>'
                            f'</div>', unsafe_allow_html=True)
            with k_c3:
                st.markdown(f'<div class="kpi-card kpi-parc">'
                            f'<div class="label">Militares Pendentes</div>'
                            f'<div class="value">{fmt_num(rpm_metrics["Militares Pendentes"])}</div>'
                            f'<div class="sub">Militar com avaliações não encerradas</div>'
                            f'</div>', unsafe_allow_html=True)
            with k_c4:
                avg_grade_str = f"{avg_grade:.2f}".replace(".", ",")
                st.markdown(f'<div class="kpi-card kpi-ca">'
                            f'<div class="label">Média das Notas</div>'
                            f'<div class="value">{avg_grade_str}</div>'
                            f'<div class="sub">Média aritmética final consolidada</div>'
                            f'</div>', unsafe_allow_html=True)
                            
            st.markdown("<div style='margin-bottom: 12px;'></div>", unsafe_allow_html=True)
            
            # Linha 2: 4 Quadrantes de Status
            q_c1, q_c2, q_c3, q_c4 = st.columns(4)
            with q_c1:
                st.markdown(f'<div class="kpi-card kpi-enc">'
                            f'<div class="label">🟢 Encerradas</div>'
                            f'<div class="value">{fmt_num(enc_evals)}</div>'
                            f'<div class="sub">Avaliações finalizadas</div>'
                            f'</div>', unsafe_allow_html=True)
            with q_c2:
                st.markdown(f'<div class="kpi-card kpi-aberta">'
                            f'<div class="label">🔴 Abertas</div>'
                            f'<div class="value">{fmt_num(rpm_metrics["Abertas"])}</div>'
                            f'<div class="sub">Sem nota ou AV1 pendente</div>'
                            f'</div>', unsafe_allow_html=True)
            with q_c3:
                st.markdown(f'<div class="kpi-card kpi-parc">'
                            f'<div class="label">🟠 Parcialmente Enc.</div>'
                            f'<div class="value">{fmt_num(rpm_metrics["Parc. Encerradas"])}</div>'
                            f'<div class="sub">Aguardando AV2</div>'
                            f'</div>', unsafe_allow_html=True)
            with q_c4:
                st.markdown(f'<div class="kpi-card kpi-hom">'
                            f'<div class="label">🟡 Homologação</div>'
                            f'<div class="value">{fmt_num(rpm_metrics["Homologação"])}</div>'
                            f'<div class="sub">Aguardando Homologador</div>'
                            f'</div>', unsafe_allow_html=True)
                            
            st.markdown("<div style='margin-bottom: 15px;'></div>", unsafe_allow_html=True)
            st.markdown("##### 📁 Detalhamento de Unidades Subordinadas")
            
            sub_rows = []
            unique_subs = sorted(df_rpm_evals["Unidade Principal (Avaliado)"].dropna().unique())
            for sub in unique_subs:
                sub_evals = df_rpm_evals[df_rpm_evals["Unidade Principal (Avaliado)"] == sub]
                sub_audit = df_rpm_audit[df_rpm_audit["Nome Unidade Principal"] == sub]
                
                metrics = compute_metrics(sub_evals, sub_audit)
                row = {"Unidade Subordinada": sub}
                row.update(metrics)
                sub_rows.append(row)
                
            if sub_rows:
                df_sub_table = pd.DataFrame(sub_rows)
                
                df_sub_table_disp = df_sub_table.rename(columns={
                    "Avaliações Realizadas": "Total Aval.",
                    "Comissão Atual": "CA",
                    "Nota Provisória": "NP",
                    "Encerradas": "Enc.",
                    "Abertas": "Aber.",
                    "Parc. Encerradas": "Parc. Enc.",
                    "Homologação": "Hom.",
                    "AV1 Pendente": "AV1 Pend.",
                    "AV2 Pendente": "AV2 Pend.",
                    "HOM Pendente": "HOM Pend.",
                    "Militares Encerrados": "Mil. Enc.",
                    "Militares Pendentes": "Mil. Pend.",
                    "Média Notas": "Média"
                })
                
                st.dataframe(df_sub_table_disp.style.format({
                    "Média": lambda x: f"{x:.2f}".replace(".", ",")
                }), use_container_width=True, hide_index=True)
            else:
                st.info("Nenhuma unidade subordinada encontrada.")


if active_page == "Painel Administrador" and st.session_state.user_role == "ADMINISTRADOR":
    st.markdown("### ⚙️ Painel Administrador")
    
    tab_pending, tab_active, tab_logs = st.tabs([
        "⏳ Solicitações de Cadastro Pendentes",
        "👥 Gerenciar/Alterar Usuários Cadastrados",
        "📜 Auditoria"
    ])
    
    # ── 1) Solicitações de Cadastro Pendentes ────────────────────────────────
    with tab_pending:
        st.markdown("#### ⏳ Solicitações de Cadastro Pendentes")
        pend_list = db_get_pending_users()
        
        if not pend_list:
            st.info("Não há solicitações pendentes no momento.")
        else:
            for pm, name, rank, rpm, unit, function, created_at in pend_list:
                pm_str = str(pm)
                with st.container():
                    st.markdown(f"**{rank} {name}** (PM: `{pm_str}`)")
                    st.markdown(f"RPM: `{rpm}` | Unidade: `{unit}` | Função: `{function}` | Data: `{created_at}`")
                    
                    c_role = st.selectbox("Selecione o perfil de acesso:", ["P1", "SADM", "Gestor", "ADMINISTRADOR"], key=f"role_{pm_str}")
                    
                    col_ap, col_rec, _ = st.columns([1, 1, 4])
                    if col_ap.button("✅ Autorizar Acesso", key=f"ap_{pm_str}", type="primary"):
                        target_rpm = "Gestor" if c_role in ("Gestor", "ADMINISTRADOR") else rpm
                        db_approve_user(pm_str, c_role, target_rpm)
                        log_action("ADM", "AUTORIZAR_ACESSO", f"Usuario {pm_str} ({name}) aprovado como {c_role}")
                        st.success(f"Acesso de {name} autorizado com sucesso!")
                        st.rerun()
                    if col_rec.button("❌ Recusar", key=f"rec_{pm_str}", type="secondary"):
                        db_reject_user(pm_str)
                        log_action("ADM", "RECUSAR_CADASTRO", f"Cadastro do usuario {pm_str} ({name}) recusado")
                        st.warning(f"Cadastro de {name} recusado!")
                        st.rerun()
                    st.markdown("---")
                    
    # ── 2) Gerenciar/Alterar Usuários Cadastrados ────────────────────────────
    with tab_active:
        st.markdown("#### 👥 Gerenciar / Alterar Usuários Cadastrados")
        
        # --- Sincronização manual com SIGEF ---
        if st.button("🔄 Sincronizar Cadastros com SIGEF (Atualiza Órgãos/Unidades)", use_container_width=True, key="btn_sync_sigef"):
            with st.spinner("⏳ Sincronizando dados com a base SIGEF..."):
                sync_users_with_sigef()
            st.success("✅ Cadastros sincronizados com a base SIGEF com sucesso!")
            st.rerun()
            
        st.markdown("---")
        
        # --- SIMULADOR DE VISÃO DE TELA ---
        st.markdown("##### 🕵️ Simulador de Visão de Tela")
        sim_users = db_get_simulator_users()
        
        if st.session_state.get("simulation_active", False):
            st.warning(f"🕵️ **Simulação Ativa**: Você está visualizando o sistema como **{st.session_state.simulated_name}** ({st.session_state.simulated_role}).")
            if st.button("❌ Parar Simulação / Voltar ao normal", type="secondary", use_container_width=True, key="btn_stop_sim"):
                st.session_state.simulation_active = False
                st.session_state.simulated_pm = ""
                st.session_state.simulated_name = ""
                st.session_state.simulated_role = ""
                st.session_state.simulated_rpm = ""
                st.session_state.simulated_unit = ""
                st.session_state.active_page = "Painel Administrador"
                log_action("ADM", "ENCERRAR_SIMULACAO", "Simulacao desativada")
                st.rerun()
        else:
            sim_options = []
            for u in sim_users:
                sim_options.append(f"{u[2]} {u[1]} ({u[3]} - {u[4]}) [PM: {u[0]}]")
            
            selected_sim_user = st.selectbox(
                "Selecione um usuário cadastrado para simular a visão dele:",
                sim_options,
                index=None,
                placeholder="Escolha um usuário...",
                key="sim_user_select"
            )
            
            if selected_sim_user:
                pm_match = re.search(r'\[PM:\s*(\w+)\]', selected_sim_user)
                if pm_match:
                    target_pm = pm_match.group(1)
                    if st.button("🎭 Simular visão deste usuário", type="primary", use_container_width=True, key="btn_start_sim"):
                        for u in sim_users:
                            if str(u[0]).strip() == str(target_pm).strip():
                                st.session_state.simulation_active = True
                                st.session_state.simulated_pm = u[0]
                                st.session_state.simulated_name = f"{u[2]} {u[1]}"
                                st.session_state.simulated_role = u[3]
                                st.session_state.simulated_rpm = u[4]
                                st.session_state.simulated_unit = u[5]
                                st.session_state.active_page = "Análise Gráfica"
                                log_action("ADM", "INICIAR_SIMULACAO", f"Simulando usuario {u[0]}")
                                st.success(f"Iniciando simulação de visão de: {u[2]} {u[1]}")
                                st.rerun()
                                
        st.markdown("---")
        
        # --- LISTA E ALTERAÇÃO DE USUÁRIOS ATIVOS ---
        st.markdown("##### 👥 Relação de Usuários Cadastrados")
        
        # Filtro por Nº PM
        filter_pm = st.text_input("🔍 Consultar por Nº PM (deixe em branco para ver todos):", "", key="active_users_filter_pm").strip()
        
        active_list = db_get_active_users()
        if not active_list:
            st.info("Nenhum usuário ativo cadastrado.")
        else:
            df_act = pd.DataFrame(active_list, columns=["Nº PM", "Posto/Grad.", "Nome", "RPM", "Unidade Principal", "Setor", "Perfil", "Data Cadastro"])
            
            # Reordenar para Nº PM / Posto/Grad. / Nome primeiro
            col_order = ["Nº PM", "Posto/Grad.", "Nome", "RPM", "Unidade Principal", "Setor", "Perfil", "Data Cadastro"]
            df_act = df_act[[c for c in col_order if c in df_act.columns]]
            
            # Filtragem se houver termo digitado
            if filter_pm:
                df_act = df_act[df_act["Nº PM"].astype(str).str.contains(filter_pm)]
                
            if df_act.empty:
                st.info("Nenhum usuário cadastrado encontrado com o Nº PM informado.")
            else:
                st.dataframe(clean_none_values(df_act), use_container_width=True, hide_index=True)
                
                st.markdown("##### ⚙️ Gerenciar / Alterar Cadastro:")
                
                user_options = [f"{row['Posto/Grad.']} {row['Nome']} (PM: {row['Nº PM']})" for _, row in df_act.iterrows()]
                selected_user_label = st.selectbox("Escolha o usuário para gerenciar:", user_options, key="manage_user_select")
                
                if selected_user_label:
                    m_pm = selected_user_label.split(" (PM: ")[1].rstrip(")")
                    matching_rows = df_act[df_act["Nº PM"].astype(str).str.strip() == str(m_pm).strip()]
                    if not matching_rows.empty:
                        row_sel = matching_rows.iloc[0]
                        
                        st.write(f"**Dados Atuais:** Perfil: `{row_sel['Perfil']}` | RPM: `{row_sel['RPM']}`")
                        
                        col_role, col_rpm = st.columns(2)
                        with col_role:
                            new_role = st.selectbox("Alterar Perfil de Acesso:", ["P1", "SADM", "Gestor", "ADMINISTRADOR"], 
                                                    index=["P1", "SADM", "Gestor", "ADMINISTRADOR"].index(row_sel["Perfil"]) if row_sel["Perfil"] in ["P1", "SADM", "Gestor", "ADMINISTRADOR"] else 0,
                                                    key=f"edit_role_{m_pm}")
                        with col_rpm:
                            rpm_choices = ["Gestor"] + [f"{i} RPM" for i in range(1, 20)] + [
                                "AM-ALMG", "AM-TJMG", "APM", "AUD SET", "CME", "COMAVE", "CPE", 
                                "CPM", "DAL", "DCO", "DEE", "DF", "DINT", "DOP", "DPS", "DRH", "DTS", 
                                "EMPM/SCG", "GCG", "GMG"
                            ]
                            try:
                                cur_index = rpm_choices.index(row_sel["RPM"])
                            except ValueError:
                                cur_index = 0
                                
                            new_rpm = st.selectbox("Alterar RPM / Diretoria / UDG:", rpm_choices, index=cur_index, key=f"edit_rpm_{m_pm}")
                        
                        col_save, col_rev = st.columns(2)
                        with col_save:
                            if st.button("💾 Salvar Alterações", type="primary", use_container_width=True, key=f"save_edit_{m_pm}"):
                                db_update_user_role_rpm(m_pm, new_role, new_rpm)
                                log_action("ADM", "ALTERAR_CADASTRO", f"Usuario {m_pm} alterado para Perfil: {new_role}, RPM: {new_rpm}")
                                st.success("✅ Alterações salvas com sucesso!")
                                st.rerun()
                        with col_rev:
                            if st.button("🚫 Revogar Acesso", type="secondary", use_container_width=True, key=f"revoke_edit_{m_pm}"):
                                db_revoke_user(m_pm)
                                log_action("ADM", "REVOGAR_ACESSO", f"Acesso do usuario {m_pm} revogado")
                                st.warning("🚫 Acesso revogado com sucesso!")
                                st.rerun()
                    else:
                        st.error("Erro ao selecionar usuário. O cache pode estar desatualizado.")
                        st.stop()
                        
    # ── 3) Auditoria ─────────────────────────────────────────────────────────
    with tab_logs:
        st.markdown("#### 📜 Auditoria de Atividades / Logs")
    
        # Buscar lista de usuários únicos que possuem ações nos logs
        logged_pms = db_get_logs_pms()
    
        # Mapear PM para Nome de forma amigável
        user_map = {"Todos": "Todos os Usuários"}
        for pm_id in logged_pms:
            u_row = db_get_user_info(pm_id)
            if u_row:
                user_map[pm_id] = f"{u_row[0]} {u_row[1]} (PM: {pm_id})"
            else:
                user_map[pm_id] = f"PM: {pm_id}"
            
        # Componente de filtro de militares
        sel_log_user = st.selectbox(
            "Filtrar logs por militar:",
            list(user_map.keys()),
            format_func=lambda x: user_map[x],
            key="filter_log_user"
        )
    
        # Filtro de período por data
        col_d1, col_d2 = st.columns(2)
        with col_d1:
            start_date = st.date_input("Data de início:", value=now_br().date() - timedelta(days=7), key="log_start_date")
        with col_d2:
            end_date = st.date_input("Data de fim:", value=now_br().date(), key="log_end_date")
        
        start_str = start_date.strftime("%Y-%m-%d")
        end_str = end_date.strftime("%Y-%m-%d")
    
        df_logs = db_get_logs_df(sel_log_user, start_str, end_str)
    
        if df_logs.empty:
            st.info("Nenhum log registrado para a seleção.")
        else:
            df_logs_disp = df_logs.copy()
            df_logs_disp.index = range(1, len(df_logs_disp) + 1)
            st.dataframe(clean_none_values(df_logs_disp), use_container_width=True)
        
        dl_log_xlsx = df_to_xlsx(df_logs)
        dl_log_csv = df_logs.to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig")
    
        col_l1, col_l2 = st.columns(2)
        with col_l1:
            st.download_button(
                "⬇️ Baixar Logs Filtrados (Excel .xlsx)",
                dl_log_xlsx,
                f"logs_{sel_log_user}_{now_br().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_logs_xlsx"
            )
        with col_l2:
            st.download_button(
                "⬇️ Baixar Logs Filtrados (CSV)",
                dl_log_csv,
                f"logs_{sel_log_user}_{now_br().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv",
                key="dl_logs_csv"
            )

st.markdown("---")
st.markdown(f"<center><small>AADP 2026 · Polícia Militar de Minas Gerais · "
            f"Resolução 5458/2025 · {now_br().strftime('%d/%m/%Y')}<br>"
            f"DIRETORIA DE RECURSO HUMANOS - DRH6</small></center>",
            unsafe_allow_html=True)

print(f"[AADP PROFILE] Script finished in {time.time() - _prof_start:.4f} seconds", flush=True)