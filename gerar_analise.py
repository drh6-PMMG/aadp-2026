# -*- coding: utf-8 -*-
"""
Script de Geração Local/CLI Consolidada do AADP 2026.
Usa a mesma engine openpyxl do app_aadp.py para paridade total de regras e gráficos.
"""
import csv, sys, io, os, zipfile, re, unicodedata
from datetime import datetime
from collections import defaultdict, Counter
import pandas as pd

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# ──────────────────────────── CONFIGURAÇÕES ───────────────────────────────────
BASE_DIR        = r"c:\Users\guilh\Downloads\analise AADP 2026"
AVALIACOES_FILE = os.path.join(BASE_DIR, "avaliacoes.csv")
SIGEF_FILE      = os.path.join(BASE_DIR, "SIGEF.csv")
OUT_DIR         = os.path.join(BASE_DIR, "Resultado_AADP_2026")
os.makedirs(OUT_DIR, exist_ok=True)

SITUACOES_ALVO = {
    "ATIV. DIRECAO GERAL", "ATIV. FIM DESTACADO", "ATIV. FIM NA SEDE",
    "ATIV. MEIO", "ATIVIDADE MEIO", "DISP MED DEFINITIVA",
    "GESTANTE/LAC/ADOTANT", "QUADRO ESPECIALISTA",
}

CONCEITO_FAIXA = {
    "nivel superior de desempenho":       (9.00, 10.00),
    "nivel alto de desempenho":           (7.00,  8.99),
    "nivel intermediario de desempenho":  (6.00,  6.99),
    "nivel baixo de desempenho":          (3.00,  5.99),
    "nivel inferior de desempenho":       (0.00,  2.99),
}

def normaliza(texto: str) -> str:
    t = unicodedata.normalize("NFD", texto.lower())
    return "".join(c for c in t if unicodedata.category(c) != "Mn")

def is_empty(v) -> bool:
    return not v or str(v).strip() in ("", "-")

def concordam(conceito: str, nota_str: str):
    if is_empty(conceito) or is_empty(nota_str):
        return None
    try:
        nota = float(str(nota_str).replace(",", "."))
    except ValueError:
        return None
    faixa = CONCEITO_FAIXA.get(normaliza(conceito.strip()))
    if faixa is None:
        return None
    return faixa[0] <= nota <= faixa[1]

def calc_cert_hom(j: str, l: str) -> str:
    if is_empty(j) or is_empty(l):
        return "-"
    c = concordam(j, l)
    return "NÃO" if c is True else ("SIM" if c is False else "-")

def calc_status(j: str, l: str, n: str) -> str:
    has_j = not is_empty(j)
    has_l = not is_empty(l)
    has_n = not is_empty(n)
    if not has_j:
        return "Aberta"
    if has_j and not has_l:
        return "Parcialmente Encerrada"
    c = concordam(j, l)
    if c is True:
        return "Encerrada"
    elif c is False:
        return "Encerrada" if has_n else "Homologação"
    return "Encerrada"

def rpm_sort_key(name):
    m = re.match(r'^(\d+)\s+RPM', str(name))
    return (0, int(m.group(1)), "") if m else (1, 0, str(name))

# ── Carregar Lotação SIGEF ────────────────────────────────────────────────────
print("Carregando SIGEF.csv ...")
sigef_unidade = {}
with open(SIGEF_FILE, encoding="cp1252", errors="replace") as f:
    reader = csv.reader(f, delimiter=";")
    next(reader)
    for row in reader:
        if len(row) > 9:
            nrpm = row[0].strip().lstrip("0") or "0"
            sigef_unidade[nrpm] = row[9].strip()

# ── Processar avaliacoes.csv ──────────────────────────────────────────────────
print("Carregando avaliacoes.csv ...")
rows = []
with open(AVALIACOES_FILE, encoding="cp1252", errors="replace") as f:
    reader = csv.reader(f, delimiter=";")
    next(reader)
    for row in reader:
        while len(row) < 50:
            row.append("")
        sit = row[7].strip()
        if sit not in SITUACOES_ALVO:
            continue
        nrpm = row[0].strip(); local = row[5].strip()
        j = row[9].strip(); l = row[11].strip(); n = row[13].strip()
        sc = "Comissão Atual" if local.upper().strip() == sigef_unidade.get(nrpm.lstrip("0") or "0","").upper().strip() else "Nota Provisória"
        rows.append({
            "nrPM (Avaliado)": nrpm, "Nome (Avaliado)": row[1].strip(),
            "Posto/Grad. (Avaliado)": row[2].strip(),
            "Unidade RPM (Avaliado)": row[3].strip(),
            "Unidade Principal (Avaliado)": row[4].strip(),
            "Local/Unidade (Avaliado)": local, "Situação Funcional": sit,
            "Data AV1": row[8].strip(), "Conceito Geral": j,
            "Data AV2": row[10].strip(), "Nota Geral": l,
            "Certificação Homologador": calc_cert_hom(j, l),
            "Data HOM": row[12].strip(), "Nota Homologação": n,
            "nrPM (Av1)": row[26].strip(), "Nome (Av1)": row[27].strip(),
            "Posto (Av1)": row[28].strip(), "RPM (Av1)": row[29].strip(),
            "Unid. Principal (Av1)": row[30].strip(), "Local (Av1)": row[31].strip(),
            "Quadro (Av1)": row[32].strip(), "Situação (Av1)": row[33].strip(),
            "nrPM (Av2)": row[34].strip(), "Nome (Av2)": row[35].strip(),
            "Posto (Av2)": row[36].strip(), "RPM (Av2)": row[37].strip(),
            "Unid. Principal (Av2)": row[38].strip(), "Local (Av2)": row[39].strip(),
            "Quadro (Av2)": row[40].strip(), "Situação (Av2)": row[41].strip(),
            "nrPM (Hom)": row[42].strip(), "Nome (Hom)": row[43].strip(),
            "Posto (Hom)": row[44].strip(), "RPM (Hom)": row[45].strip(),
            "Unid. Principal (Hom)": row[46].strip(), "Local (Hom)": row[47].strip(),
            "Quadro (Hom)": row[48].strip(), "Situação (Hom)": row[49].strip(),
            "Situação Comissão": sc,
            "Status Avaliação": calc_status(j, l, n),
        })

df_full = pd.DataFrame(rows)
print(f"  {len(df_full):,} registros processados.")

# ──────────────────────────── ENGINE EXCEL (openpyxl) ──────────────────────────
STATUS_BG_XL = {
    "Encerrada":             "70AD47",
    "Homologação":           "FFD966",
    "Parcialmente Encerrada":"FF8C00",
    "Aberta":                "FF4444",
}
SIT_BG_XL = {"Comissão Atual": "4472C4", "Nota Provisória": "FFC000"}

COLS_XLS = [
    "nrPM (Avaliado)", "Posto/Grad. (Avaliado)", "Nome (Avaliado)",
    "Unidade RPM (Avaliado)", "Unidade Principal (Avaliado)", "Local/Unidade (Avaliado)",
    "Situação Funcional",
    "Data AV1", "Data AV2", "Data HOM", "Certificação Homologador",
    "nrPM (Av1)", "Posto (Av1)", "Nome (Av1)", "RPM (Av1)", "Unid. Principal (Av1)",
    "nrPM (Av2)", "Posto (Av2)", "Nome (Av2)", "RPM (Av2)", "Unid. Principal (Av2)",
    "nrPM (Hom)", "Posto (Hom)", "Nome (Hom)", "RPM (Hom)", "Unid. Principal (Hom)",
    "Situação Comissão", "Status Avaliação",
]

def _xl_styles():
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    thin = Side(border_style="thin", color="CCCCCC")
    return {
        "hdr_fill":  PatternFill("solid", fgColor="1F3864"),
        "hdr_font":  Font(bold=True, color="FFFFFF", name="Calibri", size=10),
        "hdr_al":    Alignment(horizontal="center", vertical="center", wrap_text=True),
        "brd":       Border(left=thin, right=thin, top=thin, bottom=thin),
        "data_font": Font(name="Calibri", size=9),
        "center":    Alignment(horizontal="center", vertical="center"),
        "left":      Alignment(vertical="center"),
        "title_font":Font(bold=True, color="FFFFFF", name="Calibri", size=13),
        "title_al":  Alignment(horizontal="center", vertical="center"),
    }

def _write_title(ws, titulo, n_cols, s):
    from openpyxl.styles import PatternFill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(n_cols, 1))
    c = ws.cell(1, 1, titulo)
    c.fill = PatternFill("solid", fgColor="1F3864")
    c.font = s["title_font"]; c.alignment = s["title_al"]
    ws.row_dimensions[1].height = 22

def _write_headers(ws, cols, s):
    for i, col in enumerate(cols, 1):
        c = ws.cell(2, i, col)
        c.fill = s["hdr_fill"]; c.font = s["hdr_font"]
        c.alignment = s["hdr_al"]; c.border = s["brd"]
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

def _write_data_rows(ws, df, cols, s):
    from openpyxl.styles import PatternFill, Font
    for r, (_, row) in enumerate(df.iterrows(), 3):
        for ci, col in enumerate(cols, 1):
            val = row.get(col, "")
            cell = ws.cell(r, ci, str(val) if pd.notna(val) and val != "" else "")
            cell.font = s["data_font"]; cell.border = s["brd"]
            cell.alignment = s["center"] if ci > 7 else s["left"]
            if col == "Status Avaliação" and val in STATUS_BG_XL:
                cell.fill = PatternFill("solid", fgColor=STATUS_BG_XL[val])
                cell.font = Font(bold=True, name="Calibri", size=9,
                                  color="FFFFFF" if val == "Aberta" else "000000")
            elif col == "Situação Comissão" and val in SIT_BG_XL:
                cell.fill = PatternFill("solid", fgColor=SIT_BG_XL[val])
                cell.font = Font(bold=True, name="Calibri", size=9,
                                  color="FFFFFF" if val == "Comissão Atual" else "000000")

def _auto_widths(ws, df, cols):
    from openpyxl.utils import get_column_letter
    for ci, col in enumerate(cols, 1):
        max_len = 0
        if col in df.columns and not df.empty:
            max_val = df[col].astype(str).str.len().max()
            if pd.notna(max_val):
                max_len = int(max_val)
        max_len = max(len(str(col)), max_len)
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len * 0.92, 8), 40)

def _write_data_sheet(ws, df, titulo, cols, s):
    df_c = df[[c for c in cols if c in df.columns]].reset_index(drop=True)
    actual_cols = list(df_c.columns)
    _write_title(ws, titulo, len(actual_cols), s)
    _write_headers(ws, actual_cols, s)
    _write_data_rows(ws, df_c, actual_cols, s)
    _auto_widths(ws, df_c, actual_cols)

def _write_avaliadores_sheet(ws, df_unit, s):
    from openpyxl.styles import PatternFill, Font, Alignment
    _write_title(ws, "AVALIADORES PENDENTES — LOTADOS NA UNIDADE", 14, s)

    # AV1
    df_ab = df_unit[df_unit["Status Avaliação"] == "Aberta"]
    av1 = defaultdict(lambda: {"nome":"","posto":"","rpm":"","unid":"","CA":0,"NP":0})
    for _, r in df_ab.iterrows():
        k = str(r.get("nrPM (Av1)","")).strip()
        if not k: continue
        av1[k].update(nome=r.get("Nome (Av1)",""), posto=r.get("Posto (Av1)",""),
                      rpm=r.get("RPM (Av1)",""), unid=r.get("Unid. Principal (Av1)",""))
        if r["Situação Comissão"] == "Comissão Atual": av1[k]["CA"] += 1
        else: av1[k]["NP"] += 1

    # AV2
    df_pe = df_unit[df_unit["Status Avaliação"].isin(["Aberta","Parcialmente Encerrada"])]
    av2 = defaultdict(lambda: {"nome":"","posto":"","rpm":"","unid":"","CA_ab":0,"CA_pe":0,"NP_ab":0,"NP_pe":0})
    for _, r in df_pe.iterrows():
        k = str(r.get("nrPM (Av2)","")).strip()
        if not k: continue
        av2[k].update(nome=r.get("Nome (Av2)",""), posto=r.get("Posto (Av2)",""),
                      rpm=r.get("RPM (Av2)",""), unid=r.get("Unid. Principal (Av2)",""))
        is_ca = r["Situação Comissão"] == "Comissão Atual"
        if r["Status Avaliação"] == "Aberta":
            if is_ca: av2[k]["CA_ab"] += 1
            else: av2[k]["NP_ab"] += 1
        else:
            if is_ca: av2[k]["CA_pe"] += 1
            else: av2[k]["NP_pe"] += 1

    # Homologador
    df_hom = df_unit[df_unit["Status Avaliação"] == "Homologação"]
    hom = defaultdict(lambda: {"nome":"","posto":"","rpm":"","unid":"","CA":0,"NP":0})
    for _, r in df_hom.iterrows():
        k = str(r.get("nrPM (Hom)","")).strip() or "N/I"
        hom[k].update(nome=r.get("Nome (Hom)",""), posto=r.get("Posto (Hom)",""),
                      rpm=r.get("RPM (Hom)",""), unid=r.get("Unid. Principal (Hom)",""))
        if r["Situação Comissão"] == "Comissão Atual": hom[k]["CA"] += 1
        else: hom[k]["NP"] += 1

    row_num = 3
    # AV1 Header
    titles_av1 = ["Nº PM","Posto","Nome","RPM","Unidade","CA—Aberta","NP—Aberta","Total AV1"]
    fill_av1 = PatternFill("solid", fgColor="1F3864")
    for ci, h in enumerate(titles_av1, 1):
        c = ws.cell(row_num, ci, h)
        c.fill = fill_av1; c.font = s["hdr_font"]; c.alignment = s["hdr_al"]; c.border = s["brd"]
    ws.merge_cells(start_row=row_num-1, start_column=1, end_row=row_num-1, end_column=8)
    lbl = ws.cell(row_num-1, 1, "AVALIADOR 1 — Avaliações Em Aberto")
    lbl.fill = PatternFill("solid", fgColor="2E5090"); lbl.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    lbl.alignment = Alignment(horizontal="center", vertical="center")
    row_num += 1
    for k, d in sorted(av1.items(), key=lambda x: -(x[1]["CA"]+x[1]["NP"])):
        tot = d["CA"]+d["NP"]
        if tot == 0: continue
        for ci, val in enumerate([k, d["posto"], d["nome"], d["rpm"], d["unid"],
                                    d["CA"], d["NP"], tot], 1):
            c = ws.cell(row_num, ci, val)
            c.font = s["data_font"]; c.border = s["brd"]; c.alignment = s["center"]
        row_num += 1

    row_num += 2
    # AV2 Header
    titles_av2 = ["Nº PM","Posto","Nome","RPM","Unidade","CA—Ab.","CA—PE","NP—Ab.","NP—PE","Total AV2"]
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)
    lbl2 = ws.cell(row_num, 1, "AVALIADOR 2 — Em Aberto + Parcialmente Encerrada")
    lbl2.fill = PatternFill("solid", fgColor="2E5090"); lbl2.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    lbl2.alignment = Alignment(horizontal="center", vertical="center")
    row_num += 1
    for ci, h in enumerate(titles_av2, 1):
        c = ws.cell(row_num, ci, h)
        c.fill = fill_av1; c.font = s["hdr_font"]; c.alignment = s["hdr_al"]; c.border = s["brd"]
    row_num += 1
    for k, d in sorted(av2.items(), key=lambda x: -(x[1]["CA_ab"]+x[1]["CA_pe"]+x[1]["NP_ab"]+x[1]["NP_pe"])):
        tot = d["CA_ab"]+d["CA_pe"]+d["NP_ab"]+d["NP_pe"]
        if tot == 0: continue
        for ci, val in enumerate([k, d["posto"], d["nome"], d["rpm"], d["unid"],
                                    d["CA_ab"], d["CA_pe"], d["NP_ab"], d["NP_pe"], tot], 1):
            c = ws.cell(row_num, ci, val)
            c.font = s["data_font"]; c.border = s["brd"]; c.alignment = s["center"]
        row_num += 1

    row_num += 2
    # HOM Header
    titles_hom = ["Nº PM","Posto","Nome","RPM","Unidade","CA—Pend.","NP—Pend.","Total HOM"]
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
    lbl3 = ws.cell(row_num, 1, "HOMOLOGADOR — Avaliações com Divergência Aguardando Nota")
    lbl3.fill = PatternFill("solid", fgColor="7B3F00"); lbl3.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    lbl3.alignment = Alignment(horizontal="center", vertical="center")
    row_num += 1
    for ci, h in enumerate(titles_hom, 1):
        c = ws.cell(row_num, ci, h)
        c.fill = PatternFill("solid", fgColor="7B3F00"); c.font = s["hdr_font"]
        c.alignment = s["hdr_al"]; c.border = s["brd"]
    row_num += 1
    for k, d in sorted(hom.items(), key=lambda x: -(x[1]["CA"]+x[1]["NP"])):
        tot = d["CA"]+d["NP"]
        if tot == 0: continue
        for ci, val in enumerate([k, d["posto"], d["nome"], d["rpm"], d["unid"],
                                    d["CA"], d["NP"], tot], 1):
            c = ws.cell(row_num, ci, val)
            c.font = s["data_font"]; c.border = s["brd"]; c.alignment = s["center"]
        row_num += 1

    from openpyxl.utils import get_column_letter
    for ci in range(1, 15):
        ws.column_dimensions[get_column_letter(ci)].width = 18

def _write_analise_sheet(ws, df, titulo, s):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList

    _write_title(ws, f"ANÁLISE — {titulo}", 6, s)

    STATUS_ORD = ["Aberta", "Parcialmente Encerrada", "Homologação", "Encerrada"]
    ca  = df[df["Situação Comissão"] == "Comissão Atual"]
    np_ = df[df["Situação Comissão"] == "Nota Provisória"]
    total = len(df)

    fill_ca  = PatternFill("solid", fgColor="4472C4")
    fill_np  = PatternFill("solid", fgColor="FFC000")
    fill_tot = PatternFill("solid", fgColor="1F3864")
    fill_st  = {
        "Aberta":                 PatternFill("solid", fgColor="FF4444"),
        "Parcialmente Encerrada": PatternFill("solid", fgColor="FF8C00"),
        "Homologação":            PatternFill("solid", fgColor="FFD966"),
        "Encerrada":              PatternFill("solid", fgColor="70AD47"),
    }
    white_bold  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    black_bold  = Font(bold=True, color="000000", name="Calibri", size=11)
    center_al   = Alignment(horizontal="center", vertical="center")
    thin = s["brd"]

    r = 3
    headers = ["STATUS", "COMISSÃO ATUAL", "NOTA PROVISÓRIA", "TOTAL", "%"]
    col_fills = [fill_tot, fill_ca, fill_np, fill_tot, fill_tot]
    col_fonts = [white_bold, white_bold, black_bold, white_bold, white_bold]
    for ci, (h, fll, fnt) in enumerate(zip(headers, col_fills, col_fonts), 1):
        c = ws.cell(r, ci, h)
        c.fill = fll; c.font = fnt; c.alignment = center_al; c.border = thin
        ws.column_dimensions[ws.cell(r, ci).column_letter].width = 26 if ci == 1 else 18
    ws.row_dimensions[r].height = 28
    r += 1

    data_start_row = r
    for st in STATUS_ORD:
        ca_n  = int((ca["Status Avaliação"] == st).sum())
        np_n  = int((np_["Status Avaliação"] == st).sum())
        tot_n = ca_n + np_n
        pct   = f"{tot_n/total*100:.1f}%" if total > 0 else "0%"

        fll_st = fill_st.get(st, PatternFill())
        use_white = st == "Aberta"

        c0 = ws.cell(r, 1, st)
        c0.fill = fll_st; c0.border = thin
        c0.font = white_bold if use_white else black_bold
        c0.alignment = center_al

        c1 = ws.cell(r, 2, ca_n)
        c1.fill = fill_ca; c1.font = white_bold; c1.alignment = center_al; c1.border = thin

        c2 = ws.cell(r, 3, np_n)
        c2.fill = fill_np; c2.font = black_bold; c2.alignment = center_al; c2.border = thin

        c3 = ws.cell(r, 4, tot_n)
        c3.fill = fill_tot; c3.font = white_bold; c3.alignment = center_al; c3.border = thin

        c4 = ws.cell(r, 5, pct)
        c4.font = Font(name="Calibri", size=11); c4.alignment = center_al; c4.border = thin

        ws.row_dimensions[r].height = 22
        r += 1
    data_end_row = r - 1

    ca_tot = len(ca); np_tot = len(np_)
    for ci, val in enumerate(["TOTAL GERAL", ca_tot, np_tot, total, "100%"], 1):
        c = ws.cell(r, ci, val)
        c.fill = fill_tot; c.font = white_bold; c.alignment = center_al; c.border = thin
    ws.row_dimensions[r].height = 26
    r += 2

    for txt, fll, fnt in [
        ("COMISSÃO ATUAL — Policial lotado na unidade avaliadora", fill_ca, white_bold),
        ("NOTA PROVISÓRIA — Policial transferido; nota pode mudar", fill_np, black_bold),
    ]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        lc = ws.cell(r, 1, txt)
        lc.fill = fll; lc.font = fnt
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1
    sit_data_row = r
    ws.cell(r, 7, "Situação");  ws.cell(r, 8, "Qtd")
    r += 1
    ws.cell(r, 7, "Comissão Atual");  ws.cell(r, 8, ca_tot)
    r += 1
    ws.cell(r, 7, "Nota Provisória"); ws.cell(r, 8, np_tot)
    sit_end_row = r

    # Gráfico 1: Pizza por Status
    pie1 = PieChart()
    pie1.title  = "Status das Avaliações"
    pie1.style  = 10
    pie1.width  = 14
    pie1.height = 10
    data1 = Reference(ws, min_col=4, min_row=data_start_row, max_row=data_end_row)
    cats1 = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
    pie1.add_data(data1)
    pie1.set_categories(cats1)
    pie1.series[0].title = None
    SLICE_COLORS = ["FF4444", "FF8C00", "FFD966", "70AD47"]
    for idx, hex_color in enumerate(SLICE_COLORS):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = hex_color
        pie1.series[0].dPt.append(pt)
    pie1.dataLabels = DataLabelList()
    pie1.dataLabels.showPercent     = True
    pie1.dataLabels.showCatName     = True
    pie1.dataLabels.showVal         = True
    pie1.dataLabels.showLeaderLines = True
    ws.add_chart(pie1, "G3")

    # Gráfico 2: Pizza CA vs NP
    pie2 = PieChart()
    pie2.title  = "Situação da Comissão"
    pie2.style  = 10
    pie2.width  = 14
    pie2.height = 10
    data2 = Reference(ws, min_col=8, min_row=sit_data_row + 1, max_row=sit_end_row)
    cats2 = Reference(ws, min_col=7, min_row=sit_data_row + 1, max_row=sit_end_row)
    pie2.add_data(data2)
    pie2.set_categories(cats2)
    pie2.series[0].title = None
    for idx, hex_color in enumerate(["4472C4", "FFC000"]):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = hex_color
        pie2.series[0].dPt.append(pt)
    pie2.dataLabels = DataLabelList()
    pie2.dataLabels.showPercent     = True
    pie2.dataLabels.showCatName     = True
    pie2.dataLabels.showVal         = True
    pie2.dataLabels.showLeaderLines = True
    ws.add_chart(pie2, "G23")

def _build_workbook(df_unit: pd.DataFrame, titulo: str) -> bytes:
    from openpyxl import Workbook
    import io
    s = _xl_styles()
    wb = Workbook()

    cols = [c for c in COLS_XLS if c in df_unit.columns]

    # Aba 1 — Geral
    ws1 = wb.active; ws1.title = "Geral"
    _write_data_sheet(ws1, df_unit, f"AVALIAÇÕES — {titulo}", cols, s)

    # Aba 2 — Avaliações Pendentes
    ws2 = wb.create_sheet("Avaliações Pendentes")
    df_pend = df_unit[df_unit["Status Avaliação"].isin(["Aberta", "Parcialmente Encerrada", "Homologação"])]
    _write_data_sheet(ws2, df_pend, f"AVALIAÇÕES PENDENTES — {titulo}", cols, s)

    # Aba 3 — Avaliadores Pendentes
    ws3 = wb.create_sheet("Avaliadores Pendentes")
    _write_avaliadores_sheet(ws3, df_unit, s)

    # Aba 4 — Análise (tabela + gráficos de pizza)
    ws4 = wb.create_sheet("Análise")
    _write_analise_sheet(ws4, df_unit, titulo, s)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()

# ─────────────────────── MAIN (CLI) ───────────────────────────────────────────
import argparse

parser = argparse.ArgumentParser(description="Gera relatórios AADP 2026 em Excel/ZIP localmente.")
parser.add_argument("--mode",   choices=["all","geral","units"], default="all",
                    help="all=tudo, geral=somente Geral, units=unidades específicas")
parser.add_argument("--units",  default="",
                    help="Lista de RPMs separadas por vírgula (usado com --mode units)")
parser.add_argument("--output", default=None,
                    help="Caminho do arquivo ZIP de saída")
args = parser.parse_args()

zip_out = args.output or os.path.join(BASE_DIR, "Resultado_AADP_2026.zip")
units_req = [u.strip() for u in args.units.split(",") if u.strip()] if args.units else []

# Mapeamento por RPM
rpm_map = defaultdict(list)
for _, r in df_full.iterrows():
    rpm = str(r["Unidade RPM (Avaliado)"]).strip()
    if rpm and rpm != "nan":
        rpm_map[rpm].append(r)

arquivos_zip = []

# ── Geral ─────────────────────────────────────────────────────────────────────
if args.mode in ("all", "geral"):
    print("\n=== Gerando Planilha GERAL ===")
    geral_path = os.path.join(OUT_DIR, "Analise_Avaliacoes_Geral.xlsx")
    xlsx_bytes = _build_workbook(df_full, "GERAL — AADP 2026")
    with open(geral_path, "wb") as f_out:
        f_out.write(xlsx_bytes)
    print(f"  OK → {geral_path}")
    arquivos_zip.append(geral_path)

# ── Por RPM ───────────────────────────────────────────────────────────────────
if args.mode in ("all", "units"):
    target_rpms = units_req if units_req else sorted(rpm_map.keys(), key=rpm_sort_key)
    print(f"\n=== Gerando Planilhas por Unidade RPM ({len(target_rpms)} unidades) ===")
    for rpm in target_rpms:
        df_rpm = df_full[df_full["Unidade RPM (Avaliado)"] == rpm].copy()
        if df_rpm.empty:
            print(f"  AVISO: nenhum registro para '{rpm}'")
            continue
        safe  = re.sub(r'[^\w]', '_', str(rpm))
        fname = f"Analise_Avaliacoes_{safe}.xlsx"
        fp    = os.path.join(OUT_DIR, fname)
        print(f"  {fname}  ({len(df_rpm):,} reg.)")
        xlsx_bytes = _build_workbook(df_rpm, rpm)
        with open(fp, "wb") as f_out:
            f_out.write(xlsx_bytes)
        arquivos_zip.append(fp)

# ── Compactar ZIP ─────────────────────────────────────────────────────────────
print("\n=== Compactando ZIP ===")
with zipfile.ZipFile(zip_out, "w", zipfile.ZIP_DEFLATED) as zf:
    for fp in arquivos_zip:
        zf.write(fp, arcname=os.path.basename(fp))

print(f"  ZIP → {zip_out}")
print(f"\n=== CONCLUÍDO ===")
print(f"  Total registros : {len(df_full):,}")
sc = df_full["Status Avaliação"].value_counts().to_dict()
print(f"  Aberta          : {sc.get('Aberta', 0):,}")
print(f"  Parc. Encerrada : {sc.get('Parcialmente Encerrada', 0):,}")
print(f"  Homologação     : {sc.get('Homologação', 0):,}")
print(f"  Encerrada       : {sc.get('Encerrada', 0):,}")
print(f"  Planilhas       : {len(arquivos_zip)}")
