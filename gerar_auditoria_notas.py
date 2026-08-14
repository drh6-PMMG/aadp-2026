# -*- coding: utf-8 -*-
"""
Script de Auditoria de Notas AADP 2026
Gera planilha completa a partir de geral.csv + SIGEF.csv
com enriquecimento de MOVIMENTAÇÕES, PRESO e PUNIÇÃO.
Data de referência: 2026-07-24
"""

import csv, sys, io, os, re, unicodedata, zipfile
from datetime import datetime, date
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# ─── CONFIGURAÇÕES ────────────────────────────────────────────────────────────
BASE_DIR         = r"C:\Users\guilh\Downloads\analise AADP 2026"
GERAL_FILE       = os.path.join(BASE_DIR, "geral.csv")
SIGEF_FILE       = os.path.join(BASE_DIR, "SIGEF.csv")
MOV_FILE         = os.path.join(BASE_DIR, "MOVIMENTAÇÕES.xlsx")
PRESO_FILE       = os.path.join(BASE_DIR, "PRESO.xlsx")
PUN_FILE         = os.path.join(BASE_DIR, "PUNIÇÃO.xlsx")
OUT_DIR          = os.path.join(BASE_DIR, "Resultado_AADP_2026")
OUT_FILE         = os.path.join(OUT_DIR, "Auditoria_Notas_AADP2026.xlsx")
TODAY            = date.today()

os.makedirs(OUT_DIR, exist_ok=True)

SITUACOES_ALVO = {
    "ATIV. DIRECAO GERAL", "ATIV. FIM DESTACADO", "ATIV. FIM NA SEDE",
    "ATIV. MEIO", "ATIVIDADE MEIO", "DISP MED DEFINITIVA",
    "GESTANTE/LAC/ADOTANT", "QUADRO ESPECIALISTA",
}

CONCEITO_FAIXA = {
    "nivel superior de desempenho":       (9.00, 10.00),
    "nivel alto de desempenho":           (7.00,  8.99),
    "nivel intermediario de desempenho":  (6.00,  6.99),
    "nivel baixo de desempenho":          (3.00,  5.99),
    "nivel inferior de desempenho":       (0.00,  2.99),
}

# ─── FUNÇÕES UTILITÁRIAS ──────────────────────────────────────────────────────
def normaliza(texto: str) -> str:
    t = unicodedata.normalize("NFD", str(texto).lower())
    return "".join(c for c in t if unicodedata.category(c) != "Mn")

def is_empty(v) -> bool:
    if v == 0 or v == 0.0:
        return False
    return not v or str(v).strip() in ("", "-", "nan", "none", "NaT")

def concordam(conceito: str, nota_str: str):
    if is_empty(conceito) or is_empty(nota_str):
        return None
    try:
        nota = float(str(nota_str).replace(",", "."))
    except ValueError:
        return None
    faixa = CONCEITO_FAIXA.get(normaliza(conceito.strip()))
    if faixa is None:
        return None
    return faixa[0] <= nota <= faixa[1]

def calc_cert_hom(j: str, l: str) -> str:
    if is_empty(j) or is_empty(l):
        return "-"
    c = concordam(j, l)
    return "NÃO" if c is True else ("SIM" if c is False else "-")

def calc_status(j: str, l: str, n: str) -> str:
    if is_empty(j):
        return "Aberta"
    if is_empty(l):
        return "Parcialmente Encerrada"
    c = concordam(j, l)
    if c is True:
        return "Encerrada"
    elif c is False:
        return "Encerrada" if not is_empty(n) else "Homologação"
    return "Parcialmente Encerrada"

def fmt_date(val) -> str:
    """Formata datetime/date para string dd/mm/aaaa."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%d/%m/%Y")
    s = str(val).strip()
    if s in ("", "NaT", "nan", "00/00/0000"):
        return ""
    # Tenta parsear
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s.split(" ")[0], fmt.split(" ")[0]).strftime("%d/%m/%Y")
        except:
            pass
    return s

def parse_date_safe(val):
    """Retorna date ou None."""
    if val is None:
        return None
    if isinstance(val, (datetime,)):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, float) and pd.isna(val):
        return None
    s = str(val).strip()
    if s in ("", "NaT", "nan", "00/00/0000", "None"):
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(s[:10], fmt[:10]).date()
        except:
            pass
    return None

def add_business_days(start_date, num_days):
    from datetime import timedelta
    curr = start_date
    added = 0
    while added < num_days:
        curr += timedelta(days=1)
        if curr.weekday() < 5: # Mon-Fri
            added += 1
    return curr

def parse_float(s):
    if not s or str(s).strip() in ("", "-", "nan", "none", "None", "<NA>"):
        return None
    try:
        return float(str(s).replace(",", "."))
    except ValueError:
        return None

# ─── 1. SIGEF — mapeia nrPM → unidade ────────────────────────────────────────
print("Carregando SIGEF.csv ...")
sigef_unidade = {}
with open(SIGEF_FILE, encoding="cp1252", errors="replace") as f:
    reader = csv.reader(f, delimiter=";")
    next(reader)
    for row in reader:
        if len(row) > 9:
            nrpm = row[0].strip().lstrip("0") or "0"
            sigef_unidade[nrpm] = row[9].strip()

# ─── 2. MOVIMENTAÇÕES — mapeia nrPM → lista de movimentações ─────────────────
print("Carregando MOVIMENTAÇÕES.xlsx ...")
df_mov = pd.read_excel(MOV_FILE, dtype=str)
# Limpeza de nomes de colunas
df_mov.columns = [c.strip() for c in df_mov.columns]

# Mapeia: nrPM → lista de {data, motivo}
mov_map = defaultdict(list)
for _, row in df_mov.iterrows():
    nrpm = str(row.get("Numero Servidor", "")).strip().lstrip("0") or "0"
    dt_raw = row.get("Dt movimentacao", row.get(" Dt movimentacao", ""))
    motivo = str(row.get("Motivo movimentacao", "")).strip()
    dt = parse_date_safe(dt_raw)
    if nrpm and nrpm != "0":
        mov_map[nrpm].append({"data": dt, "motivo": motivo})

# Para cada militar, pega a movimentação mais recente
mov_mais_recente = {}
for nrpm, lista in mov_map.items():
    lista_valida = [x for x in lista if x["data"] is not None]
    if lista_valida:
        mais_recente = max(lista_valida, key=lambda x: x["data"])
        mov_mais_recente[nrpm] = mais_recente
    elif lista:
        mov_mais_recente[nrpm] = lista[0]

# ─── 3. PRESO — mapeia nrPM → dados de prisão ────────────────────────────────
print("Carregando PRESO.xlsx ...")
df_preso = pd.read_excel(PRESO_FILE, dtype=str)
df_preso.columns = [c.strip() for c in df_preso.columns]

# Mapeia: nrPM → lista de registros (pode haver mais de um por militar)
preso_map = defaultdict(list)
for _, row in df_preso.iterrows():
    nrpm = str(row.get("NUMERO", "")).strip().lstrip("0") or "0"
    data_inicio_raw = row.get("DATA INICIO", "")
    data_fim_raw    = row.get("DATA TERMINO", "")
    sit_func        = str(row.get("SITUACAO FUNCIONAL", "")).strip()

    data_inicio = parse_date_safe(data_inicio_raw)
    data_fim    = parse_date_safe(data_fim_raw)

    if nrpm and nrpm != "0":
        preso_map[nrpm].append({
            "data_inicio": data_inicio,
            "data_fim":    data_fim,
            "sit_func":    sit_func,
        })

# Para cada militar, se há múltiplos registros, considera a prisão mais recente
# com data de início não nula
preso_info = {}
for nrpm, lista in preso_map.items():
    lista_valida = [x for x in lista if x["data_inicio"] is not None]
    if lista_valida:
        mais_recente = max(lista_valida, key=lambda x: x["data_inicio"])
        preso_info[nrpm] = mais_recente

# ─── 4. PUNIÇÃO — mapeia nrPM → dados agregados ──────────────────────────────
print("Carregando PUNIÇÃO.xlsx ...")
df_pun = pd.read_excel(PUN_FILE, dtype=str)
df_pun.columns = [c.strip() for c in df_pun.columns]

pun_map = defaultdict(lambda: {"qtd_punicoes": 0, "qtd_ativadas": 0, "pontos": 0})

for _, row in df_pun.iterrows():
    nrpm = str(row.get("MATRICULA", "")).strip().lstrip("0") or "0"
    if not nrpm or nrpm == "0":
        continue

    # DATA PUNIÇÃO — coluna J (índice 9)
    data_pun_raw = row.get("DATA PUNICAO", "")
    data_pun = parse_date_safe(data_pun_raw)

    if data_pun is None:
        continue  # Somente conta punições com data válida

    pun_map[nrpm]["qtd_punicoes"] += 1

    # DATA ATIVAÇÃO — coluna K: se não for "00/00/0000" e tiver data válida → ativada
    data_ativ_raw = str(row.get("DATA ATIVACAO", "")).strip()
    if data_ativ_raw not in ("00/00/0000", "", "nan", "None", "NaT"):
        data_ativ = parse_date_safe(data_ativ_raw)
        if data_ativ is not None:
            pun_map[nrpm]["qtd_ativadas"] += 1

    # PONTOS — coluna O
    pontos_raw = row.get("PONTOS", "0")
    try:
        pontos = int(float(str(pontos_raw).replace(",", ".")))
    except:
        pontos = 0
    pun_map[nrpm]["pontos"] += pontos

# ─── 5. GERAL.CSV — processar e montar DataFrame principal ────────────────────
print("Carregando geral.csv ...")
rows = []
with open(GERAL_FILE, encoding="cp1252", errors="replace") as f:
    reader = csv.reader(f, delimiter=";")
    headers = next(reader)

    # Índices das colunas no geral.csv (baseado na estrutura confirmada)
    # [0]  Id do CDP          → não usado diretamente
    # [1]  nrPM (Avaliado)
    # [2]  Nome Completo (Avaliado)
    # [3]  Função
    # [4]  Atividades
    # [5]  Data do CDP
    # [6]  Posto/Graduação (Avaliado)
    # [7]  Unidade RPM (Avaliado)
    # [8]  Unidade Principal (Avaliado)
    # [9]  Local/Unidade (Avaliado)
    # [10] Quadro Atual (Avaliado)
    # [11] Situação Funcional Atual (Avaliado)
    # [28] nrPM (Avaliador 1)
    # [29] Nome Completo (Avaliador 1)
    # [30] Posto/Graduação (Avaliador 1)
    # [31] Unidade RPM (Avaliador 1)
    # [32] Unidade Principal (Avaliador 1)
    # [33] Local/Unidade (Avaliador 1)
    # [34] Quadro Atual (Avaliador 1)
    # [35] Situação Funcional Atual (Avaliador 1)
    # [36] Data da Avaliação 1
    # [37] nrPM (Avaliador 2)
    # [38] Nome Completo (Avaliador 2)
    # [39] Posto/Graduação (Avaliador 2)
    # [40] Unidade RPM (Avaliador 2)
    # [41] Unidade Principal (Avaliador 2)
    # [42] Local/Unidade (Avaliador 2)
    # [43] Quadro Atual (Avaliador 2)
    # [44] Situação Funcional Atual (Avaliador 2)
    # [45] Data da Avaliação 2
    # [46] Conceito Geral
    # [47] Nota Geral
    # [48] Competência 1
    # [49] Conceito (Competência 1)
    # [50] Nota (Competência 1)
    # [51] Competência 2
    # [52] Conceito (Competência 2)
    # [53] Nota (Competência 2)
    # [54] Competência 3
    # [55] Conceito (Competência 3)
    # [56] Nota (Competência 3)
    # [57] Competência 4
    # [58] Conceito (Competência 4)
    # [59] Nota (Competência 4)
    # [62] nrPM (Homologador)
    # [63] Nome Completo (Homologador)
    # [64] Posto/Graduação (Homologador)
    # [65] Unidade RPM (Homologador)
    # [66] Unidade Principal (Homologador)
    # [67] Local/Unidade (Homologador)
    # [68] Quadro Atual (Homologador)
    # [69] Situação Funcional Atual (Homologador)
    # [70] Nota da Homologação
    # [71] Data da Homologação

    for row in reader:
        while len(row) < 198:
            row.append("")

        sit = row[11].strip()
        if sit not in SITUACOES_ALVO:
            continue

        nrpm  = row[1].strip()
        local = row[9].strip()
        nrpm_key = nrpm.lstrip("0") or "0"

        # Conceito, Nota, Homologação
        j = row[46].strip()   # Conceito Geral
        l = row[47].strip()   # Nota Geral
        n = row[70].strip()   # Nota Homologação

        # Situação Comissão
        sigef_unit = sigef_unidade.get(nrpm_key, "")
        sc = "Comissão Atual" if local.upper().strip() == sigef_unit.upper().strip() else "Nota Provisória"

        # Lógica de Recurso para o Status da Avaliação
        status_av = calc_status(j, l, n)
        n_f4 = parse_float(row[86])
        n_f3 = parse_float(row[82])
        n_f2 = parse_float(row[78])
        n_f1 = parse_float(row[74])
        
        r_f4 = row[85].strip()
        r_f3 = row[81].strip()
        r_f2 = row[77].strip()
        r_f1 = row[73].strip()
        
        dt_f4 = row[88].strip()  # Data Cadastro (Fase 4)
        dt_f3 = row[84].strip()  # Data Cadastro (Fase 3)
        dt_f2 = row[80].strip()  # Data Cadastro (Fase 2)
        dt_f1 = row[76].strip()  # Data Cadastro (Fase 1)
        
        # Nota original da comissão (AV2 ou Homologado se houve divergência)
        original_grade_str = n if not is_empty(n) else l
        
        c_val = concordam(j, l)
        has_appeal = (r_f1 not in ("", "-")) or (n_f1 is not None)
        ref_date = date.today()
        
        if status_av in ("Aberta", "Parcialmente Encerrada"):
            pass
        elif not has_appeal:
            if c_val is False:
                # Houve discordância: necessita passar para o homologador
                if is_empty(n):
                    status_av = "Homologação"
                else:
                    dt_base = parse_date_safe(row[71]) # Data Homologação
                    if dt_base is not None:
                        deadline = add_business_days(dt_base, 5)
                        if ref_date <= deadline:
                            status_av = "EM PRAZO DE RECURSO"
                        else:
                            status_av = "Encerrada"
                    else:
                        status_av = "Encerrada"
            else:
                # Não houve discordância: prazo de 5 dias úteis a partir da data de AV2
                dt_base = parse_date_safe(row[45]) # Data Avaliação 2
                if dt_base is not None:
                    deadline = add_business_days(dt_base, 5)
                    if ref_date <= deadline:
                        status_av = "EM PRAZO DE RECURSO"
                    else:
                        status_av = "Encerrada"
                else:
                    status_av = "Encerrada"
        else:
            # Houve recurso
            if n_f4 is not None:
                status_av = "Encerrada"
            elif r_f3 not in ("", "-") or n_f3 is not None:
                status_av = "Encerrada"
            elif r_f2 not in ("", "-"):
                if n_f2 is not None:
                    status_av = "Encerrada"
                else:
                    status_av = "AUTORIDADE RECURSAL"
            else:
                status_av = "RECONSIDERAÇÃO COMISSÃO"

        rows.append({
            "nrPM (Avaliado)":              nrpm,
            "Nome Completo (Avaliado)":     row[2].strip(),
            "Posto/Graduação (Avaliado)":   row[6].strip(),
            "Unidade RPM (Avaliado)":       row[7].strip(),
            "Unidade Principal (Avaliado)": row[8].strip(),
            "Local/Unidade (Avaliado)":     local,
            "Quadro Atual (Avaliado)":      row[10].strip(),
            "Situação Funcional Atual":     sit,
            # Avaliação 1
            "Data Avaliação 1":             row[36].strip(),
            "Conceito Geral":               j,
            # Avaliação 2
            "Data Avaliação 2":             row[45].strip(),
            "Nota Geral":                   l,
            "Certificação Homologador":     calc_cert_hom(j, l),
            # Homologação
            "Data Homologação":             row[71].strip(),
            "Nota Homologação":             n,
            # Competências
            "Competência 1":                row[48].strip(),
            "Conceito (Comp. 1)":           row[49].strip(),
            "Nota (Comp. 1)":               row[50].strip(),
            "Competência 2":                row[51].strip(),
            "Conceito (Comp. 2)":           row[52].strip(),
            "Nota (Comp. 2)":               row[53].strip(),
            "Competência 3":                row[54].strip(),
            "Conceito (Comp. 3)":           row[55].strip(),
            "Nota (Comp. 3)":               row[56].strip(),
            "Competência 4":                row[57].strip(),
            "Conceito (Comp. 4)":           row[58].strip(),
            "Nota (Comp. 4)":               row[59].strip(),
            # Avaliador 1
            "nrPM (Avaliador 1)":               row[28].strip(),
            "Nome (Avaliador 1)":               row[29].strip(),
            "Posto (Avaliador 1)":              row[30].strip(),
            "RPM (Avaliador 1)":                row[31].strip(),
            "Unid. Principal (Avaliador 1)":    row[32].strip(),
            "Local (Avaliador 1)":              row[33].strip(),
            "Quadro (Avaliador 1)":             row[34].strip(),
            "Situação (Avaliador 1)":           row[35].strip(),
            # Avaliador 2
            "nrPM (Avaliador 2)":               row[37].strip(),
            "Nome (Avaliador 2)":               row[38].strip(),
            "Posto (Avaliador 2)":              row[39].strip(),
            "RPM (Avaliador 2)":                row[40].strip(),
            "Unid. Principal (Avaliador 2)":    row[41].strip(),
            "Local (Avaliador 2)":              row[42].strip(),
            "Quadro (Avaliador 2)":             row[43].strip(),
            "Situação (Avaliador 2)":           row[44].strip(),
            # Homologador
            "nrPM (Homologador)":               row[62].strip(),
            "Nome (Homologador)":               row[63].strip(),
            "Posto (Homologador)":              row[64].strip(),
            "RPM (Homologador)":                row[65].strip(),
            "Unid. Principal (Homologador)":    row[66].strip(),
            "Local (Homologador)":              row[67].strip(),
            "Quadro (Homologador)":             row[68].strip(),
            "Situação (Homologador)":           row[69].strip(),
            # Análise
            "Situação Comissão":            sc,
            "Status Avaliação":             status_av,
            # Recursos (Fase 1 a 4)
            "Recurso Fase 1":               row[73].strip(),
            "Nota (Fase 1)":                row[74].strip(),
            "Data Recurso 1":               row[76].strip(),
            "Recurso Fase 2":               row[77].strip(),
            "Nota (Fase 2)":                row[78].strip(),
            "Data Recurso 2":               row[80].strip(),
            "Recurso Fase 3":               row[81].strip(),
            "Nota (Fase 3)":                row[82].strip(),
            "Data Recurso 3":               row[84].strip(),
            "Recurso Fase 4":               row[85].strip(),
            "Nota (Fase 4)":                row[86].strip(),
            "Data Recurso 4":               row[88].strip(),
        })

df_full = pd.DataFrame(rows)
print(f"  {len(df_full):,} registros processados do geral.csv")

# ─── 6. ENRIQUECIMENTO COM MOVIMENTAÇÕES, PRESO, PUNIÇÃO ─────────────────────
print("Enriquecendo dados com MOVIMENTAÇÕES, PRESO e PUNIÇÃO ...")

# Colunas extras
data_transf_list   = []
motivo_transf_list = []
data_inicio_preso_list  = []
data_fim_preso_list     = []
sit_func_preso_list     = []
dias_preso_list    = []
qtd_pun_list       = []
qtd_ativadas_list  = []
pontos_pun_list    = []

for _, row in df_full.iterrows():
    nrpm_key = str(row["nrPM (Avaliado)"]).strip().lstrip("0") or "0"

    # ── MOVIMENTAÇÕES ──
    mov = mov_mais_recente.get(nrpm_key)
    if mov:
        data_transf_list.append(fmt_date(mov["data"]) if mov["data"] else "")
        motivo_transf_list.append(mov["motivo"])
    else:
        data_transf_list.append("")
        motivo_transf_list.append("")

    # ── PRESO ──
    preso = preso_info.get(nrpm_key)
    if preso:
        dt_inicio = preso["data_inicio"]
        dt_fim    = preso["data_fim"]
        sit_f     = preso["sit_func"]

        data_inicio_preso_list.append(fmt_date(dt_inicio) if dt_inicio else "")
        data_fim_preso_list.append(fmt_date(dt_fim) if dt_fim else "")
        sit_func_preso_list.append(sit_f)

        # Dias preso
        if dt_inicio:
            ref = dt_fim if dt_fim else TODAY
            dias = (ref - dt_inicio).days
            dias_preso_list.append(max(0, dias))
        else:
            dias_preso_list.append("")
    else:
        data_inicio_preso_list.append("")
        data_fim_preso_list.append("")
        sit_func_preso_list.append("")
        dias_preso_list.append("")

    # ── PUNIÇÃO ──
    pun = pun_map.get(nrpm_key)
    if pun and pun["qtd_punicoes"] > 0:
        qtd_pun_list.append(pun["qtd_punicoes"])
        qtd_ativadas_list.append(pun["qtd_ativadas"])
        pontos_pun_list.append(pun["pontos"])
    else:
        qtd_pun_list.append("")
        qtd_ativadas_list.append("")
        pontos_pun_list.append("")

df_full["DATA DA TRANSFERÊNCIA"]   = data_transf_list
df_full["MOTIVO DA TRANSFERÊNCIA"] = motivo_transf_list
df_full["DATA INÍCIO DA PRISÃO"]   = data_inicio_preso_list
df_full["DATA FIM DA PRISÃO"]      = data_fim_preso_list
df_full["SITUAÇÃO FUNCIONAL (PRESO)"] = sit_func_preso_list
df_full["DIAS PRESO"]              = dias_preso_list
df_full["QTDADE DE PUNIÇÕES"]      = qtd_pun_list
df_full["QTDADE DE ATIVADAS"]      = qtd_ativadas_list
df_full["PONTOS PERDIDOS"]         = pontos_pun_list

print(f"  Militares com movimentação : {sum(1 for x in data_transf_list if x)}")
print(f"  Militares com prisão       : {sum(1 for x in data_inicio_preso_list if x)}")
print(f"  Militares com punição      : {sum(1 for x in qtd_pun_list if x != '')}")

# ─── 7. COLUNAS PARA EXPORTAÇÃO ───────────────────────────────────────────────
COLS_EXPORT = [
    # Identificação do avaliado
    "nrPM (Avaliado)",
    "Nome Completo (Avaliado)",
    "Posto/Graduação (Avaliado)",
    "Unidade RPM (Avaliado)",
    "Unidade Principal (Avaliado)",
    "Local/Unidade (Avaliado)",
    "Quadro Atual (Avaliado)",
    "Situação Funcional Atual",
    # Avaliação
    "Data Avaliação 1",
    "Conceito Geral",
    "Data Avaliação 2",
    "Nota Geral",
    "Certificação Homologador",
    "Data Homologação",
    "Nota Homologação",
    # Competências
    "Competência 1", "Conceito (Comp. 1)", "Nota (Comp. 1)",
    "Competência 2", "Conceito (Comp. 2)", "Nota (Comp. 2)",
    "Competência 3", "Conceito (Comp. 3)", "Nota (Comp. 3)",
    "Competência 4", "Conceito (Comp. 4)", "Nota (Comp. 4)",
    # Avaliador 1
    "nrPM (Avaliador 1)", "Nome (Avaliador 1)", "Posto (Avaliador 1)",
    "RPM (Avaliador 1)", "Unid. Principal (Avaliador 1)",
    "Local (Avaliador 1)", "Quadro (Avaliador 1)", "Situação (Avaliador 1)",
    # Avaliador 2
    "nrPM (Avaliador 2)", "Nome (Avaliador 2)", "Posto (Avaliador 2)",
    "RPM (Avaliador 2)", "Unid. Principal (Avaliador 2)",
    "Local (Avaliador 2)", "Quadro (Avaliador 2)", "Situação (Avaliador 2)",
    # Homologador
    "nrPM (Homologador)", "Nome (Homologador)", "Posto (Homologador)",
    "RPM (Homologador)", "Unid. Principal (Homologador)",
    "Local (Homologador)", "Quadro (Homologador)", "Situação (Homologador)",
    # Análise
    "Situação Comissão",
    "Status Avaliação",
    # Recursos
    "Recurso Fase 1", "Nota (Fase 1)", "Data Recurso 1",
    "Recurso Fase 2", "Nota (Fase 2)", "Data Recurso 2",
    "Recurso Fase 3", "Nota (Fase 3)", "Data Recurso 3",
    "Recurso Fase 4", "Nota (Fase 4)", "Data Recurso 4",
    # ── NOVAS COLUNAS ──
    "DATA DA TRANSFERÊNCIA",
    "MOTIVO DA TRANSFERÊNCIA",
    "DATA INÍCIO DA PRISÃO",
    "DATA FIM DA PRISÃO",
    "SITUAÇÃO FUNCIONAL (PRESO)",
    "DIAS PRESO",
    "QTDADE DE PUNIÇÕES",
    "QTDADE DE ATIVADAS",
    "PONTOS PERDIDOS",
]

# Garante que todas as colunas existam
for col in COLS_EXPORT:
    if col not in df_full.columns:
        df_full[col] = ""

# ─── 8. ESTILOS OPENPYXL ──────────────────────────────────────────────────────
STATUS_BG = {
    "Encerrada":              "70AD47",
    "Homologação":            "FFD966",
    "Parcialmente Encerrada": "FF8C00",
    "Aberta":                 "FF4444",
    "EM PRAZO DE RECURSO":    "BDC3C7",
    "EM RECURSO (FASE 1)":    "F5B041",
    "EM RECURSO (FASE 2)":    "E67E22",
    "EM RECURSO (FASE 3)":    "D35400",
}
SIT_BG = {
    "Comissão Atual":  "4472C4",
    "Nota Provisória": "FFC000",
}

NOVAS_COLUNAS = {
    "DATA DA TRANSFERÊNCIA",
    "MOTIVO DA TRANSFERÊNCIA",
    "DATA INÍCIO DA PRISÃO",
    "DATA FIM DA PRISÃO",
    "SITUAÇÃO FUNCIONAL (PRESO)",
    "DIAS PRESO",
    "QTDADE DE PUNIÇÕES",
    "QTDADE DE ATIVADAS",
    "PONTOS PERDIDOS",
}

thin = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

S = {
    "hdr_fill":   PatternFill("solid", fgColor="1F3864"),
    "nova_fill":  PatternFill("solid", fgColor="7B3F00"),  # Marrom para novas colunas
    "hdr_font":   Font(bold=True, color="FFFFFF", name="Calibri", size=9),
    "hdr_al":     Alignment(horizontal="center", vertical="center", wrap_text=True),
    "data_font":  Font(name="Calibri", size=9),
    "center":     Alignment(horizontal="center", vertical="center"),
    "left":       Alignment(vertical="center"),
    "title_font": Font(bold=True, color="FFFFFF", name="Calibri", size=13),
    "title_al":   Alignment(horizontal="center", vertical="center"),
}

def write_sheet(ws, df, titulo, cols, show_all=True):
    """Escreve os dados em uma aba do workbook."""
    df_c = df[[c for c in cols if c in df.columns]].reset_index(drop=True)
    actual_cols = list(df_c.columns)
    n_cols = len(actual_cols)

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(n_cols, 1))
    c = ws.cell(1, 1, titulo)
    c.fill = PatternFill("solid", fgColor="1F3864")
    c.font = S["title_font"]
    c.alignment = S["title_al"]
    ws.row_dimensions[1].height = 22

    # Cabeçalhos
    for ci, col in enumerate(actual_cols, 1):
        cell = ws.cell(2, ci, col)
        fill = S["nova_fill"] if col in NOVAS_COLUNAS else S["hdr_fill"]
        cell.fill = fill
        cell.font = S["hdr_font"]
        cell.alignment = S["hdr_al"]
        cell.border = BORDER
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "A3"

    # Dados
    for r_idx, (_, row) in enumerate(df_c.iterrows(), 3):
        for ci, col in enumerate(actual_cols, 1):
            val = row.get(col, "")
            txt = "" if (pd.isna(val) if isinstance(val, float) else False) else str(val) if val != "" else ""
            cell = ws.cell(r_idx, ci, txt)
            cell.font = S["data_font"]
            cell.border = BORDER
            cell.alignment = S["center"] if ci > 8 else S["left"]

            if col == "Status Avaliação" and val in STATUS_BG:
                cell.fill = PatternFill("solid", fgColor=STATUS_BG[val])
                cell.font = Font(bold=True, name="Calibri", size=9,
                                 color="FFFFFF" if val == "Aberta" else "000000")
            elif col == "Situação Comissão" and val in SIT_BG:
                cell.fill = PatternFill("solid", fgColor=SIT_BG[val])
                cell.font = Font(bold=True, name="Calibri", size=9,
                                 color="FFFFFF" if val == "Comissão Atual" else "000000")
            elif col in NOVAS_COLUNAS and txt:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")  # Amarelo claro para novas colunas com dados
                cell.font = Font(bold=False, name="Calibri", size=9, color="000000")

    # Largura das colunas
    for ci, col in enumerate(actual_cols, 1):
        max_len = len(str(col))
        if col in df_c.columns and not df_c.empty:
            m = df_c[col].astype(str).str.len().max()
            if pd.notna(m):
                max_len = max(max_len, int(m))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len * 0.9, 8), 45)

    return len(df_c)

# ─── 9. GERAR WORKBOOK ────────────────────────────────────────────────────────
print(f"\nGerando {OUT_FILE} ...")
wb = Workbook()

# Aba 1 — Geral (todos os registros)
ws1 = wb.active
ws1.title = "Geral"
n = write_sheet(ws1, df_full,
                f"AUDITORIA DE NOTAS — AADP 2026 (Gerado em {TODAY.strftime('%d/%m/%Y')})",
                COLS_EXPORT)
print(f"  Aba 'Geral': {n} registros")

# Aba 2 — Pendentes
ws2 = wb.create_sheet("Pendentes")
df_pend = df_full[df_full["Status Avaliação"].isin(["Aberta", "Parcialmente Encerrada", "Homologação"])].copy()
n = write_sheet(ws2, df_pend,
                f"AVALIAÇÕES PENDENTES — AADP 2026",
                COLS_EXPORT)
print(f"  Aba 'Pendentes': {n} registros")

# Aba 3 — Com Transferência
ws3 = wb.create_sheet("Com Transferência")
df_transf = df_full[df_full["DATA DA TRANSFERÊNCIA"] != ""].copy()
n = write_sheet(ws3, df_transf,
                f"MILITARES COM TRANSFERÊNCIA",
                COLS_EXPORT)
print(f"  Aba 'Com Transferência': {n} registros")

# Aba 4 — Com Prisão
ws4 = wb.create_sheet("Com Prisão")
df_pres = df_full[df_full["DATA INÍCIO DA PRISÃO"] != ""].copy()
n = write_sheet(ws4, df_pres,
                f"MILITARES COM REGISTRO DE PRISÃO",
                COLS_EXPORT)
print(f"  Aba 'Com Prisão': {n} registros")

# Aba 5 — Com Punição
ws5 = wb.create_sheet("Com Punição")
df_punidos = df_full[df_full["QTDADE DE PUNIÇÕES"] != ""].copy()
n = write_sheet(ws5, df_punidos,
                f"MILITARES COM PUNIÇÕES",
                COLS_EXPORT)
print(f"  Aba 'Com Punição': {n} registros")

# Aba 6 — Resumo estatístico
ws6 = wb.create_sheet("Resumo")

def escreve_resumo(ws, df):
    ws.merge_cells("A1:F1")
    t = ws.cell(1, 1, f"RESUMO ESTATÍSTICO — AUDITORIA NOTAS AADP 2026 — {TODAY.strftime('%d/%m/%Y')}")
    t.fill = PatternFill("solid", fgColor="1F3864")
    t.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    STATUS_ORDEM = [
        "Aberta", "Parcialmente Encerrada", "Homologação", "Encerrada",
        "EM PRAZO DE RECURSO", "EM RECURSO (FASE 1)", "EM RECURSO (FASE 2)", "EM RECURSO (FASE 3)"
    ]
    STATUS_CORES = {
        "Aberta":                 "FF4444",
        "Parcialmente Encerrada": "FF8C00",
        "Homologação":            "FFD966",
        "Encerrada":              "70AD47",
        "EM PRAZO DE RECURSO":    "BDC3C7",
        "EM RECURSO (FASE 1)":    "F5B041",
        "EM RECURSO (FASE 2)":    "E67E22",
        "EM RECURSO (FASE 3)":    "D35400",
    }

    r = 3
    # Seção: Status das Avaliações
    ws.merge_cells(f"A{r}:F{r}")
    h = ws.cell(r, 1, "STATUS DAS AVALIAÇÕES")
    h.fill = PatternFill("solid", fgColor="2E5090")
    h.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 20
    r += 1

    hcols = ["Status", "Total", "%"]
    for ci, hc in enumerate(hcols, 1):
        c = ws.cell(r, ci, hc)
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[r].height = 20
    r += 1

    total = len(df)
    for st in STATUS_ORDEM:
        cnt = int((df["Status Avaliação"] == st).sum())
        pct = f"{cnt/total*100:.2f}%" if total > 0 else "0%"
        cor = STATUS_CORES.get(st, "FFFFFF")
        for ci, val in enumerate([st, cnt, pct], 1):
            c = ws.cell(r, ci, val)
            c.fill = PatternFill("solid", fgColor=cor)
            c.font = Font(bold=True, name="Calibri", size=10,
                          color="FFFFFF" if st == "Aberta" else "000000")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
        ws.row_dimensions[r].height = 18
        r += 1

    for ci, val in enumerate(["TOTAL GERAL", total, "100%"], 1):
        c = ws.cell(r, ci, val)
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[r].height = 20
    r += 2

    # Seção: Situação Comissão
    ws.merge_cells(f"A{r}:F{r}")
    h2 = ws.cell(r, 1, "SITUAÇÃO DA COMISSÃO")
    h2.fill = PatternFill("solid", fgColor="2E5090")
    h2.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    h2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 20
    r += 1

    for ci, hc in enumerate(["Situação", "Total", "%"], 1):
        c = ws.cell(r, ci, hc)
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[r].height = 20
    r += 1

    for sit_c, cor in [("Comissão Atual", "4472C4"), ("Nota Provisória", "FFC000")]:
        cnt = int((df["Situação Comissão"] == sit_c).sum())
        pct = f"{cnt/total*100:.2f}%" if total > 0 else "0%"
        for ci, val in enumerate([sit_c, cnt, pct], 1):
            c = ws.cell(r, ci, val)
            c.fill = PatternFill("solid", fgColor=cor)
            c.font = Font(bold=True, name="Calibri", size=10,
                          color="FFFFFF" if sit_c == "Comissão Atual" else "000000")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1

    # Seção: Extras
    ws.merge_cells(f"A{r}:F{r}")
    h3 = ws.cell(r, 1, "DADOS COMPLEMENTARES")
    h3.fill = PatternFill("solid", fgColor="7B3F00")
    h3.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    h3.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 20
    r += 1

    extras = [
        ("Total de avaliados", total),
        ("Com transferência", int((df["DATA DA TRANSFERÊNCIA"] != "").sum())),
        ("Com prisão", int((df["DATA INÍCIO DA PRISÃO"] != "").sum())),
        ("Com punição", int((df["QTDADE DE PUNIÇÕES"] != "").sum())),
        ("Total de punições (soma)", int(pd.to_numeric(df["QTDADE DE PUNIÇÕES"], errors="coerce").fillna(0).sum())),
        ("Total pontos perdidos (soma)", int(pd.to_numeric(df["PONTOS PERDIDOS"], errors="coerce").fillna(0).sum())),
    ]
    for label, val in extras:
        c1 = ws.cell(r, 1, label)
        c2 = ws.cell(r, 2, val)
        for c in [c1, c2]:
            c.font = Font(name="Calibri", size=10)
            c.border = BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 18
        r += 1

    for ci in range(1, 7):
        ws.column_dimensions[get_column_letter(ci)].width = 35 if ci == 1 else 15

escreve_resumo(ws6, df_full)
print(f"  Aba 'Resumo': OK")

# Salva
wb.save(OUT_FILE)
print(f"\n✅ Arquivo gerado com sucesso:")
print(f"   {OUT_FILE}")
print(f"\n📊 Estatísticas:")
print(f"   Total registros : {len(df_full):,}")
sc = df_full["Status Avaliação"].value_counts().to_dict()
print(f"   Aberta          : {sc.get('Aberta', 0):,}")
print(f"   Parc. Encerrada : {sc.get('Parcialmente Encerrada', 0):,}")
print(f"   Homologação     : {sc.get('Homologação', 0):,}")
print(f"   Encerrada       : {sc.get('Encerrada', 0):,}")
print(f"   Em Prazo Recurso: {sc.get('EM PRAZO DE RECURSO', 0):,}")
print(f"   Em Recurso F1   : {sc.get('EM RECURSO (FASE 1)', 0):,}")
print(f"   Em Recurso F2   : {sc.get('EM RECURSO (FASE 2)', 0):,}")
print(f"   Em Recurso F3   : {sc.get('EM RECURSO (FASE 3)', 0):,}")
cc = df_full["Situação Comissão"].value_counts().to_dict()
print(f"   Comissão Atual  : {cc.get('Comissão Atual', 0):,}")
print(f"   Nota Provisória : {cc.get('Nota Provisória', 0):,}")
print(f"   Com transferência: {int((df_full['DATA DA TRANSFERÊNCIA'] != '').sum()):,}")
print(f"   Com prisão       : {int((df_full['DATA INÍCIO DA PRISÃO'] != '').sum()):,}")
print(f"   Com punição      : {int((df_full['QTDADE DE PUNIÇÕES'] != '').sum()):,}")
